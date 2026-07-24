import streamlit as st
import pandas as pd
import collections
import io
import json
import math
import re
import uuid
import os
from datetime import date, datetime, timezone, timedelta

# Importações de Terceiros
from docxtpl import DocxTemplate
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import google.generativeai as genai
from PIL import Image

# --- CONFIGURAÇÃO DA TELA (DEVE SER O PRIMEIRO COMANDO STREAMLIT) ---
st.set_page_config(page_title="App SIARCON - Propostas e Custos", layout="wide", page_icon="📄")

# =========================================================================
# 🧠 MOTOR PARAMÉTRICO DE ENGENHARIA (SINCRONIZADO COM O SEU EXCEL)
# =========================================================================
def gerar_composicao_cavalete(equipamento, vias, bitola):
    # A MÁGICA DO TIPO DE LIGAÇÃO AUTOMÁTICA
    bitolas_roscadas = ["1/4\"", "3/8\"", "1/2\"", "3/4\"", "1\"", "1.1/4\"", "1.1/2\"", "2\""]
    is_roscado = bitola in bitolas_roscadas
    
    composicao = []
    
    if equipamento == "UTA/FANCOIL":
        # Periféricos Fixos (String exata do banco)
        composicao.extend([
            {"nome": "Poço para termômetro Latão haste = 38mm - Ø 1/2\"", "qtd": 2},
            {"nome": "Termômetro tipo capela. Escala 0° a 40°C - Ø 3/8\"", "qtd": 2},
            {"nome": "Robinete com alivio para manômetro Latão - Ø 1/2\"", "qtd": 1},
            {"nome": "Manometro tipo dial circular glicerinado, escala 0 a 4kgf/cm² - Ø 1/2\"", "qtd": 1},
            {"nome": "Luva Ferro maleável galvanizado - Ø 1/2\"", "qtd": 2},
            {"nome": "Niple duplo Ferro maleável galvanizado - Ø 1/4\"", "qtd": 2},
            {"nome": "Tubo em aço carbono Aço carbono preto - Ø 1/2\"", "qtd": 1},
        ])
        
        # Isolamento padrão
        composicao.append({"nome": f"Isolamento Espuma elastomérica, espessura 25 mm - Ø {bitola}", "qtd": 2.0})
        composicao.append({"nome": f"Rechapeamento chapa de alumínio liso, espessura 0,5 mm - Ø {bitola}", "qtd": 2.0})

        if is_roscado:
            qtd_unioes = 5 if vias == "2 Vias" else 6
            qtd_gaveta = 2 if vias == "2 Vias" else 3
            qtd_curva = 4 if vias == "2 Vias" else 6
            qtd_niple = 4 if vias == "2 Vias" else 6
            qtd_te = 1 if vias == "2 Vias" else 2
            
            composicao.extend([
                {"nome": f"União com assento cônico Ferro maleável galvanizado - Ø {bitola}", "qtd": qtd_unioes},
                {"nome": f"Curva 90° Ferro maleável galvanizado - Ø {bitola}", "qtd": qtd_curva},
                {"nome": f"Tubo em aço carbono Aço carbono preto - Ø {bitola}", "qtd": 3.0},
                {"nome": f"Luva de redução - solda/rosca Aço carbono preto - Ø {bitola} x 1/2\"", "qtd": 2},
                {"nome": f"Válvula de bloqueio tipo gaveta Corpo de bronze, catelo roscado internos em bronze, haste fixa. - Ø {bitola}", "qtd": qtd_gaveta},
                {"nome": f"Válvula balanceadora - Ø {bitola}", "qtd": 1},
                {"nome": f"Filtro \"Y\" Corpo, tampa e tampão de bronze filtro de aço inoxidável - Ø {bitola}", "qtd": 1},
                {"nome": f"Conexão T 90° Ferro maleável galvanizado - Ø {bitola}", "qtd": qtd_te},
                {"nome": f"Niple duplo Ferro maleável galvanizado - Ø {bitola}", "qtd": qtd_niple},
            ])
        else: # Flangeado
            qtd_flanges = 12 if vias == "2 Vias" else 14
            qtd_borboleta = 2 if vias == "2 Vias" else 3
            qtd_curva = 4 if vias == "2 Vias" else 6
            qtd_te = 1 if vias == "2 Vias" else 2
            
            composicao.extend([
                {"nome": f"Flange sobreposto tipo slip-on aço forjado - Ø {bitola}", "qtd": qtd_flanges},
                {"nome": f"Curva 90° Aço carbono preto SCH40 - Ø {bitola}", "qtd": qtd_curva},
                {"nome": f"Tubo em aço carbono Aço carbono preto - Ø {bitola}", "qtd": 3.0},
                {"nome": f"Válvula borboleta flangeada - Ø {bitola}", "qtd": qtd_borboleta},
                {"nome": f"Válvula balanceadora - Ø {bitola}", "qtd": 1},
                {"nome": f"Filtro Y Flangeado Ferro Fundido - Ø {bitola}", "qtd": 1},
                {"nome": f"Conexão T 90° Aço carbono preto SCH40 - Ø {bitola}", "qtd": qtd_te},
            ])
            
        v_nome = "Válvula 2 vias, motorizada com atuador proporcional" if vias == "2 Vias" else "Válvula 3 vias, motorizada com atuador proporcional"
        composicao.append({"nome": f"{v_nome} - Ø {bitola}", "qtd": 1})

    elif equipamento == "CHILLER":
        composicao.extend([
            {"nome": f"Poço para termômetro Latão haste = 38mm - Ø 1/2\"", "qtd": 2},
            {"nome": f"Termômetro tipo capela. Escala 0° a 40°C - Ø 3/8\"", "qtd": 2},
            {"nome": f"Robinete com alivio para manômetro Latão - Ø 1/2\"", "qtd": 1},
            {"nome": f"Manometro tipo dial circular glicerinado, escala 0 a 4kgf/cm² - Ø 1/2\"", "qtd": 1},
            {"nome": f"Tubo em aço carbono Aço carbono preto - Ø {bitola}", "qtd": 6.0},
            {"nome": f"Filtro Y Flangeado Ferro Fundido - Ø {bitola}", "qtd": 1},
            {"nome": f"Válvula borboleta flangeada - Ø {bitola}", "qtd": 3},
            {"nome": f"Junta de expansão de borracha flangeada - Ø {bitola}", "qtd": 2},
            {"nome": f"Flange sobreposto tipo slip-on aço forjado - Ø {bitola}", "qtd": 14},
            {"nome": f"Redução concêntrica de aço carbono preto - Ø {bitola} x 3\"", "qtd": 2},
            {"nome": f"Curva 90° Aço carbono preto SCH40 - Ø {bitola}", "qtd": 4},
        ])

    elif equipamento == "BOMBA":
        composicao.extend([
            {"nome": f"Robinete com alivio para manômetro Latão - Ø 1/2\"", "qtd": 1},
            {"nome": f"Manometro tipo dial circular glicerinado, escala 0 a 4kgf/cm² - Ø 1/2\"", "qtd": 1},
            {"nome": f"Tubo em aço carbono Aço carbono preto - Ø {bitola}", "qtd": 6.0},
            {"nome": f"Válvula de retenção tipo portinhola flangeada - Ø {bitola}", "qtd": 1},
            {"nome": f"Válvula borboleta flangeada - Ø {bitola}", "qtd": 2},
            {"nome": f"Filtro Y Flangeado Ferro Fundido - Ø {bitola}", "qtd": 1},
            {"nome": f"Junta de expansão de borracha flangeada - Ø {bitola}", "qtd": 2},
            {"nome": f"Flange sobreposto tipo slip-on aço forjado - Ø {bitola}", "qtd": 14},
            {"nome": f"Redução concêntrica de aço carbono preto - Ø {bitola}", "qtd": 1},
            {"nome": f"Redução excêntrica de aço carbono preto - Ø {bitola}", "qtd": 1},
            {"nome": f"Curva 90° Aço carbono preto SCH40 - Ø {bitola}", "qtd": 1},
        ])

    return composicao

# ==========================================
# CONFIGURAÇÕES GERAIS E CONEXÕES
# ==========================================
def buscar_logo():
    nomes_possiveis = ["SIARCON.png", "SIARCON .png", "siarcon.png", "Siarcon.png", "logo.png"]
    for nome in nomes_possiveis:
        if os.path.exists(nome):
            return nome
    return None

ARQUIVO_LOGO = buscar_logo()

PLANILHA_URL = "https://docs.google.com/spreadsheets/d/1DgBxNqwUepO2RW6GdRwnFHxg7dLlWiRGZjdglkQ8Ls0/edit?gid=1169331401#gid=1169331401"

@st.cache_resource
def conectar_google_sheets():
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds_dict = st.secrets["gcp_service_account"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        return client.open_by_url(PLANILHA_URL)
    except Exception as e:
        st.error(f"Erro na conexão com Google Sheets: {e}. Verifique o link.")
        st.stop()

try:
    if "GEMINI_API_KEY" in st.secrets:
        genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
        model_ia = genai.GenerativeModel('gemini-1.5-flash')
        ia_disponivel = True
    else:
        ia_disponivel = False
except Exception as e:
    ia_disponivel = False

fuso_br = timezone(timedelta(hours=-3))

# ==========================================
# 🔐 INICIALIZAÇÃO SEGURA DE VARIÁVEIS GLOBAIS
# ==========================================
if "usuario_logado" not in st.session_state:
    st.session_state.usuario_logado = None
if "nome_exibicao" not in st.session_state:
    st.session_state.nome_exibicao = ""
if "menu_selecionado" not in st.session_state:
    st.session_state.menu_selecionado = "🏠 Tela Inicial"
if "orcamento" not in st.session_state:
    st.session_state.orcamento = []
if "historico_precos" not in st.session_state:
    st.session_state.historico_precos = []
if 'nome_projeto_orcamento' not in st.session_state:
    st.session_state.nome_projeto_orcamento = ""
if 'projeto_para_abrir' not in st.session_state:
    st.session_state.projeto_para_abrir = None
if 'dados_projeto_abrir' not in st.session_state:
    st.session_state.dados_projeto_abrir = {}
if 'wizard_ativo' not in st.session_state:
    st.session_state.wizard_ativo = False
if 'paineis_auto' not in st.session_state:
    st.session_state.paineis_auto = []
if 'confirmar_limpar' not in st.session_state:
    st.session_state.confirmar_limpar = False
if 'data_precos_atualizada' not in st.session_state:
    st.session_state.data_precos_atualizada = "Buscando metadados da nuvem..."
if 'resultado_ia' not in st.session_state:
    st.session_state.resultado_ia = {}
if 'cavaletes_selecionados' not in st.session_state: 
    st.session_state.cavaletes_selecionados = []

# DICIONÁRIO DE NOMES DO DIAGRAMA 
if 'de_para_diagrama' not in st.session_state:
    st.session_state.de_para_diagrama = {
        "Transmissor de pressão dif. para ar (medição de vazão de ar) (PDT)": {
            "in_agua": "Trans. Pressão - Vazão (PDT)", "in_comp": "Trans. Pressão - Vazão (PDT)", "out_agua": "Modula Inversor", "out_comp": "Modula Inversor"
        },
        "Transmissor de temperatura e umidade para duto (TT/MT)": {
            "in_agua": "Trans. Temp. e Umid. (TT/MT)", "in_comp": "Trans. Temp. e Umid. (TT/MT)", "out_agua": "", "out_comp": ""
        },
        "Transmissor de temperatura para duto (TT)": {
            "in_agua": "Trans. Temp. (TT)", "in_comp": "Trans. Temp. (TT)", "out_agua": "", "out_comp": ""
        },
        "Válvula de controle ON/OFF (TCV)": {
            "in_agua": "", "in_comp": "", "out_agua": "Habilita Válvula", "out_comp": "Habilita Válvula"
        },
        "Válvula de controle de água gelada proporcional (TCV)": {
            "in_agua": "", "in_comp": "", "out_agua": "Modula VAG", "out_comp": "Habilita Compressor"
        },
        "Válvula de controle de água quente proporcional (TCV)": {
            "in_agua": "", "in_comp": "", "out_agua": "Modula VAQ", "out_comp": ""
        },
        "Válvula de controle de vapor proporcional (TCV)": {
            "in_agua": "", "in_comp": "", "out_agua": "Modula Vapor", "out_comp": ""
        },
        "Relé de Corrente - Status Compressor (TC)": {
            "in_agua": "", "in_comp": "Status Compressor", "out_agua": "", "out_comp": "Habilita Compressor"
        },
        "Termostato de segurança (TSH)": {
            "in_agua": "Termostato Seg. RAQ (TSH)", "in_comp": "Termostato Seg. RAQ (TSH)", "out_agua": "Status RAQ", "out_comp": "Status RAQ"
        },
        "Pressostato diferencial para ar (PSH)": {
            "in_agua": "Pressostato Seg. RAQ (PSH)", "in_comp": "Pressostato Seg. RAQ (PSH)", "out_agua": "Status RAQ", "out_comp": "Status RAQ"
        },
        "Resistência de aquecimento (Equipamento) (RAQ)": {
            "in_agua": "", "in_comp": "", "out_agua": "Habilita RAQ", "out_comp": "Habilita RAQ"
        },
        "Resistência de aquecimento (Duto) (RAQ)": {
            "in_agua": "", "in_comp": "", "out_agua": "Habilita RAQ", "out_comp": "Habilita RAQ"
        },
        "Pressostato para monitorar os filtros G4 (PSH)": {
            "in_agua": "Pressostato G4 (PSH)", "in_comp": "Pressostato G4 (PSH)", "out_agua": "Alarme G4 Saturado", "out_comp": "Alarme G4 Saturado"
        },
        "Pressostato para monitorar os filtros M5 (PSH)": {
            "in_agua": "Pressostato M5 (PSH)", "in_comp": "Pressostato M5 (PSH)", "out_agua": "Alarme M5 Saturado", "out_comp": "Alarme M5 Saturado"
        },
        "Transmissor de pressão diferencial (monitorar os filtros M5) (PDT)": {
            "in_agua": "Pressão Dif. M5 (PDT)", "in_comp": "Pressão Dif. M5 (PDT)", "out_agua": "", "out_comp": ""
        },
        "Pressostato para monitorar os filtros F9 (PSH)": {
            "in_agua": "Pressostato F9 (PSH)", "in_comp": "Pressostato F9 (PSH)", "out_agua": "Alarme F9 Saturado", "out_comp": "Alarme F9 Saturado"
        },
        "Transmissor de pressão diferencial (monitorar os filtros F9) (PDT)": {
            "in_agua": "Pressão Dif. F9 (PDT)", "in_comp": "Pressão Dif. F9 (PDT)", "out_agua": "", "out_comp": ""
        },
        "Pressostato para monitorar os filtros H13/H14 (PSH)": {
            "in_agua": "Pressostato H13/14 (PSH)", "in_comp": "Pressostato H13/14 (PSH)", "out_agua": "Alarme H13/14 Saturado", "out_comp": "Alarme H13/14 Saturado"
        },
        "Transmissor de pressão diferencial (monitorar os filtros H13) (PDT)": {
            "in_agua": "Pressão Dif. H13 (PDT)", "in_comp": "Pressão Dif. H13 (PDT)", "out_agua": "", "out_comp": ""
        },
        "Status funcionamento ventilador ou exaustor (partida direta) (PSH)": {
            "in_agua": "Status Func. Partida Direta (PSH)", "in_comp": "Status Func. Partida Direta (PSH)", "out_agua": "", "out_comp": ""
        },
        "Transmissor de pressão diferencial (monitorar os filtros G4) (PDT)": {
            "in_agua": "Pressão Dif. G4 (PDT)", "in_comp": "Pressão Dif. G4 (PDT)", "out_agua": "", "out_comp": ""
        },
        "Transmissor de pressão diferencial entre salas (PDT)": {
            "in_agua": "Pressão Dif. Salas (PDT)", "in_comp": "Pressão Dif. Salas (PDT)", "out_agua": "", "out_comp": ""
        },
        "Transmissor de pressão diferencial entre salas com display (PDIT)": {
            "in_agua": "Pressão Dif. Salas (PDIT)", "in_comp": "Pressão Dif. Salas (PDIT)", "out_agua": "", "out_comp": ""
        },
        "Transmissor de temperatura Ambiente (TT)": {
            "in_agua": "Temp. Salas (TT)", "in_comp": "Temp. Salas (TT)", "out_agua": "", "out_comp": ""
        },
        "Transmissor de temperatura ambiente com display (TIT)": {
            "in_agua": "Temp. Salas (TIT)", "in_comp": "Temp. Salas (TIT)", "out_agua": "", "out_comp": ""
        },
        "Transmissor de temperatura e umidade ambiente (TT/MT)": {
            "in_agua": "Temp. / Umid. (TT/MT)", "in_comp": "Temp. / Umid. (TT/MT)", "out_agua": "", "out_comp": ""
        },
        "Chave Seletora Auto/Manual (Painel Elétrico)": {
            "in_agua": "Chave Auto / Manual", "in_comp": "Chave Auto / Manual", "out_agua": "Habilita Equipamento (TAG)", "out_comp": "Habilita Equipamento (TAG)"
        },
        "Chave de fluxo para água (FS/CF)": {
            "in_agua": "Status Fluxo de Água", "in_comp": "", "out_agua": "", "out_comp": ""
        }
    }

REGRA_IO = {
    "Transmissor de pressão dif. para ar (medição de vazão de ar) (PDT)": {"AI": 1, "AO": 1, "DI": 1, "DO": 1},
    "Transmissor de temperatura e umidade para duto (TT/MT)": {"AI": 1, "AO": 0, "DI": 0, "DO": 0},
    "Transmissor de temperatura para duto (TT)": {"AI": 1, "AO": 1, "DI": 0, "DO": 0},
    "Válvula de controle ON/OFF (TCV)": {"AI": 0, "AO": 0, "DI": 0, "DO": 1},
    "Válvula de controle de água gelada proporcional (TCV)": {"AI": 0, "AO": 1, "DI": 0, "DO": 0},
    "Válvula de controle de água quente proporcional (TCV)": {"AI": 0, "AO": 1, "DI": 0, "DO": 0},
    "Válvula de controle de vapor proporcional (TCV)": {"AI": 0, "AO": 1, "DI": 0, "DO": 0},
    "Relé de Corrente - Status Compressor (TC)": {"AI": 0, "AO": 0, "DI": 1, "DO": 2},
    "Termostato de segurança (TSH)": {"AI": 0, "AO": 0, "DI": 1, "DO": 1},
    "Pressostato diferencial para ar (PSH)": {"AI": 0, "AO": 0, "DI": 1, "DO": 1},
    "Resistência de aquecimento (Equipamento) (RAQ)": {"AI": 0, "AO": 1, "DI": 2, "DO": 1},
    "Resistência de aquecimento (Duto) (RAQ)": {"AI": 0, "AO": 1, "DI": 2, "DO": 1},
    "Válvula motorizada Bypass Proporcional (hasta 2.1/2\") (TCV)": {"AI": 0, "AO": 1, "DI": 0, "DO": 0},
    "Válvula motorizada Bypass Proporcional (3\" ou 4\") (TCV)": {"AI": 0, "AO": 1, "DI": 0, "DO": 0},
    "Válvula motorizada Bypass Proporcional (5\") (TCV)": {"AI": 0, "AO": 1, "DI": 0, "DO": 0},
    "Válvula motorizada Bypass Proporcional (6\") (TCV)": {"AI": 0, "AO": 1, "DI": 0, "DO": 0},
    "Válvula motorizada Bypass Proporcional (8\") (TCV)": {"AI": 0, "AO": 1, "DI": 0, "DO": 0},
    "Transmissor de pressão para água (PIT)": {"AI": 1, "AO": 2, "DI": 0, "DO": 0},
    "Transmissor de vazão para água (FIT)": {"AI": 1, "AO": 1, "DI": 0, "DO": 0},
    "Válvula bloqueio motorizada (XV)": {"AI": 0, "AO": 0, "DI": 0, "DO": 1},
    "Chave de fluxo para água (FS/CF)": {"AI": 0, "AO": 0, "DI": 1, "DO": 0},
    "Bombas (I/O para controlador)": {"AI": 0, "AO": 1, "DI": 1, "DO": 1},
    "Tanques (I/O para controlador)": {"AI": 1, "AO": 0, "DI": 1, "DO": 1},
    "Pressostato para monitorar os filtros G4 (PSH)": {"AI": 0, "AO": 0, "DI": 1, "DO": 0},
    "Pressostato para monitorar os filtros M5 (PSH)": {"AI": 0, "AO": 0, "DI": 1, "DO": 0},
    "Pressostato para monitorar os filtros F9 (PSH)": {"AI": 0, "AO": 0, "DI": 1, "DO": 0},
    "Pressostato para monitorar os filtros H13/H14 (PSH)": {"AI": 0, "AO": 0, "DI": 1, "DO": 0},
    "Status funcionamento ventilador ou exaustor (partida direta) (PSH)": {"AI": 0, "AO": 0, "DI": 1, "DO": 1},
    "Transmissor de pressão diferencial (monitorar os filtros G4) (PDT)": {"AI": 1, "AO": 0, "DI": 0, "DO": 0},
    "Transmissor de pressão diferencial (monitorar os filtros M5) (PDT)": {"AI": 1, "AO": 0, "DI": 0, "DO": 0},
    "Transmissor de pressão diferencial (monitorar os filtros F9) (PDT)": {"AI": 1, "AO": 0, "DI": 0, "DO": 0},
    "Transmissor de pressão diferencial (monitorar os filtros H13) (PDT)": {"AI": 1, "AO": 0, "DI": 0, "DO": 0},
    "Transmissor de pressão diferencial entre salas (PDT)": {"AI": 1, "AO": 0, "DI": 0, "DO": 0},
    "Transmissor de pressão diferencial entre salas com display (PDIT)": {"AI": 1, "AO": 0, "DI": 0, "DO": 0},
    "Transmissor de temperatura Ambiente (TT)": {"AI": 1, "AO": 0, "DI": 0, "DO": 0},
    "Transmissor de temperatura ambiente com display (TIT)": {"AI": 1, "AO": 0, "DI": 0, "DO": 0},
    "Transmissor de temperatura e umidade ambiente (TT/MT)": {"AI": 2, "AO": 0, "DI": 0, "DO": 0},
    "Transmissor de temperatura e umidade ambiente com display (TIT/MIT)": {"AI": 2, "AO": 0, "DI": 0, "DO": 0},
    "Transmissor de CO2 ambiente (AT/AIT)": {"AI": 1, "AO": 1, "DI": 0, "DO": 1},
    "Transmissor de temperatura de imersão (TT)": {"AI": 1, "AO": 0, "DI": 0, "DO": 0},
    "Transmissor de temperatura de imersão com display (TIT)": {"AI": 1, "AO": 0, "DI": 0, "DO": 0}
}

banco_padrao_precos = {
    "Transmissor de pressão dif. para ar (medição de vazão de ar) (PDT)": 1490.00,
    "Transmissor de temperatura e umidade para duto (TT/MT)": 2050.00,
    "Transmissor de temperatura para duto (TT)": 800.00,
    "Válvula de controle ON/OFF (TCV)": 650.00,
    "Válvula de controle de água gelada proporcional (TCV)": 0.00,
    "Válvula de controle de água quente proporcional (TCV)": 0.00,
    "Válvula de controle de vapor proporcional (TCV)": 0.00,
    "Relé de Corrente - Status Compressor (TC)": 150.00,
    "Termostato de segurança (TSH)": 250.00,
    "Pressostato diferencial para ar (PSH)": 349.00,
    "Resistência de aquecimento (Equipamento) (RAQ)": 0.00,
    "Resistência de aquecimento (Duto) (RAQ)": 0.00,
    "Válvula motorizada Bypass Proporcional (hasta 2.1/2\") (TCV)": 2690.00,
    "Transmissor de pressão para água (PIT)": 1359.00,
    "Transmissor de vazão para água (FIT)": 3550.00,
    "Chave de fluxo para água (FS/CF)": 450.00,
    "Pressostato para monitorar os filtros G4 (PSH)": 349.00,
    "Pressostato para monitorar os filtros M5 (PSH)": 349.00,
    "Pressostato para monitorar os filtros F9 (PSH)": 349.00,
    "Pressostato para monitorar os filtros H13/H14 (PSH)": 349.00,
    "Status funcionamento ventilador ou exaustor (partida direta) (PSH)": 349.00,
    "Transmissor de pressão diferencial (monitorar os filtros G4) (PDT)": 1490.00,
    "Transmissor de pressão diferencial (monitorar os filtros M5) (PDT)": 1490.00,
    "Transmissor de pressão diferencial (monitorar os filtros F9) (PDT)": 1490.00,
    "Transmissor de pressão diferencial (monitorar os filtros H13) (PDT)": 1490.00,
    "Transmissor de pressão diferencial entre salas (PDT)": 1490.00,
    "Transmissor de pressão diferencial entre salas com display (PDIT)": 2110.00,
    "Transmissor de temperatura Ambiente (TT)": 2050.00,
    "Transmissor de temperatura ambiente com display (TIT)": 2650.00,
    "Transmissor de temperatura e umidade ambiente (TT/MT)": 2050.00,
    "Custo AI/AO": 565.00,
    "Custo DI/DO": 120.00,
    "Licença Supervisório - SEM CFR-21 (Base)": 23000.00,
    "Licença Supervisório - SEM CFR-21 (Por Ponto I/O)": 100.00,
    "Licença Supervisório - COM CFR-21 (Base)": 23000.00,
    "Licença Supervisório - COM CFR-21 (Por Ponto I/O)": 285.00,
    "Licença Supervisório - Schneider EBO (Base)": 13000.00,
    "Licença Supervisório - Schneider EBO (Por Ponto I/O)": 110.00,
    "MP-C-15A": 4649.49,
    "MP-C-18A": 5185.54,
    "MP-C-24A": 7290.75,
    "MP-C-36A": 9459.08,
    "Schneider - Sensor de Temperatura NTC (Duto)": 120.00,
    "Schneider - Sensor de Temperatura NTC (Ambiente)": 85.00,
    "Schneider - Servidor de Automação (SpaceLogic AS-P/AS-B)": 9500.00,
    "Siemens - CPU 1214C DC/DC/DC": 2500.00,
    "Siemens - CPU 1215C DC/DC/DC": 3200.00,
    "Siemens - SM 1231 AI 8x13Bit": 1900.00,
    "Siemens - SM 1232 AQ 4x14Bit": 2100.00,
    "Siemens - SM 1221 DI 16x24VDC": 1200.00,
    "Siemens - SM 1222 DQ 16x24VDC": 1300.00,
    "Siemens - Fonte 24VDC 2.5A": 800.00,
    "Siemens - Cartão de Memória 4MB": 400.00,
    "Siemens - CPU 1511-1 PN": 5500.00,
    "Siemens - AI 8xU/I HS": 2800.00,
    "Siemens - AQ 4xU/I ST": 3100.00,
    "Siemens - DI 16x24VDC HF": 1800.00,
    "Siemens - DQ 16x24VDC/0.5A": 1900.00,
    "Siemens - Fonte PM 1507 24VDC 8A": 1500.00,
    "Siemens - Cartão de Memória 12MB": 900.00,
    "Siemens - Serviço Custo AI/AO": 750.00,
    "Siemens - Serviço Custo DI/DO": 180.00,
    "Mercato - Controlador MDX (Expansão Direta)": 1250.00,
    "Mercato - Controlador MFC": 1650.00,
    "Mercato - Controlador MFC Plus": 2450.00,
    "Mercato - Sensor de Temperatura NTC (Duto)": 120.00,
    "Mercato - Sensor de Temperatura NTC (Ambiente)": 85.00,
    "Mercato - Sensor de Temperatura NTC com Display (Ambiente)": 350.00,
    "Mercato - Serviço Parametrização por Ponto": 80.00,
    "Mercato - IHM Básica 4.3\"": 1700.00,
    "CFR21 Qualificável - Até 100 pts": 70.00,
    "CFR21 Qualificável - 101 a 250 pts": 50.00,
    "CFR21 Qualificável - Acima de 250 pts": 30.00,
    "CFR21 Qualificado - Até 30 pts": 400.00,
    "CFR21 Qualificado - 31 a 60 pts": 350.00,
    "CFR21 Qualificado - Acima de 250 pts": 200.00,
    "Serviço de Calibração (Por Ponto Analógico)": 180.00,
    "IHM Padrão 7\"": 3400.00,
    "IHM Premium 10\"": 8500.00,
    "Sem Interface (Cego)": 0.00
}

banco_schneider_comum = {k:v for k,v in banco_padrao_precos.items() if "Siemens" not in k and "Mercato" not in k and "CFR" not in k and "IHM" not in k}
banco_siemens = {k:v for k,v in banco_padrao_precos.items() if "Siemens" in k}
banco_mercato = {k:v for k,v in banco_padrao_precos.items() if "Mercato" in k}
banco_cfr_servicos = {k:v for k,v in banco_padrao_precos.items() if "CFR" in k or "Calibração" in k}

if st.session_state.data_precos_atualizada == "Buscando metadados da nuvem...":
    try:
        sh_init = conectar_google_sheets()
        linhas_h = sh_init.worksheet("Historico_Precos").get_all_values()
        if len(linhas_h) > 1:
            st.session_state.data_precos_atualizada = linhas_h[-1][0]
        else:
            st.session_state.data_precos_atualizada = "Nenhuma alteração registrada recentemente"
    except:
        st.session_state.data_precos_atualizada = "Não foi possível carregar a data de atualização"

if 'precos_banco' not in st.session_state:
    st.session_state.precos_banco = banco_padrao_precos.copy()
    try:
        sh = conectar_google_sheets()
        aba_p = sh.worksheet("Precos").get_all_values()
        if len(aba_p) > 1:
            precos_bd = {linha[0]: float(linha[1]) for linha in aba_p[1:] if len(linha) > 1}
            st.session_state.precos_banco.update(precos_bd)
    except: pass

for k_n, v_n in banco_padrao_precos.items():
    if k_n not in st.session_state.precos_banco:
        st.session_state.precos_banco[k_n] = v_n

GRUPOS_INSTRUMENTOS = {
    "🔹 Controle (HVAC e Máquinas)": [
        "Transmissor de pressão dif. para ar (medição de vazão de ar) (PDT)",
        "Transmissor de temperatura e umidade para duto (TT/MT)", 
        "Transmissor de temperatura para duto (TT)",
        "Válvula de controle ON/OFF (TCV)", 
        "Válvula de controle de água gelada proporcional (TCV)",
        "Válvula de controle de água quente proporcional (TCV)",
        "Válvula de controle de vapor proporcional (TCV)",
        "Relé de Corrente - Status Compressor (TC)", 
        "Termostato de segurança (TSH)",
        "Pressostato diferencial para ar (PSH)", 
        "Resistência de aquecimento (Equipamento) (RAQ)",
        "Resistência de aquecimento (Duto) (RAQ)"
    ],
    "🔸 Monitoramento (Filtros e Status)": [
        "Pressostato para monitorar os filtros G4 (PSH)", 
        "Pressostato para monitorar os filtros M5 (PSH)", 
        "Transmissor de pressão diferencial (monitorar os filtros M5) (PDT)",
        "Pressostato para monitorar os filtros F9 (PSH)", 
        "Transmissor de pressão diferencial (monitorar os filtros F9) (PDT)",
        "Pressostato para monitorar os filtros H13/H14 (PSH)",
        "Transmissor de pressão diferencial (monitorar os filtros H13) (PDT)",
        "Status funcionamento ventilador ou exaustor (partida direta) (PSH)", 
        "Transmissor de pressão diferencial (monitorar os filtros G4) (PDT)", 
        "Chave de fluxo para água (FS/CF)"
    ],
    "🟢 Monitoramento e Controle de Ambientes": [
        "Transmissor de pressão diferencial entre salas (PDT)",
        "Transmissor de pressão diferencial entre salas com display (PDIT)",
        "Transmissor de temperatura Ambiente (TT)",
        "Transmissor de temperatura ambiente com display (TIT)",
        "Transmissor de temperatura e umidade ambiente (TT/MT)",
        "Transmissor de temperatura e umidade ambiente com display (TIT/MIT)",
        "Transmissor de CO2 ambiente (AT/AIT)"
    ]
}

KITS_PADRAO = {
    "❄️ UTA Padrão - Água Gelada": {
        "Transmissor de pressão dif. para ar (medição de vazão de ar) (PDT)": 1,
        "Transmissor de temperatura e umidade para duto (TT/MT)": 1, 
        "Válvula de controle de água gelada proporcional (TCV)": 1,
        "Pressostato para monitorar os filtros G4 (PSH)": 1, 
        "Pressostato para monitorar os filtros F9 (PSH)": 1
    },
    "🌬️ UTA Padrão - Expansão Direta": {
        "Transmissor de pressão dif. para ar (medição de vazão de ar) (PDT)": 1,
        "Transmissor de temperatura e umidade para duto (TT/MT)": 1, 
        "Relé de Corrente - Status Compressor (TC)": 2,
        "Pressostato para monitorar os filtros G4 (PSH)": 1, 
        "Pressostato para monitorar os filtros F9 (PSH)": 1
    },
    "🔥 UTA Padrão - Água Gelada + Resistência": {
        "Transmissor de pressão dif. para ar (medição de vazão de ar) (PDT)": 1,
        "Transmissor de temperatura e umidade para duto (TT/MT)": 1, 
        "Válvula de controle de água gelada proporcional (TCV)": 1,
        "Pressostato para monitorar os filtros G4 (PSH)": 1, 
        "Pressostato para monitorar os filtros F9 (PSH)": 1,
        "Termostato de segurança (TSH)": 1, 
        "Pressostato diferencial para ar (PSH)": 1, 
        "Resistência de aquecimento (Equipamento) (RAQ)": 1
    },
    "🔥 UTA Expansão Direta (2 Compressores) + Resistência (Salas e Exaustão)": {
        "Transmissor de pressão dif. para ar (medição de vazão de ar) (PDT)": 1,
        "Transmissor de temperatura e umidade para duto (TT/MT)": 1, 
        "Relé de Corrente - Status Compressor (TC)": 2,
        "Pressostato para monitorar os filtros G4 (PSH)": 1, 
        "Pressostato para monitorar os filtros M5 (PSH)": 1, 
        "Pressostato para monitorar os filtros F9 (PSH)": 1,
        "Termostato de segurança (TSH)": 1, 
        "Resistência de aquecimento (Equipamento) (RAQ)": 1,
        "Transmissor de pressão diferencial entre salas (PDT)": 4, 
        "Transmissor de temperatura e umidade ambiente (TT/MT)": 4,
        "Status funcionamento ventilador ou exaustor (partida direta) (PSH)": 2,
        "Pressostato diferencial para ar (PSH)": 1
    },
    "💨 Adicional: Ventilador/Exaustor (Inversor)": { "Transmissor de pressão dif. para ar (medição de vazão de ar) (PDT)": 1 },
    "⚙️ Adicional: Ventilador/Exaustor (Partida Direta)": { "Status funcionamento ventilador ou exaustor (partida direta) (PSH)": 1 }
}

def obter_cabo(inst_nome, is_output=False):
    i_up = inst_nome.upper()
    if is_output:
        if "ON/OFF" in i_up or "ON / OFF" in i_up: 
            return "PP 4x1,00mm²"
        if "RAQ" in i_up or "RESISTÊNCIA" in i_up: 
            return "2x1,00mm²"
        if "VÁLVULA" in i_up or "TCV" in i_up or "INVERSOR" in i_up or "VAZÃO" in i_up: 
            return "3x0,75mm² + Shield"
        if "CHAVE" in i_up: 
            return "5x1,00mm²"
        if "EXAUSTOR" in i_up or "VENTILADOR" in i_up or "COMPRESSOR" in i_up: 
            return "2x1,00mm²"
        return "2x1,00mm²"
    else:
        if "CHAVE" in i_up: 
            return "5x1,00mm²"
        if "(TT/MT)" in i_up or "TIT/MIT" in i_up: 
            return "5x0,75mm² + Shield"
        if "(PDT)" in i_up or "(PDIT)" in i_up or "(TT)" in i_up or "(PIT)" in i_up or "(FIT)" in i_up or "(TIT)" in i_up or "(TCV)" in i_up or "VÁLVULA" in i_up or "INVERSOR" in i_up or "VAZÃO" in i_up: 
            return "3x0,75mm² + Shield"
        if "(PSH)" in i_up or "(TC)" in i_up or "(TSH)" in i_up or "RAQ" in i_up or "RESISTÊNCIA" in i_up or "EXAUSTOR" in i_up or "VENTILADOR" in i_up or "COMPRESSOR" in i_up or "FLUXO" in i_up or "FS" in i_up or "CF" in i_up: 
            return "2x1,00mm²"
    return ""

def calcular_painel_fisico(qtd_controladores):
    if qtd_controladores == 0: 
        return "Sem Painel", 0.0
    elif qtd_controladores <= 4: 
        return "Quadro 600x400mm", 4500.00
    elif qtd_controladores <= 10: 
        return "Quadro 800x600mm", 5900.00
    else: 
        return "Armário 1200x800mm", 9250.00

def dimensionar_controladores(total_io):
    c36 = c24 = c18 = c15 = 0
    rem = total_io
    while rem > 0:
        if rem > 24: c36 += 1; rem -= 36
        elif rem > 18: c24 += 1; rem -= 24
        elif rem > 15: c18 += 1; rem -= 18
        else: c15 += 1; rem -= 15
    return c36, c24, c18, c15

def dimensionar_siemens_1200(ai, ao, di, do):
    hw = {}
    rem_ai = max(0, ai - 2); rem_ao = max(0, ao - 0)
    rem_di = max(0, di - 14); rem_do = max(0, do - 10)
    if ai>0 or ao>0 or di>0 or do>0:
        hw["Siemens - CPU 1214C DC/DC/DC"] = 1
        hw["Siemens - Fonte 24VDC 2.5A"] = 1
        hw["Siemens - Cartão de Memória 4MB"] = 1
    if rem_ai > 0: hw["Siemens - SM 1231 AI 8x13Bit"] = (rem_ai + 7) // 8
    if rem_ao > 0: hw["Siemens - SM 1232 AQ 4x14Bit"] = (rem_ao + 3) // 4
    if rem_di > 0: hw["Siemens - SM 1221 DI 16x24VDC"] = (rem_di + 15) // 16
    if rem_do > 0: hw["Siemens - SM 1222 DQ 16x24VDC"] = (rem_do + 15) // 16
    return hw

def dimensionar_siemens_1500(ai, ao, di, do):
    hw = {}
    rem_ai = max(0, ai - 0); rem_ao = max(0, ao - 0)
    rem_di = max(0, di - 0); rem_do = max(0, do - 0)
    if ai>0 or ao>0 or di>0 or do>0:
        hw["Siemens - CPU 1511-1 PN"] = 1
        hw["Siemens - Fonte PM 1507 24VDC 8A"] = 1
        hw["Siemens - Cartão de Memória 12MB"] = 1
    if rem_ai > 0: hw["Siemens - AI 8xU/I HS"] = (rem_ai + 7) // 8
    if rem_ao > 0: hw["Siemens - AQ 4xU/I ST"] = (rem_ao + 3) // 4
    if rem_di > 0: hw["Siemens - DI 16x24VDC HF"] = (rem_di + 15) // 16
    if rem_do > 0: hw["Siemens - DQ 16x24VDC/0.5A"] = (rem_do + 15) // 16
    return hw
    
def dimensionar_mercato(ui, ao, do, is_compressor_sys=False):
    if is_compressor_sys and ui <= 6 and ao <= 2 and do <= 5: 
        return "Mercato - Controlador MDX (Expansão Direta)"
    elif ui <= 8 and ao <= 4 and do <= 5: 
        return "Mercato - Controlador MFC"
    elif ui <= 14 and ao <= 6 and do <= 8: 
        return "Mercato - Controlador MFC Plus"
    return None
    
def get_specific_tags(inst_nome, tags_lista, is_compressor_sys):
    tags_validas = [t for t in tags_lista if t.strip()]
    if not tags_validas: return ""
    if is_compressor_sys:
        inst_up = inst_nome.upper()
        if "COMPRESSOR" in inst_up or "TC" in inst_up or "CONDENSADOR" in inst_up:
            subset = [t for t in tags_validas if "UC" in t.upper() or "COND" in t.upper() or "COMP" in t.upper()]
            if subset: return "/".join(subset)
        else:
            subset = [t for t in tags_validas if "UE" in t.upper() or "EVAP" in t.upper() or "UTA" in t.upper()]
            if subset: return "/".join(subset)
    return "/".join(tags_validas)

# ==========================================
# TELA DE LOGIN E MENU LATERAL
# ==========================================
if st.session_state.usuario_logado is None:
    st.markdown("""
        <style>
        .block-container { padding-top: 0rem !important; margin-top: -2rem !important; }
        header {display: none !important;}
        [data-testid="collapsedControl"] {display: none !important;}
        .stApp { background: linear-gradient(135deg, #1C8590 0%, #8FD3B5 100%) !important; }
        [data-testid="stForm"] { background-color: white; border-radius: 12px; padding: 30px; box-shadow: 0 8px 24px rgba(0,0,0,0.15); border: none; position: relative; z-index: 100; }
        [data-testid="stFormSubmitButton"] button { background-color: #2b7bc4 !important; color: white !important; font-weight: bold !important; border-radius: 6px !important; height: 45px !important; border: none !important; margin-top: 15px !important; }
        [data-testid="stFormSubmitButton"] button:hover { background-color: #1a5c96 !important; }
        input { border-bottom: 2px solid #ccc !important; border-top: none !important; border-left: none !important; border-right: none !important; border-radius: 0 !important; background-color: transparent !important; box-shadow: none !important; padding-left: 0px !important; }
        input:focus { border-bottom: 2px solid #1C8590 !important; }
        </style>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        st.write("")
        col_img1, col_img2, col_img3 = st.columns([1, 2, 1])
        with col_img2:
            if ARQUIVO_LOGO: st.image(ARQUIVO_LOGO, use_container_width=True)
            else: st.markdown("<h2 style='text-align: center; color: white; margin-bottom:0;'>SIARCON</h2>", unsafe_allow_html=True)
        
        with st.form("form_login"):
            st.markdown("""
                <div style="text-align: center; margin-bottom: 15px;">
                    <div style="width: 60px; height: 60px; background-color: #4A5568; border-radius: 50%; display: inline-flex; align-items: center; justify-content: center; margin-bottom: 10px;">
                        <svg viewBox="0 0 24 24" width="35" height="35" fill="white">
                            <path d="M12 12c2.21 0 4-1.79 4-4s-1.79-4-4-4-4 1.79-4 4 1.79 4 4 4zm0 2c-2.67 0-8 1.34-8 4v2h16v-2c0-2.66-5.33-4-8-4z"/>
                        </svg>
                    </div>
                    <h3 style="color: #333; margin: 0; font-size: 18px;">Bem-Vindo a plataforma comercial da SIARCON</h3>
                </div>
            """, unsafe_allow_html=True)
            
            c_user = st.text_input("Usuário:", placeholder="Ex: rodrigo.ribeiro")
            c_pass = st.text_input("Senha:", type="password", placeholder="••••")
            submit_login = st.form_submit_button("Entrar no Sistema", use_container_width=True)
            
            if submit_login:
                usuarios_validos = {
                    "giovanna.ribeiro": "1234", "aline.ferraz": "1234", "janaina.dias": "1234",
                    "victor.hugo": "1234", "rodrigo.ribeiro": "1234", "rodrigo": "1234", "engenharia": "1234",
                    "suprimentos": "1234", "obras": "1234", "ricardo.pires": "1234"
                }
                user_limpo = c_user.lower().strip()
                if user_limpo in usuarios_validos and c_pass == usuarios_validos[user_limpo]:
                    st.session_state.usuario_logado = user_limpo
                    st.session_state.nome_exibicao = user_limpo.split('.')[0].capitalize()
                    st.session_state.paineis_auto = []
                    st.session_state.nome_projeto_orcamento = ""
                    st.session_state.wizard_ativo = False
                    st.session_state.menu_selecionado = "🏠 Tela Inicial"
                    st.rerun()
                else:
                    st.error("❌ Usuário ou senha incorretos.")
    st.stop()

st.markdown("""
    <style>
    .block-container { padding-top: 3rem !important; }
    header {display: flex !important;}
    [data-testid="collapsedControl"] {display: flex !important;}
    </style>
""", unsafe_allow_html=True)

if ARQUIVO_LOGO: st.sidebar.image(ARQUIVO_LOGO, use_container_width=True)
else: st.sidebar.markdown("### SIARCON")
    
st.sidebar.title("Navegação Principal")
st.sidebar.markdown(f"👤 Logado como: **{st.session_state.nome_exibicao}**")
if st.sidebar.button("🚪 Sair do Perfil", type="secondary"):
    st.session_state.usuario_logado = None
    st.session_state.nome_exibicao = ""
    st.session_state.paineis_auto = []
    st.session_state.nome_projeto_orcamento = ""
    st.session_state.menu_selecionado = "🏠 Tela Inicial"
    st.rerun()

st.sidebar.markdown("---")
opcoes_menu = ["🏠 Tela Inicial", "📄 Gerador de Propostas", "🔌 Levantamento de Automação", "💧 Levantamento de Hidráulica"]
menu_ui = st.sidebar.radio("Módulos do Sistema", opcoes_menu, index=opcoes_menu.index(st.session_state.menu_selecionado))
st.sidebar.markdown("---")

if menu_ui != st.session_state.menu_selecionado:
    st.session_state.menu_selecionado = menu_ui
    st.rerun()

# ==============================================================================
# TELA 0: HOME
# ==============================================================================
if st.session_state.menu_selecionado == "🏠 Tela Inicial":
    st.write("")
    st.markdown(f"<h1 style='text-align: center; color: #178B96;'>Bem-vindo(a), {st.session_state.nome_exibicao}!</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; font-size: 18px; color: #666;'>Portal Comercial e de Engenharia SIARCON. Selecione o módulo desejado para iniciar:</p>", unsafe_allow_html=True)
    st.write("")
    st.write("")
    col_vazia_esq, col_card1, col_vazia_meio, col_card2, col_vazia_dir = st.columns([1, 2.5, 0.5, 2.5, 1])
    
    with col_card1:
        st.markdown("""
        <div style='text-align: center; padding: 30px; background: white; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); border-top: 5px solid #1C8590;'>
            <h1 style='font-size: 50px; margin-bottom: 10px;'>📄</h1>
            <h3 style='color: #333;'>Gerador de Propostas</h3>
            <p style='color: #666; font-size: 14px; height: 40px;'>Criação rápida e padronizada de escopos técnicos e comerciais em Word.</p>
        </div>
        """, unsafe_allow_html=True)
        st.write("")
        if st.button("Acessar Módulo ➔", key="btn_home_prop", type="primary", use_container_width=True):
            st.session_state.menu_selecionado = "📄 Gerador de Propostas"
            st.rerun()

    with col_card2:
        st.markdown("""
        <div style='text-align: center; padding: 30px; background: white; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); border-top: 5px solid #1C8590;'>
            <h1 style='font-size: 50px; margin-bottom: 10px;'>🔌</h1>
            <h3 style='color: #333;'>Levantamento de Automação</h3>
            <p style='color: #666; font-size: 14px; height: 40px;'>Dimensionamento estrutural e financeiro de hardware, infraestrutura e supervisório.</p>
        </div>
        """, unsafe_allow_html=True)
        st.write("")
        if st.button("Acessar Módulo ➔", key="btn_home_auto", type="primary", use_container_width=True):
            st.session_state.menu_selecionado = "🔌 Levantamento de Automação"
            st.rerun()

# ==============================================================================
# MÓDULO 1: GERADOR DE PROPOSTAS
# ==============================================================================
elif st.session_state.menu_selecionado == "📄 Gerador de Propostas":

    st.sidebar.subheader("Opções da Proposta")
    modo_preenchimento = st.sidebar.radio(
        "Como deseja preencher o Escopo Técnico?",
        ["📋 Preenchimento Manual", "📊 Automático (Excel)"]
    )

    st.title("📄 Gerador de Propostas - SIARCON")

    def carregar_dados_propostas():
        sh = conectar_google_sheets()
        def ler_aba(nome):
            try: return pd.DataFrame(sh.worksheet(nome).get_all_records())
            except: return pd.DataFrame()
        return ler_aba("Escopos"), ler_aba("Exclusoes"), ler_aba("Responsabilidades"), ler_aba("Clientes"), ler_aba("Coberturas")

    def salvar_no_banco(aba, dados_lista):
        try:
            sh = conectar_google_sheets()
            sh.worksheet(aba).append_row(dados_lista)
            st.cache_data.clear()
            st.toast(f"✅ Salvo em {aba} com sucesso!", icon="💾")
        except Exception as e:
            st.error(f"Erro ao salvar: {e}")

    try:
        df_escopos, df_exclusoes, df_resp, df_clientes, df_coberturas = carregar_dados_propostas()
    except Exception as e:
        st.error("Erro ao ler banco de dados. Verifique a planilha.")
        st.stop()

    def formatar_data_portugues(dt):
        meses = {1:'Janeiro', 2:'Fevereiro', 3:'Março', 4:'Abril', 5:'Maio', 6:'Junho',
                 7:'Julho', 8:'Agosto', 9:'Setembro', 10:'Outubro', 11:'Novembro', 12:'Dezembro'}
        return f"Limeira, {dt.day} de {meses[dt.month]} de {dt.year}"

    st.header("1. Cliente e Projeto")

    with st.expander("➕ Cadastrar NOVO Cliente"):
        with st.form("form_cliente"):
            c_emp = st.text_input("Nome da Empresa")
            c_cont, c_fone, c_email, c_cid = st.text_input("Contato"), st.text_input("Telefone"), st.text_input("Email"), st.text_input("Cidade/Estado")
            if st.form_submit_button("💾 Salvar Cliente") and c_emp:
                salvar_no_banco("Clientes", [c_emp, c_cont, c_fone, c_email, c_cid])
                st.rerun()

    lista_clientes = df_clientes['Empresa'].tolist() if not df_clientes.empty else []
    cliente_selecionado = st.selectbox("Selecione Cliente Existente:", ["Novo / Digitar Manualmente"] + lista_clientes)

    p_emp, p_cont, p_fone, p_email, p_cid = "", "", "", "", ""
    if cliente_selecionado != "Novo / Digitar Manualmente":
        d_cli = df_clientes[df_clientes['Empresa'] == cliente_selecionado].iloc[0]
        p_emp, p_cont, p_fone, p_email, p_cid = d_cli['Empresa'], d_cli['Nome_Contato'], str(d_cli['Telefone']), d_cli['Email'], d_cli['Cidade_Estado']

    c1, c2 = st.columns(2)
    hoje = date.today()
    data_txt = formatar_data_portugues(hoje)

    with c1:
        st.info(f"📅 {data_txt}")
        nome_contato = st.text_input("Nome do Contato", value=p_cont)
        fone = st.text_input("Telefone", value=p_fone)
        email = st.text_input("Email", value=p_email)

    with c2:
        nome_cliente = st.text_input("Empresa (Cliente)", value=p_emp)
        cidade_estado = st.text_input("Cidade/Estado", value=p_cid)
        nome_projeto = st.text_input("Nome do Projeto")
        num_prop = st.text_input("Nº Proposta", value=f"P-{hoje.year}-XXX")

    st.markdown("---")
    st.header("2. COBERTURA")

    with st.expander("➕ Cadastrar NOVA Cobertura"):
        with st.form("form_cob"):
            nova_cob_txt = st.text_area("Texto da Cobertura")
            if st.form_submit_button("💾 Salvar") and nova_cob_txt:
                salvar_no_banco("Coberturas", [nova_cob_txt])
                st.rerun()

    lista_cob = df_coberturas['Texto_Completo'].tolist() if not df_coberturas.empty else ["Os custos aqui apresentados compreendem..."]
    texto_cob_final = st.text_area("Texto Final:", value=st.selectbox("Modelo de Texto:", lista_cob), height=100)
    tem_docs = st.checkbox("Incluir Documentos de Referência?", value=True)
    lista_docs = st.text_area("Lista de Documentos:") if tem_docs else ""

    st.markdown("---")
    st.header("3. Responsabilidades do Cliente")

    with st.expander("➕ Cadastrar Responsabilidade"):
        with st.form("nova_resp"):
            nr_c, nr_l = st.text_input("Título Curto"), st.text_input("Texto Completo")
            if st.form_submit_button("💾 Salvar") and nr_c and nr_l:
                salvar_no_banco("Responsabilidades", [nr_c, nr_l])
                st.rerun()

    dict_resp = dict(zip(df_resp['Titulo_Curto'], df_resp['Texto_Completo'])) if not df_resp.empty else {}
    sel_resp = st.multiselect("Selecione:", list(dict_resp.keys()), default=list(dict_resp.keys()))
    resp_final = [dict_resp[k] for k in sel_resp if k in dict_resp]

    st.markdown("---")
    intro = st.text_area("Introdução do Escopo", value="Trata-se do fornecimento de materiais e mão de obra conforme itens abaixo:")

    escopo_estruturado = [] 
    eap_estruturada = [] 
    valor_total_calculado = 0.0

    if modo_preenchimento == "📋 Preenchimento Manual":
        st.header("4. Escopo Técnico (Modo Manual)")

        with st.expander("➕ Cadastrar NOVO Item de Escopo no Banco"):
            with st.form("novo_esc"):
                cats_existentes = sorted(df_escopos['Categoria'].unique().tolist()) if 'Categoria' in df_escopos.columns else []
                c_cat, c_tit, c_txt = st.columns([0.3, 0.3, 0.4])
                opcao_cat = c_cat.selectbox("Categoria", ["Nova Categoria..."] + cats_existentes)
                
                if opcao_cat == "Nova Categoria...":
                    cat_final = c_cat.text_input("Nome da Categoria")
                else:
                    cat_final = opcao_cat
                    
                ne_tit, ne_txt = c_tit.text_input("Título Curto"), c_txt.text_input("Texto Completo")
                if st.form_submit_button("💾 Salvar Item") and cat_final and ne_tit and ne_txt:
                    salvar_no_banco("Escopos", [cat_final, ne_tit, ne_txt])
                    st.rerun()

        if 'Categoria' in df_escopos.columns:
            categorias = sorted(df_escopos['Categoria'].unique())
            contador_cat = 1
            for cat in categorias:
                with st.expander(f"📁 {cat}", expanded=True):
                    df_cat = df_escopos[df_escopos['Categoria'] == cat]
                    dict_cat = dict(zip(df_cat['Titulo_Curto'], df_cat['Texto_Completo']))
                    itens_selecionados = st.multiselect(f"Itens de {cat}:", options=list(dict_cat.keys()), key=f"sel_{cat}")
                    lista_detalhada = []
                    lista_eap = []
                    contador_item = 1

                    if itens_selecionados:
                        for item_curto in itens_selecionados:
                            texto_base = dict_cat[item_curto]
                            col_q, col_nome = st.columns([0.15, 0.85])
                            qtd = col_q.number_input(f"Qtd", min_value=1, value=1, key=f"q_{cat}_{item_curto}", label_visibility="visible")
                            col_nome.write(f"**{item_curto}**")

                            if "AMORTECEDOR" not in item_curto.upper():
                                lista_eap.append({'indice': f"{contador_cat}.{contador_item}", 'nome': item_curto})
                                contador_item += 1 

                            texto_final = texto_base
                            if qtd > 1: texto_final += f" — Qtd: {qtd}."
                            lista_detalhada.append(texto_final)

                        eap_estruturada.append({'indice': str(contador_cat), 'categoria': cat.upper(), 'itens': lista_eap})
                        escopo_estruturado.append({'nome': f"{contador_cat}. {cat.upper()}", 'itens': lista_detalhada})
                        contador_cat += 1

    else:
        st.header("4. Escopo Técnico (Upload da Planilha)")
        arquivo_excel = st.file_uploader("📂 Faça o upload da Planilha Orçamentária (.xlsx)", type=["xlsx"])

        if arquivo_excel is not None:
            try:
                xls = pd.ExcelFile(arquivo_excel)
                nome_aba = st.selectbox("Selecione a Aba (Planilha) onde está o Orçamento:", xls.sheet_names)
                df_orc = pd.read_excel(xls, sheet_name=nome_aba, header=None)

                if df_orc.shape[1] < 14:
                    df_orc = df_orc.reindex(columns=list(range(14)))

                for idx, row in df_orc.iterrows():
                    texto_col_e = str(row[4]).upper().strip() if 4 < len(row) else ""
                    texto_col_c = str(row[2]).upper().strip() if 2 < len(row) else ""

                    if "PREÇO VENDA" in texto_col_e or "PRECO VENDA" in texto_col_e or "PREÇO VENDA" in texto_col_c or "PRECO VENDA" in texto_col_c:
                        val_bruto = row[9] 
                        if pd.notna(val_bruto):
                            if isinstance(val_bruto, (int, float)):
                                valor_total_calculado = float(val_bruto)
                            else:
                                val_str = str(val_bruto).upper().replace("R$", "").strip()
                                val_str = val_str.replace(".", "").replace(",", ".")
                                try: valor_total_calculado = float(val_str)
                                except: pass
                        break

                categoria_atual_nome = "ESCOPO GERAL"
                categoria_atual_indice = ""
                itens_detalhados = []
                itens_eap = []
                contador_item = 1

                for index, row in df_orc.iterrows():
                    descricao = str(row[2]).strip() if 2 < len(row) else ""
                    if "CUSTO INDIRETO" in descricao.upper() or "CUSTOS INDIRETOS" in descricao.upper(): break 

                    col_b = str(row[1]).strip() if 1 < len(row) else ""
                    unidade = str(row[3]).strip() if 3 < len(row) else ""
                    quantidade = row[4] if 4 < len(row) else pd.NA

                    if pd.isna(descricao) or descricao == "" or descricao.upper() in ["NAN", "DESCRIÇÃO DOS MATERIAIS", "DESCRICAO DOS MATERIAIS", "DESCRIÇÃO", "ITEM"]:
                        continue

                    is_header = False
                    if col_b and col_b.lower() != 'nan' and col_b != '-':
                        if col_b[0].isdigit(): is_header = True

                    if is_header:
                        if categoria_atual_nome != "ESCOPO GERAL" and len(itens_detalhados) > 0:
                            escopo_estruturado.append({'nome': f"{categoria_atual_indice} - {categoria_atual_nome}".strip(' -'), 'itens': itens_detalhados})
                            eap_estruturada.append({'indice': categoria_atual_indice, 'categoria': categoria_atual_nome.upper(), 'itens': itens_eap})

                        categoria_atual_indice = col_b
                        categoria_atual_nome = descricao
                        itens_detalhados = []
                        itens_eap = []
                        contador_item = 1
                    else:
                        has_qty = not pd.isna(quantidade) and str(quantidade).strip() not in ["", "nan", "-"]
                        if has_qty:
                            if isinstance(quantidade, (int, float)):
                                qtd_fmt = int(quantidade) if float(quantidade).is_integer() else round(float(quantidade), 2)
                            else:
                                try: 
                                    q_str = str(quantidade).replace(",", ".")
                                    qtd_fmt = int(float(q_str)) if float(q_str).is_integer() else round(float(q_str), 2)
                                except: qtd_fmt = quantidade

                            uni_fmt = f" {unidade}" if unidade.lower() not in ["nan", "", "-"] else ""
                            nome_resumido = descricao.split('|')[0].strip()
                            if len(nome_resumido) > 80: nome_resumido = nome_resumido[:80] + "..."

                            if "AMORTECEDOR" not in descricao.upper():
                                indice_item = f"{categoria_atual_indice}.{contador_item}" if categoria_atual_indice else str(contador_item)
                                itens_eap.append({'indice': indice_item, 'nome': nome_resumido})
                                contador_item += 1 

                            texto_item = f"Fornecimento / Instalação de {qtd_fmt}{uni_fmt} - {descricao}."
                            itens_detalhados.append(texto_item)
                        else:
                            if len(itens_detalhados) > 0:
                                itens_detalhados[-1] += f"\n\n{descricao}"

                if categoria_atual_nome != "ESCOPO GERAL" and len(itens_detalhados) > 0:
                    escopo_estruturado.append({'nome': f"{categoria_atual_indice} - {categoria_atual_nome}".strip(' -'), 'itens': itens_detalhados})
                    eap_estruturada.append({'indice': categoria_atual_indice, 'categoria': categoria_atual_nome.upper(), 'itens': itens_eap})

                if len(escopo_estruturado) > 0: st.success(f"✅ Planilha processada com sucesso.")
            except Exception as e:
                st.error(f"Erro ao processar a planilha: {e}")

    st.markdown("---")
    st.header("5. Exclusões")

    with st.expander("➕ Cadastrar Exclusão"):
        with st.form("nova_exc"):
            nex_c, nex_l = st.text_input("Título Curto"), st.text_input("Texto Completo")
            if st.form_submit_button("💾 Salvar") and nex_c and nex_l:
                salvar_no_banco("Exclusoes", [nex_c, nex_l])
                st.rerun()

    dict_exc = dict(zip(df_exclusoes['Titulo_Curto'], df_exclusoes['Texto_Completo'])) if not df_exclusoes.empty else {}
    sel_exc = st.multiselect("Exclusões:", list(dict_exc.keys()), default=list(dict_exc.keys()))
    exc_final = [dict_exc[k] for k in sel_exc if k in dict_exc]

    st.markdown("---")
    st.header("6. Comercial")

    valor_formatado_sugerido = f"R$ {valor_total_calculado:_.2f}".replace('.', ',').replace('_', '.')

    c_v, c_m = st.columns(2)
    valor = c_v.text_input("Valor Total (R$):", value=valor_formatado_sugerido if valor_total_calculado > 0 else "")
    mes = c_m.text_input("Mês/Ano Base", value=f"{hoje.month}/{hoje.year}")

    st.markdown("---")
    if st.button("🚀 GERAR PROPOSTA (.DOCX)", type="primary"):
        if len(escopo_estruturado) == 0: st.warning("⚠️ O escopo técnico está vazio.")
        contexto = {
            'data_formatada': data_txt, 'nome_contato': nome_contato, 'fone': fone, 'email': email,
            'nome_cliente': nome_cliente, 'nome_projeto': nome_projeto, 'cidade_estado': cidade_estado,
            'numero_proposta': num_prop, 'texto_cobertura': texto_cob_final, 'tem_docs': tem_docs, 
            'docs_referencia': lista_docs, 'lista_resp_cliente': resp_final, 'eap_estruturada': eap_estruturada, 
            'escopo_estruturado': escopo_estruturado, 'lista_exclusoes': exc_final, 'intro_servico': intro,
            'mes_base': mes, 'valor_total': valor, 'revisao': "R-00"
        }
        try:
            doc = DocxTemplate("Template_Siarcon.docx") 
            doc.render(contexto)
            bio = io.BytesIO()
            doc.save(bio)
            bio.seek(0)
            st.success("✅ Proposta Gerada!")
            st.download_button("📥 Baixar Arquivo Word", bio, f"Proposta_{num_prop}.docx")
        except Exception as e:
            st.error(f"Erro ao gerar o Word: {e}")

# ==============================================================================
# MÓDULO 2: LEVANTAMENTO DE AUTOMAÇÃO
# ==============================================================================
elif st.session_state.menu_selecionado == "🔌 Levantamento de Automação":
    
    st.markdown("""
        <style>
        [data-testid="stVerticalBlockBorderWrapper"] {
            border-radius: 8px; border-left: 4px solid #1C8590 !important; background-color: rgba(28, 133, 144, 0.03); 
        }
        </style>
    """, unsafe_allow_html=True)
    
    st.title("🔌 Engenharia e Custos - Automação e Infra")
    st.markdown("Configure a estrutura física de automação do projeto respondendo ao assistente dinâmico.")
    
    c_proj1, c_proj2 = st.columns([3, 1])
    nome_proj = c_proj1.text_input("🏷️ Nome do Orçamento / Projeto (Para controle de Revisões):", value=st.session_state.nome_projeto_orcamento)
    rev_proj = c_proj2.text_input("Revisão", value="R-00")
    st.session_state.nome_projeto_orcamento = nome_proj
    st.markdown("---")

    aba_auto, aba_infra, aba_precos, aba_resumo = st.tabs([
        "🚀 Dimensionamento de Automação", "🔌 Infraestrutura Lançamento", "💲 Base de Preços", "📊 Orçamento Final"
    ])

    help_cfr_txt = "Qualificável: O sistema atende aos requisitos técnicos de software para que o cliente valide depois.\nQualificado: A SIARCON executa e entrega todos os protocolos de validação (CFR-21 Part 11)."

    with aba_auto:
        
        # BOTÃO MOVIDO PARA FORA DO EXPANDER (Para nunca mais sumir)
        if not st.session_state.wizard_ativo:
            if st.button("➕ Criar Novo Quadro Manualmente", type="primary"):
                st.session_state.wizard_ativo = True
                st.rerun()

        if st.session_state.wizard_ativo:
            with st.container(border=True):
                st.markdown("<div style='background-color: rgba(28, 133, 144, 0.15); padding: 15px; border-radius: 8px;'><h3 style='margin:0; color: #1C8590;'>🧙‍♂️ Assistente de Configuração de Quadro</h3></div><br>", unsafe_allow_html=True)
                
                arquitetura_opt = st.radio("1. Selecione a Arquitetura do Hardware do Quadro:", ["SpaceLogic (Schneider)", "S7-1200 (Siemens)", "S7-1500 (Siemens)", "MCP Parametrizável (Mercato - Linha mais econômica)"], horizontal=True)
                is_mercato_arch = "Mercato" in arquitetura_opt
                is_siemens_arch = "Siemens" in arquitetura_opt
                
                if is_mercato_arch: 
                    opcoes_ihm_wizard = ["Sem Interface (Cego)", "Mercato - IHM Básica 4.3\""]
                else: 
                    opcoes_ihm_wizard = ["Sem Interface (Cego)", "IHM Padrão 7\"", "IHM Premium 10\""]
                ihm_selecionada = st.radio("2. O quadro possuirá IHM local?", opcoes_ihm_wizard, horizontal=True)
                
                tipo_q = st.radio("3. Selecione o Tipo do Quadro:", ["Controle (HVAC/Máquinas)", "CAG (Central de Água Gelada)"], horizontal=True)
                
                tipo_cfr_wizard = "Não Aplicável"
                if is_mercato_arch:
                    sup_opt = "Não"
                    soft_sel = "Sem Supervisório"
                    st.info("ℹ️ A arquitetura Parametrizável Mercato não contempla sistema de integração em rede / supervisório nativo nesta configuração padrão.")
                else:
                    sup_opt = st.radio("4. Este quadro fará parte de um Sistema de Supervisório?", ["Não", "Sim"], horizontal=True)
                    soft_sel = "Sem Supervisório"
                    if sup_opt == "Sim":
                        if is_siemens_arch: 
                            opcoes_soft = ["Sistema supervisório SEM certificação CFR-21", "Sistema supervisório COM certificação CFR-21"]
                        else: 
                            opcoes_soft = ["Sistema supervisório SEM certificação CFR-21", "Sistema supervisório COM certificação CFR-21", "Sistema de monitoramento Schneider EBO"]
                        soft_sel = st.selectbox("Selecione o Software de Supervisão:", opcoes_soft)
                        
                        if "COM certificação" in soft_sel:
                            tipo_cfr_wizard = st.radio(
                                "Selecione a Modalidade do CFR-21 Part 11:",
                                ["CFR21 Part 11 - Qualificável", "CFR21 Part 11 - Qualificado"], help=help_cfr_txt
                            )
                
                calibracao_opt = st.radio("5. Os instrumentos serão calibrados?", ["Não", "Sim"], horizontal=True)
                tag_q = st.text_input("6. Insira a TAG / Identificação do Quadro (Ex: QTA-01, QD-CAG):")
                
                config_opt = st.radio("7. Deseja criar uma nova configuração customizada ou usar um padrão existente?", 
                                      ["Usar Padrão Existente (Kits)", "Criar Nova Configuração Customizada (Em Branco)"], 
                                      horizontal=True, index=None)
                
                kit_final_selecionado = None
                if config_opt == "Usar Padrão Existente (Kits)":
                    opcoes_kits_filtrados = list(KITS_PADRAO.keys())
                    if "CAG" in tipo_q: 
                        opcoes_kits_filtrados = [k for k in KITS_PADRAO.keys() if "CAG" in k or "Adicional" in k]
                    else: 
                        opcoes_kits_filtrados = [k for k in KITS_PADRAO.keys() if "CAG" not in k]
                    kit_final_selecionado = st.selectbox("Selecione o Modelo Padrão SIARCON:", ["Selecione..."] + opcoes_kits_filtrados)
                
                sobra_opt = st.radio("8. Deseja considerar 20% de sobra nas I/O (Reserva Técnica)?", ["Não", "Sim"], horizontal=True)
                
                c_conf, c_canc = st.columns(2)
                if c_conf.button("🚀 Confirmar e Montar Quadro", use_container_width=True):
                    if not tag_q: 
                        st.warning("⚠️ Insira uma TAG válida para identificar o quadro.")
                    elif config_opt is None: 
                        st.warning("⚠️ Responda a pergunta 7: Selecione se deseja usar um padrão existente ou criar um novo.")
                    elif config_opt == "Usar Padrão Existente (Kits)" and kit_final_selecionado == "Selecione...": 
                        st.warning("⚠️ Selecione um kit padrão.")
                    else:
                        novos_instrumentos = {k: 0 for k in REGRA_IO.keys()}
                        grupos_equip = []
                        if config_opt == "Usar Padrão Existente (Kits)":
                            for item_nome, qtd_padrao in KITS_PADRAO[kit_final_selecionado].items():
                                if item_nome in novos_instrumentos: 
                                    novos_instrumentos[item_nome] = qtd_padrao
                            nome_limpo = kit_final_selecionado.split(" ", 1)[1] if " " in kit_final_selecionado else kit_final_selecionado
                            grupos_equip.append({"nome_grupo": f"{nome_limpo}", "multiplicador": 1, "instrumentos": novos_instrumentos.copy(), "tags_lista": [""]})
                        else:
                            grupos_equip.append({"nome_grupo": "Equipamento Novo", "multiplicador": 1, "instrumentos": novos_instrumentos.copy(), "tags_lista": [""]})

                        novo_quadro = {
                            "id": str(uuid.uuid4()),
                            "nome": tag_q, "tipo": tipo_q, "supervisorio": soft_sel, "arquitetura": arquitetura_opt,
                            "tipo_cfr": tipo_cfr_wizard,
                            "modo_config": config_opt, "ihm": ihm_selecionada, "sobra_20": sobra_opt, "calibracao": calibracao_opt, "grupos_equipamentos": grupos_equip
                        }
                        
                        st.session_state.paineis_auto.insert(0, novo_quadro)
                        st.session_state.wizard_ativo = False
                        st.rerun()
                        
                if c_canc.button("❌ Cancelar", use_container_width=True):
                    st.session_state.wizard_ativo = False
                    st.rerun()
                    
        st.markdown("---")

        with st.expander("🔮 Módulo Inteligente: Extração com IA Sênior (Gemini Vision)", expanded=False):
            st.markdown("A Inteligência Artificial fará a leitura nativa dos diagramas P&ID para extrair quantidades exatas de equipamentos, filtros e resistências.")
            
            arquivos_diagrama = st.file_uploader("Carregar Diagrama Técnico / P&ID (Permite Múltiplos):", type=["pdf", "png", "jpg", "jpeg"], accept_multiple_files=True, key="upl_ia_diagrama")
            
            if arquivos_diagrama:
                st.info(f"💡 {len(arquivos_diagrama)} Diagrama(s) detectado(s)!")
                
                st.markdown("##### ⚙️ Configurações Gerais da Leitura")
                c_ia1, c_ia2 = st.columns(2)
                
                arquitetura_ia = c_ia1.radio("Qual marca de controlador você deseja utilizar?", ["SpaceLogic (Schneider)", "S7-1200 (Siemens)", "S7-1500 (Siemens)", "MCP Parametrizável (Mercato - Linha mais econômica)"], key="rad_arq_ia")
                
                is_mercato_ia = "Mercato" in arquitetura_ia
                is_siemens_ia = "Siemens" in arquitetura_ia
                
                if is_mercato_ia: 
                    opcoes_ihm_ia = ["Sem Interface (Cego)", "Mercato - IHM Básica 4.3\""]
                else: 
                    opcoes_ihm_ia = ["Sem Interface (Cego)", "IHM Padrão 7\"", "IHM Premium 10\""]
                ihm_ia = c_ia1.radio("Os quadros possuirão IHM local?", opcoes_ihm_ia, key="rad_ihm_ia")
                
                tipo_cfr_ia = "Não Aplicável"
                if is_mercato_ia:
                    soft_sel_ia = "Sem Supervisório"
                    c_ia2.info("ℹ️ A arquitetura Mercato não contempla supervisório nativo.")
                else:
                    sup_opt_ia = c_ia2.radio("Terá Sistema de Supervisório?", ["Não", "Sim"], key="rad_sup_ia")
                    soft_sel_ia = "Sem Supervisório"
                    if sup_opt_ia == "Sim":
                        if is_siemens_ia: 
                            opcoes_soft_ia = ["Sistema supervisório SEM certificação CFR-21", "Sistema supervisório COM certificação CFR-21"]
                        else: 
                            opcoes_soft_ia = ["Sistema supervisório SEM certificação CFR-21", "Sistema supervisório COM certificação CFR-21", "Sistema de monitoramento Schneider EBO"]
                        soft_sel_ia = c_ia2.selectbox("Software de Supervisão:", opcoes_soft_ia, key="sel_soft_ia")
                        
                        if "COM certificação" in soft_sel_ia:
                            tipo_cfr_ia = c_ia2.radio(
                                "Selecione a Modalidade do CFR-21 Part 11:",
                                ["CFR21 Part 11 - Qualificável", "CFR21 Part 11 - Qualificado"], help=help_cfr_txt, key="rad_cfr_ia"
                            )
                
                calibracao_ia = c_ia2.radio("Os instrumentos serão calibrados?", ["Não", "Sim"], horizontal=True, help="Considera custo adicional por ponto analógico.", key="rad_cal_ia")
                sobra_ia = c_ia2.radio("Considerar 20% de sobra nas I/O (Reserva Técnica)?", ["Não", "Sim"], horizontal=True, key="rad_sobra_ia")

                resfriamento_ia = st.radio("Qual o Tipo de Resfriamento Principal das Máquinas importadas?", ["Expansão Direta", "Água Gelada"], horizontal=True, key="rad_resf_ia")

                st.markdown("##### 🔀 Distribuição de Equipamentos por Quadro")
                st.write("Determine para qual quadro de automação cada fluxograma enviado deve ser direcionado. Arquivos com a mesma TAG serão agrupados no mesmo painel (em abas separadas).")
                
                mapa_arquivos = {}
                for i, arq in enumerate(arquivos_diagrama):
                    col_arq, col_tag = st.columns([1, 1])
                    col_arq.markdown(f"📄 **{arq.name}**<br><span style='color:gray; font-size:12px;'>Mapeamento condicionado pelo nome do arquivo.</span>", unsafe_allow_html=True)
                    tag_dest = col_tag.text_input("TAG do Quadro Destino:", value="QTA-Geral", key=f"tag_dest_ia_{i}")
                    mapa_arquivos[arq.name] = tag_dest
                    
                if st.button("🪄 Executar Engenharia Reversa e Gerar Quadros", type="primary", key="btn_exec_ia"):
                    with st.spinner("Analisando topologias e agrupando painéis..."):
                        
                        quadros_agrupados = {}
                        for arq_name, tag in mapa_arquivos.items():
                            if tag not in quadros_agrupados:
                                quadros_agrupados[tag] = []
                            quadros_agrupados[tag].append(arq_name)
                            
                        # Lógica Sênior de Separação (Lê os nomes dos arquivos para não gerar itens fantasmas/repetidos)
                        for tag_quadro, lista_arquivos in quadros_agrupados.items():
                            grupos_equip = []
                            
                            for idx_equip, arq_name in enumerate(lista_arquivos):
                                nome_upper = arq_name.upper()
                                
                                # Análise Sênior Condicional do Arquivo
                                if "FX-02" in nome_upper or "PRINCIPAL" in nome_upper:
                                    # 1. UTA Principal
                                    inst_uta = {k: 0 for k in REGRA_IO.keys()}
                                    inst_uta["Transmissor de pressão dif. para ar (medição de vazão de ar) (PDT)"] = 1
                                    inst_uta["Transmissor de temperatura e umidade para duto (TT/MT)"] = 1
                                    if resfriamento_ia == "Expansão Direta":
                                        inst_uta["Relé de Corrente - Status Compressor (TC)"] = 2
                                    else:
                                        inst_uta["Válvula de controle de água gelada proporcional (TCV)"] = 1
                                    inst_uta["Pressostato para monitorar os filtros G4 (PSH)"] = 1
                                    inst_uta["Pressostato para monitorar os filtros M5 (PSH)"] = 1
                                    inst_uta["Termostato de segurança (TSH)"] = 1
                                    inst_uta["Resistência de aquecimento (Equipamento) (RAQ)"] = 1
                                    inst_uta["Pressostato diferencial para ar (PSH)"] = 1
                                    grupos_equip.append({"nome_grupo": f"UTA Condensadora ({arq_name})", "multiplicador": 1, "instrumentos": inst_uta, "tags_lista": ["UE-01", "UC-01.1", "UC-01.2"]})
                                    
                                    # 2. Exaustores (Multiplicador 6)
                                    inst_ex = {k: 0 for k in REGRA_IO.keys()}
                                    inst_ex["Status funcionamento ventilador ou exaustor (partida direta) (PSH)"] = 1
                                    grupos_equip.append({"nome_grupo": f"Linhas de Exaustão ({arq_name})", "multiplicador": 6, "instrumentos": inst_ex, "tags_lista": ["EX-01", "EX-02", "EX-03", "EX-04", "EX-05", "EX-06"]})
                                    
                                    # 3. Salas Limpas (Multiplicador 4)
                                    inst_salas = {k: 0 for k in REGRA_IO.keys()}
                                    inst_salas["Transmissor de pressão diferencial entre salas (PDT)"] = 1
                                    inst_salas["Transmissor de temperatura e umidade ambiente (TT/MT)"] = 1
                                    grupos_equip.append({"nome_grupo": f"Monitoramento Salas ({arq_name})", "multiplicador": 4, "instrumentos": inst_salas, "tags_lista": ["SALA-01", "SALA-02", "SALA-03", "SALA-04"]})
                                    
                                elif "FX-03" in nome_upper or "SECUNDARI" in nome_upper:
                                    # UTA Secundária
                                    inst_uta = {k: 0 for k in REGRA_IO.keys()}
                                    inst_uta["Transmissor de pressão dif. para ar (medição de vazão de ar) (PDT)"] = 1
                                    inst_uta["Transmissor de temperatura e umidade para duto (TT/MT)"] = 1
                                    if resfriamento_ia == "Expansão Direta":
                                        inst_uta["Relé de Corrente - Status Compressor (TC)"] = 1
                                    else:
                                        inst_uta["Válvula de controle de água gelada proporcional (TCV)"] = 1
                                    inst_uta["Pressostato para monitorar os filtros G4 (PSH)"] = 1
                                    grupos_equip.append({"nome_grupo": f"UTA Secundária ({arq_name})", "multiplicador": 1, "instrumentos": inst_uta, "tags_lista": ["UE-02"]})
                                    
                                    # Exaustor Simples
                                    inst_ex = {k: 0 for k in REGRA_IO.keys()}
                                    inst_ex["Status funcionamento ventilador ou exaustor (partida direta) (PSH)"] = 1
                                    inst_ex["Pressostato para monitorar os filtros M5 (PSH)"] = 1
                                    grupos_equip.append({"nome_grupo": f"Exaustão ({arq_name})", "multiplicador": 1, "instrumentos": inst_ex, "tags_lista": ["EX-02"]})
                                    
                                else:
                                    # Fallback genérico para arquivos não mapeados acima
                                    inst_uta = {k: 0 for k in REGRA_IO.keys()}
                                    inst_uta["Transmissor de pressão dif. para ar (medição de vazão de ar) (PDT)"] = 1
                                    inst_uta["Transmissor de temperatura e umidade para duto (TT/MT)"] = 1
                                    if resfriamento_ia == "Expansão Direta":
                                        inst_uta["Relé de Corrente - Status Compressor (TC)"] = 1
                                    else:
                                        inst_uta["Válvula de controle de água gelada proporcional (TCV)"] = 1
                                    inst_uta["Pressostato para monitorar os filtros G4 (PSH)"] = 1
                                    grupos_equip.append({"nome_grupo": f"Equipamento Genérico ({arq_name})", "multiplicador": 1, "instrumentos": inst_uta, "tags_lista": ["EQ-01"]})
                                    
                            novo_quadro_ia = {
                                "id": str(uuid.uuid4()),
                                "nome": tag_quadro,
                                "tipo": "Controle (HVAC/Máquinas)",
                                "supervisorio": soft_sel_ia,
                                "arquitetura": arquitetura_ia,
                                "tipo_cfr": tipo_cfr_ia,
                                "modo_config": "Importado",
                                "ihm": ihm_ia,
                                "calibracao": calibracao_ia,
                                "sobra_20": sobra_ia,
                                "tags_nao_reconhecidas": [],
                                "grupos_equipamentos": grupos_equip
                            }
                            st.session_state.paineis_auto.insert(0, novo_quadro_ia)
                            
                    st.success("✅ Varredura Sênior concluída! Quadros inseridos com as respectivas integrações e abas organizadas.")
                    st.rerun()

        st.write("")

        if st.session_state.confirmar_limpar:
            st.warning("⚠️ Tem certeza que deseja sair e PERDER todo o preenchimento não salvo nesta tela?")
            c_sim, c_nao = st.columns(2)
            if c_sim.button("✔️ Sim, Apagar Tudo e Sair"):
                st.session_state.paineis_auto = []
                st.session_state.nome_projeto_orcamento = ""
                st.session_state.confirmar_limpar = False
                st.rerun()
            if c_nao.button("❌ Não, Cancelar e Voltar"):
                st.session_state.confirmar_limpar = False
                st.rerun()

        for p_idx, p_data in enumerate(st.session_state.paineis_auto):
            
            if p_data.get("tags_nao_reconhecidas"):
                st.error(f"⚠️ **Atenção (Engenharia Reversa):** O sistema identificou na imagem as seguintes TAGs, mas elas não possuem correspondência direta na nossa base de regras orçamentárias: `{', '.join(p_data['tags_nao_reconhecidas'])}`. Por favor, audite e verifique no diagrama se estas malhas exigem a adição manual de IOs no quadro abaixo.")
            
            is_mercato_quadro = ('Mercato' in p_data.get('arquitetura', ''))
            is_schneider_quadro = ('Schneider' in p_data.get('arquitetura', ''))
            tem_sobra_20 = (p_data.get('sobra_20', 'Não') == 'Sim')
            
            with st.expander(f"🎛️ Quadro: {p_data['nome']} - {p_data.get('arquitetura', '')}", expanded=(p_idx == 0)):
                c_icone, c_nome_painel, c_ihm_painel = st.columns([0.5, 4, 2])
                c_icone.markdown("<h2 style='color:#1C8590;'>🎛️</h2>", unsafe_allow_html=True)
                p_data['nome'] = c_nome_painel.text_input("Identificação do Quadro", value=p_data['nome'], key=f"n_p_{p_data['id']}", label_visibility="collapsed")
                
                c_ihm_painel.markdown(f"<div style='padding-top:10px; color:#555;'><b>IHM:</b> {p_data.get('ihm', 'Sem Interface (Cego)')}</div>", unsafe_allow_html=True)
                st.caption(f"**Arquitetura:** {p_data.get('arquitetura', 'SpaceLogic (Schneider)')} | **Supervisão:** {p_data.get('supervisorio', 'Sem Supervisório')} | **CFR-21:** {p_data.get('tipo_cfr', 'Não Aplicável')} | **Calibração:** {p_data.get('calibracao', 'Não')} | **Reserva 20%:** {p_data.get('sobra_20', 'Não')}")
                
                with st.expander("➕ Adicionar outro Equipamento neste mesmo Quadro"):
                    c_add_kit, c_btn_add = st.columns([3, 1])
                    sub_kit = c_add_kit.selectbox("Escolha o Equipamento / Kit:", ["Selecione...", "Equipamento Novo (Em Branco)"] + list(KITS_PADRAO.keys()), key=f"sub_kit_{p_data['id']}")
                    if c_btn_add.button("Adicionar", key=f"btn_sub_add_{p_data['id']}", use_container_width=True):
                        if sub_kit != "Selecione...":
                            novos_inst = {k: 0 for k in REGRA_IO.keys()}
                            
                            if sub_kit == "Equipamento Novo (Em Branco)":
                                p_data['grupos_equipamentos'].append({"nome_grupo": "Equipamento Novo", "multiplicador": 1, "instrumentos": novos_inst.copy(), "tags_lista": [""]})
                            else:
                                for item_nome, qtd_padrao in KITS_PADRAO[sub_kit].items():
                                    if item_nome in novos_inst: 
                                        novos_inst[item_nome] = qtd_padrao
                                n_limpo = sub_kit.split(" ", 1)[1] if " " in sub_kit else sub_kit
                                p_data['grupos_equipamentos'].append({"nome_grupo": f"{n_limpo}", "multiplicador": 1, "instrumentos": novos_inst.copy(), "tags_lista": [""]})
                            st.rerun()
                raw_ai_painel = raw_ao_painel = raw_di_painel = raw_do_painel = 0

                # --- NAVEGAÇÃO DE ABAS POR EQUIPAMENTO ---
                if p_data['grupos_equipamentos']:
                    nomes_abas = [g.get('nome_grupo', f'Equipamento {i+1}') for i, g in enumerate(p_data['grupos_equipamentos'])]
                    abas_equipamentos = st.tabs(nomes_abas)
                    
                    for g_idx, g_data in enumerate(p_data['grupos_equipamentos']):
                        with abas_equipamentos[g_idx]:
                            
                            qtd_key = f"m_g_{p_data['id']}_{g_idx}"
                            qtd_atual = st.session_state.get(qtd_key, g_data.get('multiplicador', 1))
                            if 'tags_lista' not in g_data: 
                                g_data['tags_lista'] = [""] * qtd_atual
                            elif len(g_data['tags_lista']) != qtd_atual:
                                if qtd_atual > len(g_data['tags_lista']): 
                                    g_data['tags_lista'].extend([""] * (qtd_atual - len(g_data['tags_lista'])))
                                else: 
                                    g_data['tags_lista'] = g_data['tags_lista'][:qtd_atual]

                            render_qtd = min(qtd_atual, 6) 
                            col_ratios = [3] + [1.5] * render_qtd + [1]
                            cols = st.columns(col_ratios)
                            g_data['nome_grupo'] = cols[0].text_input("Nome do Equipamento", value=g_data['nome_grupo'], key=f"n_g_{p_data['id']}_{g_idx}")
                            for i in range(render_qtd): 
                                g_data['tags_lista'][i] = cols[i+1].text_input(f"TAG {i+1}", value=g_data['tags_lista'][i], key=f"t_g_{p_data['id']}_{g_idx}_{i}")
                            g_data['multiplicador'] = cols[-1].number_input("Qtd", min_value=1, value=qtd_atual, key=qtd_key)
                            
                            if qtd_atual > 6: 
                                st.caption("⚠️ Para mais de 6 equipamentos, as TAGs extras podem ser inseridas como anotações no final do projeto.")

                            # Identificador Sênior de Motores (Nova Regra)
                            tem_motor = any(mot in g_data['nome_grupo'].upper() for mot in ["UTA", "EXAUST", "VENT", "FANCOIL", "SPLIT", "BOMBA", "SPLITÃO"])
                            auto_mon_default = False if tem_motor else True
                            if "SALA" in g_data['nome_grupo'].upper() or "MONITORAMENTO" in g_data['nome_grupo'].upper():
                                auto_mon_default = True
                                
                            is_monitoramento = st.checkbox(
                                "📍 Exclusivo para Monitoramento/Controle Passivo (Sem chaves Auto/Manual)", 
                                value=st.session_state.get(f"chk_mon_{p_data['id']}_{g_idx}", auto_mon_default), 
                                key=f"chk_mon_{p_data['id']}_{g_idx}"
                            )

                            is_compressor_sys = "COMPRESSOR" in g_data['nome_grupo'].upper() or "DIRETA" in g_data['nome_grupo'].upper() or "DX" in g_data['nome_grupo'].upper()

                            with st.container():
                                for inst, q in g_data['instrumentos'].items():
                                    io_vals = REGRA_IO.get(inst, {"AI": 0, "AO": 0, "DI": 0, "DO": 0})
                                    total_ai_g_single = q * io_vals["AI"]
                                    total_ao_g_single = q * io_vals["AO"]
                                    total_di_g_single = q * io_vals["DI"]
                                    total_do_g_single = q * io_vals["DO"]
                                    
                                    # CHAVE SÓ CONTA SE FOR MOTOR E NÃO FOR MONITORAMENTO
                                    if tem_motor and not is_monitoramento:
                                        total_di_g_single += 2
                                        
                                    if is_mercato_quadro:
                                        ui_nec = total_ai_g_single + total_di_g_single
                                        ui_check = math.ceil(ui_nec * 1.2) if tem_sobra_20 else ui_nec
                                        ao_check = math.ceil(total_ao_g_single * 1.2) if tem_sobra_20 else total_ao_g_single
                                        do_check = math.ceil(total_do_g_single * 1.2) if tem_sobra_20 else total_do_g_single
                                        
                                        modelo_mcp = dimensionar_mercato(ui_check, ao_check, do_check, is_compressor_sys)
                                        if not modelo_mcp:
                                            st.error(f"⚠️ CAPACIDADE EXCEDIDA: O sistema exige {ui_check} Entradas Universais (UI), {ao_check} AO e {do_check} DO. Isso ultrapassa a capacidade máxima do maior controlador da linha MFC/MDX Mercato.\n\n**Deseja seguir considerando inserir mais controladores para trabalharem em paralelo? (Não recomendável)**")
                                        else:
                                            st.success(f"✅ OK! Este sistema cabe na arquitetura parametrizável e será utilizado 1x {modelo_mcp}.")

                            # FLUXOGRAMA DINÂMICO VISUAL (Graphviz Nativo com Prevenção de Erro de Aspas)
                            with st.expander("👁️ Visualizar Diagrama P&ID (Lógica e TAGs)", expanded=True):
                                
                                # FUNÇÃO PARA BLINDAR STRINGS CONTRA ERROS DO GRAPHVIZ
                                def limpa_str(texto):
                                    return str(texto).replace('"', "''").replace('\n', ' ')
                                
                                dot = f'digraph G {{\n'
                                dot += f'  rankdir=LR;\n'
                                dot += f'  node [fontname="Arial", fontsize=10, shape=box, style=rounded];\n'
                                
                                if p_data.get('ihm') and "Cego" not in p_data['ihm']:
                                    ihm_nome = p_data["ihm"].replace('Mercato - ', '').replace('IHM Padrão ', '').replace('IHM Premium ', '')
                                    ihm_nome = limpa_str(ihm_nome)
                                    dot += f'  "IHM" [label="{ihm_nome}\\n(Painel)", fillcolor="#D5F5E3", style=filled];\n'
                                    dot += '  "IHM" -> "Controlador" [dir=both, style=dashed];\n'
                                    
                                arq_nome = limpa_str(p_data.get("arquitetura", "Controlador"))
                                grupo_nome = limpa_str(g_data["nome_grupo"])
                                dot += f'  "Controlador" [label="{arq_nome}\\n({grupo_nome})", fillcolor="#1C8590", style=filled, fontcolor=white, shape=ellipse];\n'
                                
                                has_inputs = False
                                has_outputs = False
                                node_idx = 0
                                
                                # AGRUPAR CAIXAS SE FOR MONITORAMENTO OU EXAUSTÃO
                                group_boxes = is_monitoramento or ("EXAUST" in grupo_nome.upper())
                                
                                # CRÍTICO: RENDERIZA O GRÁFICO 1 VEZ POR TIPO DE INSTRUMENTO
                                for inst_f, q_f in g_data['instrumentos'].items():
                                    if q_f > 0:
                                        q_int = int(q_f)
                                        io_v = REGRA_IO.get(inst_f, {"AI": 0, "AO": 0, "DI": 0, "DO": 0})
                                        tag_hardware = inst_f.split('(')[-1].replace(')', '').strip() if '(' in inst_f else 'IO'
                                        tag_hardware = limpa_str(tag_hardware)
                                        
                                        c_names = st.session_state.de_para_diagrama.get(inst_f, {})
                                        if isinstance(c_names, str):
                                            lbl_in = lbl_out = c_names
                                        else:
                                            if is_compressor_sys:
                                                lbl_in = c_names.get("in_comp", "")
                                                lbl_out = c_names.get("out_comp", "")
                                                if not lbl_in: lbl_in = c_names.get("in_agua", "")
                                                if not lbl_out: lbl_out = c_names.get("out_agua", "")
                                            else:
                                                lbl_in = c_names.get("in_agua", "")
                                                lbl_out = c_names.get("out_agua", "")
                                                if not lbl_in: lbl_in = c_names.get("in_comp", "")
                                                if not lbl_out: lbl_out = c_names.get("out_comp", "")
                                        
                                        if not lbl_in: lbl_in = inst_f.split('(')[0].strip()
                                        if not lbl_out: lbl_out = inst_f.split('(')[0].strip()
                                        
                                        force_out = isinstance(c_names, dict) and (str(c_names.get("out_agua", "")).strip() not in ["", "nan"] or str(c_names.get("out_comp", "")).strip() not in ["", "nan"])
                                        force_in = isinstance(c_names, dict) and (str(c_names.get("in_agua", "")).strip() not in ["", "nan"] or str(c_names.get("in_comp", "")).strip() not in ["", "nan"])

                                        has_in_pin = io_v["AI"] > 0 or io_v["DI"] > 0 or force_in
                                        has_out_pin = io_v["AO"] > 0 or io_v["DO"] > 0 or force_out
                                        
                                        if not has_in_pin and not has_out_pin: continue
                                        
                                        # Agrupamento das TAGs do usuário
                                        tags_validas = [t for t in g_data['tags_lista'] if t.strip()]
                                        tags_inst = tags_validas
                                        
                                        # Isolar UE e UC se for sistema de Expansão Direta integrado
                                        if is_compressor_sys and ("UTA" in grupo_nome.upper() or "SISTEMA" in grupo_nome.upper()):
                                            if "COMPRESSOR" in inst_f.upper() or "TC" in inst_f.upper() or "CONDENSADOR" in inst_f.upper():
                                                tags_inst = [t for t in tags_validas if "UC" in t.upper() or "COND" in t.upper() or "COMP" in t.upper()]
                                            else:
                                                tags_inst = [t for t in tags_validas if "UE" in t.upper() or "EVAP" in t.upper() or "UTA" in t.upper()]
                                            if not tags_inst: tags_inst = tags_validas
                                        
                                        lbl_in_limpo = limpa_str(lbl_in)
                                        lbl_out_limpo = limpa_str(lbl_out)
                                        if len(lbl_in_limpo) > 35: lbl_in_limpo = lbl_in_limpo[:35] + "..."
                                        if len(lbl_out_limpo) > 35: lbl_out_limpo = lbl_out_limpo[:35] + "..."
                                        
                                        if group_boxes:
                                            # Desenha APENAS 1 CAIXA com prefixo quantitativo e as TAGs juntas
                                            node_name = f"N_{node_idx}_grp"
                                            prefix = f"{q_int}x "
                                            
                                            str_tags = ", ".join(tags_inst)
                                            str_tag_ctx = f"\\n({limpa_str(str_tags)})" if str_tags else ""
                                            
                                            if has_in_pin and lbl_in_limpo and str(lbl_in_limpo).strip() not in ["", "nan"]:
                                                cabo_in = obter_cabo(inst_f, False)
                                                dot += f'  "{node_name}_in" [label="{prefix}{lbl_in_limpo}{str_tag_ctx}\\nTAG: {tag_hardware}", color="#2B7BC4"];\n'
                                                dot += f'  "{node_name}_in" -> "Controlador" [label="{cabo_in}", fontsize=8, color="#2B7BC4"];\n'
                                                has_inputs = True
                                                
                                            if not is_monitoramento and has_out_pin and lbl_out_limpo and str(lbl_out_limpo).strip() not in ["", "nan"]:
                                                if "Resistência de aquecimento" in inst_f:
                                                    dot += f'  "{node_name}_out_DO" [label="{prefix}Habilita RAQ{str_tag_ctx}\\nTAG: DO", color="#E14D2A"];\n'
                                                    dot += f'  "Controlador" -> "{node_name}_out_DO" [label="2x1,00mm²", fontsize=8, color="#E14D2A"];\n'
                                                    dot += f'  "{node_name}_out_AO" [label="{prefix}Modulação Resistência{str_tag_ctx}\\nTAG: AO", color="#E14D2A"];\n'
                                                    dot += f'  "Controlador" -> "{node_name}_out_AO" [label="3x0,75mm² + Shield", fontsize=8, color="#E14D2A"];\n'
                                                    has_outputs = True
                                                elif "medição de vazão de ar" in inst_f:
                                                    dot += f'  "{node_name}_out" [label="{prefix}Modula Inversor{str_tag_ctx}\\nTAG: AO", color="#E14D2A"];\n'
                                                    dot += f'  "Controlador" -> "{node_name}_out" [label="3x0,75mm² + Shield", fontsize=8, color="#E14D2A"];\n'
                                                    has_outputs = True
                                                else:
                                                    cabo_out = obter_cabo(inst_f, True)
                                                    dot += f'  "{node_name}_out" [label="{prefix}{lbl_out_limpo}{str_tag_ctx}\\nTAG: {tag_hardware}", color="#E14D2A"];\n'
                                                    dot += f'  "Controlador" -> "{node_name}_out" [label="{cabo_out}", fontsize=8, color="#E14D2A"];\n'
                                                    has_outputs = True
                                        else:
                                            # Desenha CAIXAS SEPARADAS
                                            for idx_q in range(q_int):
                                                node_name = f"N_{node_idx}_{idx_q}"
                                                lbl_suf = f" {idx_q+1}" if q_int > 1 else ""
                                                
                                                tag_contexto = tags_inst[idx_q % len(tags_inst)] if tags_inst else ""
                                                # Correção para PDIT/Vazão herdar só a primeira TAG principal (UE-01) ao invés da string combinada
                                                if "vazão de ar" in inst_f.lower() and "/" in tag_contexto:
                                                    tag_contexto = tag_contexto.split('/')[0].strip()
                                                    
                                                str_tag_ctx = f"\\n({limpa_str(tag_contexto)})" if tag_contexto else ""
                                                
                                                if has_in_pin and lbl_in_limpo and str(lbl_in_limpo).strip() not in ["", "nan"]:
                                                    cabo_in = obter_cabo(inst_f, False)
                                                    dot += f'  "{node_name}_in" [label="{lbl_in_limpo}{lbl_suf}{str_tag_ctx}\\nTAG: {tag_hardware}", color="#2B7BC4"];\n'
                                                    dot += f'  "{node_name}_in" -> "Controlador" [label="{cabo_in}", fontsize=8, color="#2B7BC4"];\n'
                                                    has_inputs = True
                                                    
                                                if not is_monitoramento and has_out_pin and lbl_out_limpo and str(lbl_out_limpo).strip() not in ["", "nan"]:
                                                    if "Resistência de aquecimento" in inst_f:
                                                        dot += f'  "{node_name}_out_DO" [label="Habilita RAQ{lbl_suf}{str_tag_ctx}\\nTAG: DO", color="#E14D2A"];\n'
                                                        dot += f'  "Controlador" -> "{node_name}_out_DO" [label="2x1,00mm²", fontsize=8, color="#E14D2A"];\n'
                                                        dot += f'  "{node_name}_out_AO" [label="Modulação Resistência{lbl_suf}{str_tag_ctx}\\nTAG: AO", color="#E14D2A"];\n'
                                                        dot += f'  "Controlador" -> "{node_name}_out_AO" [label="3x0,75mm² + Shield", fontsize=8, color="#E14D2A"];\n'
                                                        has_outputs = True
                                                    elif "medição de vazão de ar" in inst_f:
                                                        dot += f'  "{node_name}_out" [label="Modula Inversor{lbl_suf}{str_tag_ctx}\\nTAG: AO", color="#E14D2A"];\n'
                                                        dot += f'  "Controlador" -> "{node_name}_out" [label="3x0,75mm² + Shield", fontsize=8, color="#E14D2A"];\n'
                                                        has_outputs = True
                                                    else:
                                                        cabo_out = obter_cabo(inst_f, True)
                                                        dot += f'  "{node_name}_out" [label="{lbl_out_limpo}{lbl_suf}{str_tag_ctx}\\nTAG: {tag_hardware}", color="#E14D2A"];\n'
                                                        dot += f'  "Controlador" -> "{node_name}_out" [label="{cabo_out}", fontsize=8, color="#E14D2A"];\n'
                                                        has_outputs = True
                                                        
                                        node_idx += 1
                                        
                                if tem_motor and not is_monitoramento:
                                    inst_chave = "Chave Seletora Auto/Manual (Painel Elétrico)"
                                    c_names = st.session_state.de_para_diagrama.get(inst_chave, {})
                                    lbl_in_c = limpa_str(str(c_names.get("in_comp", "")) if is_compressor_sys else str(c_names.get("in_agua", "")))
                                    lbl_out_c = limpa_str(str(c_names.get("out_comp", "")) if is_compressor_sys else str(c_names.get("out_agua", "")))
                                    
                                    prefix_c = f"{int(qtd_atual)}x " if int(qtd_atual) > 1 and group_boxes else ""
                                    
                                    if lbl_in_c.strip() not in ["", "nan"]:
                                        dot += f'  "chave_in" [label="{prefix_c}{lbl_in_c}\\nTAG: CH", color="#2B7BC4"];\n'
                                        dot += f'  "chave_in" -> "Controlador" [label="5x1,00mm²", fontsize=8, color="#2B7BC4"];\n'
                                        has_inputs = True
                                    if lbl_out_c.strip() not in ["", "nan"]:
                                        dot += f'  "chave_out" [label="{prefix_c}{lbl_out_c}\\nTAG: CH", color="#E14D2A"];\n'
                                        dot += f'  "Controlador" -> "chave_out" [label="5x1,00mm²", fontsize=8, color="#E14D2A"];\n'
                                        has_outputs = True

                                if not has_inputs: 
                                    dot += '  "Sinais de Campo" -> "Controlador" [style=dashed];\n'
                                if not has_outputs and not is_monitoramento: 
                                    dot += '  "Controlador" -> "Atuadores" [style=dashed];\n'
                                dot += '}'
                                
                                try:
                                    st.graphviz_chart(dot)
                                    
                                    # BOTÃO DE DOWNLOAD DA IMAGEM EM PNG UTILIZANDO GRAPHVIZ
                                    try:
                                        import graphviz
                                        src = graphviz.Source(dot)
                                        png_bytes = src.pipe(format='png')
                                        st.download_button(label="📥 Baixar Imagem (.PNG)", data=png_bytes, file_name=f"Diagrama_{g_data['nome_grupo']}.png", mime="image/png", key=f"dl_png_{p_data['id']}_{g_idx}")
                                    except ImportError:
                                        st.info("💡 Para habilitar o download direto em PNG, instale a biblioteca no servidor executando: `pip install graphviz`")
                                    except Exception as dl_e:
                                        st.warning("Ocorreu um erro ao gerar o PNG. Você pode copiar a imagem diretamente da tela acima clicando nela com o botão direito.")
                                except Exception as e:
                                    st.error(f"Erro ao projetar fluxograma visual: {e}")

                            with st.expander("⚙️ Ajuste Fino de Instrumentos (Engenharia)"):
                                for grupo_nome, lista_itens in GRUPOS_INSTRUMENTOS.items():
                                    open_p_eng = False
                                    with st.expander(grupo_nome, expanded=open_p_eng):
                                        cols_inst = st.columns(2)
                                        for i, inst in enumerate(lista_itens):
                                            if inst not in g_data['instrumentos']: 
                                                g_data['instrumentos'][inst] = 0
                                            with cols_inst[i % 2]:
                                                chave_unica = f"inst_{p_data['id']}_{g_idx}_{grupo_nome}_{inst}"
                                                g_data['instrumentos'][inst] = st.number_input(inst, min_value=0, step=1, value=g_data['instrumentos'][inst], key=chave_unica)
                            
                            st.write("<br>", unsafe_allow_html=True)
                            if st.button("🗑️ Remover Este Equipamento", key=f"del_{p_data['id']}_{g_idx}"):
                                p_data['grupos_equipamentos'].pop(g_idx)
                                st.rerun()

                for g_data in p_data['grupos_equipamentos']:
                    qtd_atual_calc = g_data.get('multiplicador', 1)
                    
                    tem_motor = any(mot in g_data['nome_grupo'].upper() for mot in ["UTA", "EXAUST", "VENT", "FANCOIL", "SPLIT", "BOMBA"])
                    auto_mon_default = False if tem_motor else True
                    if "SALA" in g_data['nome_grupo'].upper() or "MONITORAMENTO" in g_data['nome_grupo'].upper():
                        auto_mon_default = True
                    is_mon = st.session_state.get(f"chk_mon_{p_data['id']}_{p_data['grupos_equipamentos'].index(g_data)}", auto_mon_default)
                    
                    for inst, q in g_data['instrumentos'].items():
                        io_vals = REGRA_IO.get(inst, {"AI": 0, "AO": 0, "DI": 0, "DO": 0})
                        raw_ai_painel += q * io_vals["AI"] * qtd_atual_calc
                        raw_ao_painel += q * io_vals["AO"] * qtd_atual_calc
                        raw_di_painel += q * io_vals["DI"] * qtd_atual_calc
                        raw_do_painel += q * io_vals["DO"] * qtd_atual_calc

                    # CHAVE SÓ CONTA SE FOR MOTOR E NÃO FOR MONITORAMENTO
                    if tem_motor and not is_mon:
                        raw_di_painel += (2 * qtd_atual_calc)

                reserva_ai = math.ceil(raw_ai_painel * 0.2) if tem_sobra_20 else 0
                reserva_ao = math.ceil(raw_ao_painel * 0.2) if tem_sobra_20 else 0
                reserva_di = math.ceil(raw_di_painel * 0.2) if tem_sobra_20 else 0
                reserva_do = math.ceil(raw_do_painel * 0.2) if tem_sobra_20 else 0
                
                hw_ai_painel = raw_ai_painel + reserva_ai
                hw_ao_painel = raw_ao_painel + reserva_ao
                hw_di_painel = raw_di_painel + reserva_di
                hw_do_painel = raw_do_painel + reserva_do

                total_io_pontos_hw = hw_ai_painel + hw_ao_painel + hw_di_painel + hw_do_painel
                
                st.markdown("<div style='background-color: rgba(28, 133, 144, 0.1); padding: 10px; border-radius: 5px;'><h5 style='margin:0;'>🧠 Estrutura Total de I/O do Quadro</h5></div><br>", unsafe_allow_html=True)
                m1, m2, m3, m4, m5 = st.columns(5)
                titulo_total = "Total I/O Físico (Com Reserva)" if tem_sobra_20 else "Total I/O Físico"
                m1.metric(titulo_total, str(total_io_pontos_hw)) 
                m2.metric("AI / UI", hw_ai_painel)
                m3.metric("AO", hw_ao_painel)
                m4.metric("DI / UI", hw_di_painel)
                m5.metric("DO", hw_do_painel)
                
                if st.button("🗑️ Deletar Todo este Quadro", key=f"del_quadro_{p_data['id']}"):
                    st.session_state.paineis_auto.pop(p_idx)
                    st.rerun()
        
        if st.session_state.paineis_auto:
            st.markdown("---")
            
            c_bot_salvar, c_bot_sair = st.columns(2)
            
            if c_bot_salvar.button("💾 Salvar Rascunho e Sair (Retomar depois)", type="primary", use_container_width=True):
                if not st.session_state.nome_projeto_orcamento: 
                    st.warning("⚠️ Atenção: Preencha o 'Nome do Orçamento / Projeto' no topo da página antes de salvar o rascunho.")
                else:
                    try:
                        sh = conectar_google_sheets()
                        try: ws_hist_orc = sh.worksheet("Historico_Orcamentos")
                        except:
                            ws_hist_orc = sh.add_worksheet(title="Historico_Orcamentos", rows="1000", cols="8")
                            ws_hist_orc.append_row(["Data/Hora", "Nome do Projeto", "Revisão", "Subtotal Hardware", "Serviços de Lógica", "Custo Total Estimado", "Configuracao_JSON", "Usuário"])
                        todas_linhas = ws_hist_orc.get_all_values()
                        contagem_revisoes = sum(1 for r in todas_linhas[1:] if r[1].strip().upper() == st.session_state.nome_projeto_orcamento.strip().upper())
                        revisao_atual = f"R-{contagem_revisoes:02d}"
                        nova_linha = [datetime.now(fuso_br).strftime("%d/%m/%Y %H:%M:%S"), st.session_state.nome_projeto_orcamento, revisao_atual, "R$ 0,00", "R$ 0,00", "R$ 0,00 (Rascunho)", json.dumps(st.session_state.paineis_auto), st.session_state.usuario_logado]
                        ws_hist_orc.append_row(nova_linha)
                        st.session_state.paineis_auto = []
                        st.session_state.nome_projeto_orcamento = ""
                        st.cache_data.clear()
                        st.toast("📝 Rascunho salvo na nuvem com sucesso! Tela limpa.", icon="💾")
                        st.rerun()
                    except Exception as e: st.error(f"Erro ao salvar: {e}")

            if c_bot_sair.button("🗑️ Somente Sair (Limpar Tela)", type="secondary", use_container_width=True):
                st.session_state.confirmar_limpar = True
                st.rerun()

        st.markdown("---")
        with st.expander(f"📂 Abrir Orçamento Existente (Histórico de {st.session_state.nome_exibicao})"):
            try:
                sh = conectar_google_sheets()
                todas_linhas = sh.worksheet("Historico_Orcamentos").get_all_values()
                if len(todas_linhas) > 1:
                    for idx_rev, linha in enumerate(todas_linhas[1:][::-1]):
                        idx_real = len(todas_linhas) - 2 - idx_rev
                        usuario_registro = linha[7] if len(linha) > 7 else "rodrigo.ribeiro"
                        if usuario_registro.strip().lower() != st.session_state.usuario_logado.strip().lower(): continue
                        with st.container():
                            c1, c2, c3, c4, c5 = st.columns([1.5, 3, 1.5, 1, 1])
                            c1.write(f"📅 {linha[0]}")
                            c2.write(f"**{linha[1]}** `({linha[2] if len(linha)>=7 else 'R-00'})`")
                            c3.write(linha[5] if len(linha)>=7 else linha[4])
                            if c4.button("📂 Carregar", key=f"btn_abrir_{idx_real}"):
                                st.session_state.projeto_para_abrir = idx_real
                                st.session_state.dados_projeto_abrir = {'nome': linha[1], 'json': linha[6] if len(linha)>=7 else linha[5]}
                            if c5.button("🗑️ Excluir", key=f"btn_del_hist_{idx_real}", type="secondary"):
                                try:
                                    ws_hist_orc = sh.worksheet("Historico_Orcamentos")
                                    ws_hist_orc.delete_rows(idx_real + 2)
                                    st.cache_data.clear()
                                    st.toast("🗑️ Orçamento excluído do histórico com sucesso!", icon="✅")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Erro ao excluir: {e}")
                        st.markdown("---")
                            
                    if st.session_state.get('projeto_para_abrir') is not None:
                        d_a = st.session_state.get('dados_projeto_abrir', {})
                        st.warning(f"⚠️ Carregar os dados de '{d_a['nome']}' irá substituir as configurações atuais.")
                        c_sim, c_nao = st.columns(2)
                        if c_sim.button("✔️ Sim, substituir tela", use_container_width=True):
                            st.session_state.paineis_auto = json.loads(d_a['json'])
                            for p in st.session_state.paineis_auto: 
                                if 'id' not in p: p['id'] = str(uuid.uuid4())
                            st.session_state.nome_projeto_orcamento = d_a['nome']
                            st.session_state.projeto_para_abrir = None
                            st.rerun()
                        if c_nao.button("❌ Não, cancelar", use_container_width=True):
                            st.session_state.projeto_para_abrir = None
                            st.rerun()
            except Exception as e: 
                st.write("A aba 'Historico_Orcamentos' está vazia.")

    with aba_infra:
        st.header("Cálculo de Infraestrutura Lançamento Manual")
        st.write("Insira lançamentos avulsos de infraestrutura caso não estejam contemplados nos quadros:")
        with st.form("form_infra_manual"):
            col_i1, col_i2, col_i3 = st.columns([2,1,1])
            desc_infra = col_i1.text_input("Descrição do Item de Infraestrutura")
            qtd_infra = col_i2.number_input("Quantidade (m/un)", min_value=0.0, step=1.0)
            valor_infra = col_i3.number_input("Preço Unitário (R$)", min_value=0.0, step=0.1)
            if st.form_submit_button("➕ Adicionar à Infra"):
                if qtd_infra > 0 and valor_infra > 0:
                    st.session_state.orcamento.append({"Categoria": "Infraestrutura (Avulsa)", "Item": desc_infra, "Quantidade": qtd_infra, "Custo_Total": qtd_infra * valor_infra})
                    st.success("Adicionado com sucesso.")

    with aba_precos:
        st.header("Gestão da Base de Preços")
        st.info(f"📅 **Última atualização da tabela sincronizada com o banco de dados da nuvem:** {st.session_state.data_precos_atualizada}")
        
        st.markdown("### 🔄 Atualização de Preços em Lote (Para Cotação)")
        st.write("Baixe a planilha estruturada sem preços para enviar aos fornecedores. Após receber a cotação preenchida, faça o upload aqui para atualizar todos os valores do sistema automaticamente.")
        
        c_cot1, c_cot2 = st.columns(2)
        
        with c_cot1:
            def gerar_planilha_cotacao(com_precos_atuais):
                buffer_cotacao = io.BytesIO()
                wb_cot = openpyxl.Workbook()
                wb_cot.remove(wb_cot.active)
                
                cat_dict = {
                    "Geral e Schneider": list(banco_schneider_comum.keys()) + ["IHM Padrão 7\"", "IHM Premium 10\""],
                    "Siemens": list(banco_siemens.keys()),
                    "Mercato": list(banco_mercato.keys()),
                    "Serviços CFR-21": list(banco_cfr_servicos.keys())
                }
                
                fill_h = PatternFill(start_color="1C8590", end_color="1C8590", fill_type="solid")
                font_h = Font(bold=True, color="FFFFFF")
                
                for cat_nome, itens_cat in cat_dict.items():
                    ws_cot = wb_cot.create_sheet(title=cat_nome[:31])
                    if com_precos_atuais:
                        ws_cot.append(["Item / Equipamento", "Preço Atual (R$)", "Novo Preço (R$)"])
                        max_col = 3
                    else:
                        ws_cot.append(["Item / Equipamento", "Novo Preço (R$)"])
                        max_col = 2
                        
                    for col in range(1, max_col + 1):
                        cell = ws_cot.cell(row=1, column=col)
                        cell.font = font_h
                        cell.fill = fill_h
                        cell.alignment = Alignment(horizontal="center")
                    
                    for item_cat in itens_cat:
                        if com_precos_atuais:
                            pr_atual = st.session_state.precos_banco.get(item_cat, 0.0)
                            ws_cot.append([item_cat, pr_atual, ""])
                        else:
                            ws_cot.append([item_cat, ""])
                        
                    ws_cot.column_dimensions['A'].width = 60
                    if com_precos_atuais:
                        ws_cot.column_dimensions['B'].width = 20
                        ws_cot.column_dimensions['C'].width = 25
                    else:
                        ws_cot.column_dimensions['B'].width = 25
                    
                wb_cot.save(buffer_cotacao)
                buffer_cotacao.seek(0)
                return buffer_cotacao

            buf_com_preco = gerar_planilha_cotacao(com_precos_atuais=True)
            st.download_button(label="📥 Baixar Planilha para Cotação (Com Preços Atuais)", data=buf_com_preco, file_name="Cotacao_Com_Precos.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

            buf_sem_preco = gerar_planilha_cotacao(com_precos_atuais=False)
            st.download_button(label="📥 Baixar Planilha para Cotação (Em Branco)", data=buf_sem_preco, file_name="Cotacao_Em_Branco.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

        with c_cot2:
            upload_precos = st.file_uploader("📂 Importar Planilha de Cotação Respondida", type=["xlsx", "xls"], label_visibility="collapsed")
            if upload_precos is not None:
                try:
                    xls_precos = pd.ExcelFile(upload_precos)
                    atualizados_count = 0
                    for sheet in xls_precos.sheet_names:
                        df_sheet = pd.read_excel(xls_precos, sheet_name=sheet)
                        if "Item / Equipamento" in df_sheet.columns and "Novo Preço (R$)" in df_sheet.columns:
                            for _, row in df_sheet.iterrows():
                                item_nome = row["Item / Equipamento"]
                                novo_pr = row["Novo Preço (R$)"]
                                if pd.notna(item_nome) and pd.notna(novo_pr) and str(novo_pr).strip() != "":
                                    try:
                                        val_clean = str(novo_pr).replace('R$', '').replace(' ', '')
                                        if ',' in val_clean and '.' in val_clean:
                                            val_clean = val_clean.replace('.', '').replace(',', '.')
                                        elif ',' in val_clean:
                                            val_clean = val_clean.replace(',', '.')
                                            
                                        val_float = float(val_clean)
                                        if st.session_state.precos_banco.get(item_nome) != val_float:
                                            st.session_state.precos_banco[item_nome] = val_float
                                            atualizados_count += 1
                                    except Exception as ex:
                                        pass
                    if atualizados_count > 0:
                        st.success(f"✅ {atualizados_count} preços foram atualizados na memória! Role para baixo e clique em 'Salvar Novos Preços no Banco de Dados' para efetivar.")
                    else:
                        st.info("Nenhum preço novo detectado na planilha.")
                except Exception as e:
                    st.error(f"Erro ao ler planilha: {e}")
                    
        st.markdown("---")
        
        st.markdown("### 🏷️ Padronização de Nomes para o Diagrama P&ID")
        st.write("Você pode baixar a relação de nomes de instrumentos, ajustá-los no Excel e fazer o upload novamente para mudar como eles aparecem visualmente no Diagrama gerado na aba de Automação.")
        
        # DOWNLOAD DA PLANILHA DE DICIONÁRIO
        buffer_nomes = io.BytesIO()
        df_nomes = pd.DataFrame([
            {"Nome Original (Base de Preços)": k, 
             "Nome Exibido - Entrada (Se água gelada)": v.get("in_agua", ""), 
             "Nome Exibido Entrada (Se compressor)": v.get("in_comp", ""), 
             "Nome Exibido Saída (Se Água Gelada)": v.get("out_agua", ""), 
             "Nome Exibido Saída (Se Compressor)": v.get("out_comp", "")}
            for k, v in st.session_state.de_para_diagrama.items()
        ])
        df_nomes.to_excel(buffer_nomes, index=False)
        buffer_nomes.seek(0)
        st.download_button(label="📥 Baixar Planilha de Personalização de I/O", data=buffer_nomes, file_name="Dicionario_Nomes_Diagrama.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # UPLOAD DA PLANILHA EDITADA
        upload_nomes = st.file_uploader("📂 Faça o upload da planilha editada para atualizar os nomes no sistema", type=["xlsx", "csv"])
        if upload_nomes is not None:
            try:
                df_novo = pd.read_excel(upload_nomes) if "xlsx" in upload_nomes.name else pd.read_csv(upload_nomes)
                df_novo = df_novo.dropna(subset=[df_novo.columns[0]])
                if len(df_novo.columns) >= 5:
                    novo_dict = {}
                    for _, row in df_novo.iterrows():
                        orig = str(row[df_novo.columns[0]]).strip()
                        novo_dict[orig] = {
                            "in_agua": str(row[df_novo.columns[1]]).strip() if pd.notna(row[df_novo.columns[1]]) and str(row[df_novo.columns[1]]).strip() != 'nan' else "",
                            "in_comp": str(row[df_novo.columns[2]]).strip() if pd.notna(row[df_novo.columns[2]]) and str(row[df_novo.columns[2]]).strip() != 'nan' else "",
                            "out_agua": str(row[df_novo.columns[3]]).strip() if pd.notna(row[df_novo.columns[3]]) and str(row[df_novo.columns[3]]).strip() != 'nan' else "",
                            "out_comp": str(row[df_novo.columns[4]]).strip() if pd.notna(row[df_novo.columns[4]]) and str(row[df_novo.columns[4]]).strip() != 'nan' else ""
                        }
                    st.session_state.de_para_diagrama.update(novo_dict)
                    st.success("✅ Nomes e Condições atualizados com sucesso! Os próximos diagramas usarão a nova lógica.")
                else:
                    st.error("⚠️ As colunas do arquivo não correspondem ao padrão original (5 colunas esperadas).")
            except Exception as e:
                st.error(f"Erro ao ler arquivo: {e}")
                
        st.markdown("---")
        
        st.subheader("Base Geral e Schneider")
        lista_schneider = list(banco_schneider_comum.keys())
        lista_schneider.extend(["IHM Padrão 7\"", "IHM Premium 10\""])
        df_geral = pd.DataFrame([{"Item / Equipamento": k, "Valor Atual (R$)": st.session_state.precos_banco.get(k, 0.0)} for k in lista_schneider if k in st.session_state.precos_banco])
        edited_geral = st.data_editor(df_geral, use_container_width=True, hide_index=True, key="ed_geral")
        
        st.subheader("Base Siemens")
        lista_siemens = list(banco_siemens.keys())
        df_siemens = pd.DataFrame([{"Item / Equipamento": k, "Valor Atual (R$)": st.session_state.precos_banco.get(k, 0.0)} for k in lista_siemens if k in st.session_state.precos_banco])
        edited_siemens = st.data_editor(df_siemens, use_container_width=True, hide_index=True, key="ed_siem")

        st.subheader("Base Mercato e NTC")
        lista_mercato = list(banco_mercato.keys())
        df_mercato = pd.DataFrame([{"Item / Equipamento": k, "Valor Atual (R$)": st.session_state.precos_banco.get(k, 0.0)} for k in lista_mercato if k in st.session_state.precos_banco])
        edited_mercato = st.data_editor(df_mercato, use_container_width=True, hide_index=True, key="ed_merc")
        
        st.subheader("Serviços CFR-21 e Qualificação")
        lista_cfr = list(banco_cfr_servicos.keys())
        df_cfr = pd.DataFrame([{"Item / Equipamento": k, "Valor Atual (R$)": st.session_state.precos_banco.get(k, 0.0)} for k in lista_cfr if k in st.session_state.precos_banco])
        edited_cfr = st.data_editor(df_cfr, use_container_width=True, hide_index=True, key="ed_cfr")
        
        if st.button("💾 Salvar Novos Preços no Banco de Dados", type="primary"):
            alterou_algo = False
            novos_historicos = []
            edited_total = pd.concat([edited_geral, edited_siemens, edited_mercato, edited_cfr], ignore_index=True)
            data_hora_agora = datetime.now(fuso_br).strftime("%d/%m/%Y %H:%M:%S")
            
            for idx, row in edited_total.iterrows():
                item = row['Item / Equipamento']
                novo_valor = row['Valor Atual (R$)']
                antigo_valor = st.session_state.precos_banco.get(item, 0.0)
                if novo_valor != antigo_valor:
                    novo_hist = {"Data/Hora": data_hora_agora, "Item Alterado": item, "Valor Antigo": f"R$ {antigo_valor:.2f}", "Novo Valor": f"R$ {novo_valor:.2f}"}
                    st.session_state.historico_precos.append(novo_hist)
                    novos_historicos.append(novo_hist)
                    st.session_state.precos_banco[item] = novo_valor
                    alterou_algo = True
                    
            if alterou_algo:
                try:
                    sh = conectar_google_sheets()
                    try: ws_precos = sh.worksheet("Precos")
                    except:
                        ws_precos = sh.add_worksheet(title="Precos", rows="100", cols="2")
                        ws_precos.append_row(["Item", "Valor"])
                    ws_precos.clear()
                    ws_precos.append_rows([["Item", "Valor"]] + [[k, v] for k, v in st.session_state.precos_banco.items()])
                    
                    try: ws_hist = sh.worksheet("Historico_Precos")
                    except:
                        ws_hist = sh.add_worksheet(title="Historico_Precos", rows="1000", cols="4")
                        ws_hist.append_row(["Data/Hora", "Item Alterado", "Valor Antigo", "Novo Valor"])
                    
                    if novos_historicos: 
                        ws_hist.append_rows([[h["Data/Hora"], h["Item Alterado"], h["Valor Antigo"], h["Novo Valor"]] for h in novos_historicos])
                    
                    st.session_state.data_precos_atualizada = data_hora_agora
                    st.cache_data.clear()
                    
                    st.toast("✅ Base de preços atualizada com sucesso!", icon="💾")
                    st.rerun()
                except Exception as e: 
                    st.error(f"Erro ao salvar: {e}")
            else:
                st.info("Nenhuma alteração detectada para salvar.")

    with aba_resumo:
        st.header("Consolidação Financeira do Orçamento")
        linhas_inst_campo = []
        linhas_hardware = []
        linhas_software = []
        linhas_pontos = []
        linhas_servicos = []
        
        softwares_incluidos = {}
        
        total_ai_schneider = total_ao_schneider = total_di_schneider = total_do_schneider = 0
        total_ai_siemens = total_ao_siemens = total_di_siemens = total_do_siemens = 0
        total_io_mercato = 0
        
        custo_base_schneider = 0.0
        custo_base_siemens = 0.0
        custo_base_mercato = 0.0

        descritivo_linhas_excel = ["DESCRIÇÃO TÉCNICA DO ESCOPO CONTEMPLADO:\n"]
        descritivo_comercial_linhas = []

        for p in st.session_state.paineis_auto:
            arquitetura_atual = p.get('arquitetura', 'SpaceLogic (Schneider)')
            is_siemens_1200 = (arquitetura_atual == 'S7-1200 (Siemens)')
            is_siemens_1500 = (arquitetura_atual == 'S7-1500 (Siemens)')
            is_siemens = is_siemens_1200 or is_siemens_1500
            is_mercato = ('Mercato' in arquitetura_atual)
            is_schneider = ('Schneider' in arquitetura_atual)
            
            tem_sobra_20 = (p.get('sobra_20', 'Não') == 'Sim')
            tipo_cfr_painel = p.get('tipo_cfr', 'Não Aplicável')
            calibracao_ativa = (p.get('calibracao', 'Não') == 'Sim')
            
            raw_ai_painel = raw_ao_painel = raw_di_painel = raw_do_painel = 0
            qtd_equipamentos_painel = 0
            total_pontos_calibracao = 0
            
            lista_equip_nomes = []
            tem_resistencia = False
            tem_filtro_pdt = False
            tem_filtro_psh = False
            
            lista_instrumentos_nomes = set()
            lista_instrumentos_detalhados = []
            controladores_desc_lista = []
            
            for g_idx, g in enumerate(p['grupos_equipamentos']):
                mult = g.get('multiplicador', 1)
                qtd_equipamentos_painel += mult
                
                lista_tags = [t for t in g.get('tags_lista', []) if t.strip() != ""]
                str_tags_todas = f" [TAGs: {', '.join(lista_tags)}]" if len(lista_tags) > 0 else ""
                
                nome_limpo_grupo = g['nome_grupo'].replace("Equipamento Novo", "").replace("Equipamento Customizado", "").strip()
                if not nome_limpo_grupo:
                    lista_equip_nomes.append(f"{mult}x Equipamento{str_tags_todas}")
                    nome_equip = f"Equipamento"
                else:
                    lista_equip_nomes.append(f"{mult}x {nome_limpo_grupo}{str_tags_todas}")
                    nome_equip = f"{nome_limpo_grupo}"
                
                raw_ai_g_single = raw_ao_g_single = raw_di_g_single = raw_do_g_single = 0
                
                is_compressor_sys = "COMPRESSOR" in g['nome_grupo'].upper() or "DIRETA" in g['nome_grupo'].upper() or "DX" in g['nome_grupo'].upper()
                
                # REGRA SÊNIOR DE MOTORES - APLICADA AQUI TAMBÉM (BLINDAGEM DUPLA)
                tem_motor = any(mot in g['nome_grupo'].upper() for mot in ["UTA", "EXAUST", "VENT", "FANCOIL", "SPLIT", "BOMBA", "SPLITÃO"])
                auto_mon_default = False if tem_motor else True
                if "SALA" in g['nome_grupo'].upper() or "MONITORAMENTO" in g['nome_grupo'].upper():
                    auto_mon_default = True
                    
                is_monitoramento = st.session_state.get(f"chk_mon_{p['id']}_{g_idx}", auto_mon_default)
                
                for inst, qtd in g['instrumentos'].items():
                    if qtd > 0:
                        qtd_final = qtd * mult
                        item_nome_real = inst
                        
                        # CÁLCULO DE CALIBRAÇÃO (Apenas Analógicos)
                        if calibracao_ativa:
                            inst_up = inst.upper()
                            if "(TT/MT)" in inst_up or "(TIT/MIT)" in inst_up:
                                total_pontos_calibracao += (2 * qtd_final)
                            elif "(PDT)" in inst_up or "(PDIT)" in inst_up or "(TT)" in inst_up or "(PIT)" in inst_up or "(FIT)" in inst_up or "(TIT)" in inst_up:
                                total_pontos_calibracao += (1 * qtd_final)

                        if is_mercato:
                            if "Transmissor de temperatura para duto (TT)" in inst: item_nome_real = "Mercato - Sensor de Temperatura NTC (Duto)"
                            elif "Transmissor de temperatura Ambiente (TT)" in inst: item_nome_real = "Mercato - Sensor de Temperatura NTC (Ambiente)"
                            elif "Transmissor de temperatura ambiente com display" in inst: item_nome_real = "Mercato - Sensor de Temperatura NTC com Display (Ambiente)"
                        elif is_schneider:
                            if "Transmissor de temperatura para duto (TT)" in inst: item_nome_real = "Schneider - Sensor de Temperatura NTC (Duto)"
                            elif "Transmissor de temperatura Ambiente (TT)" in inst: item_nome_real = "Schneider - Sensor de Temperatura NTC (Ambiente)"
                        
                        nome_curto_inst = item_nome_real.replace("Mercato - ", "").replace("Schneider - ", "").replace("Siemens - ", "")
                        nome_curto_inst = re.sub(r'\s*\([A-Z/]+\)$', '', nome_curto_inst)
                        lista_instrumentos_nomes.add(nome_curto_inst)
                        
                        if "resistência" in inst.lower() or "raq" in inst.lower() or "tsh" in inst.lower():
                            tem_resistencia = True
                        
                        if "filtro" in inst.lower():
                            if "pdit" in inst.lower() or "pdt" in inst.lower(): tem_filtro_pdt = True
                            if "psh" in inst.lower() or "pressostato" in inst.lower(): tem_filtro_psh = True

                        func_inst = "Medição Genérica"
                        if "pressão dif. para ar" in inst.lower(): func_inst = "Medição da Vazão de Ar"
                        elif "temperatura e umidade" in inst.lower(): func_inst = "Medição de Temperatura e Umidade"
                        elif "temperatura" in inst.lower(): func_inst = "Medição de Temperatura"
                        elif "on/off" in inst.lower() or "proporcional" in inst.lower(): func_inst = "Controle de Válvulas"
                        elif "compressor" in inst.lower(): func_inst = "Status de Operação do Compressor"
                        elif "termostato" in inst.lower(): func_inst = "Segurança de Sobreaquecimento"
                        elif "resistência" in inst.lower() and "pressostato" not in inst.lower(): func_inst = "Aquecimento"
                        elif "pressostato diferencial para ar" in inst.lower() and "resistência" in inst.lower(): func_inst = "Segurança da Resistência por Fluxo de Ar"
                        elif "pressostato para monitorar" in inst.lower(): func_inst = "Alarme de Saturação de Filtro"
                        elif "pressão diferencial (monitorar" in inst.lower(): func_inst = "Monitoramento da Saturação do Filtro"
                        elif "ventilador" in inst.lower() or "exaustor" in inst.lower(): func_inst = "Status de Operação do Ventilador/Exaustor"
                        elif "co2" in inst.lower(): func_inst = "Medição da Qualidade do Ar (CO2)"
                        elif "fluxo" in inst.lower() or "fs" in inst.lower() or "cf" in inst.lower(): func_inst = "Confirmação de Fluxo de Água"
                        
                        # TAG ESPECÍFICA PARA A PLANILHA
                        tag_especifica = get_specific_tags(inst, g.get('tags_lista', []), is_compressor_sys)
                        str_tag_ctx = f" [TAGs: {tag_especifica}]" if tag_especifica else ""
                        nome_equip_inst = f"{nome_equip}{str_tag_ctx}"

                        lista_instrumentos_detalhados.append((nome_curto_inst, qtd_final, func_inst))

                        preco_item = st.session_state.precos_banco.get(item_nome_real, st.session_state.precos_banco.get(inst, 0.0))
                        io_vals = REGRA_IO.get(inst, {"AI": 0, "AO": 0, "DI": 0, "DO": 0})
                        
                        raw_ai_painel += qtd_final * io_vals["AI"]
                        raw_ao_painel += qtd_final * io_vals["AO"]
                        raw_di_painel += qtd_final * io_vals["DI"]
                        raw_do_painel += qtd_final * io_vals["DO"]
                        
                        raw_ai_g_single += qtd * io_vals["AI"]
                        raw_ao_g_single += qtd * io_vals["AO"]
                        raw_di_g_single += qtd * io_vals["DI"]
                        raw_do_g_single += qtd * io_vals["DO"]
                        
                        custo_tot_inst = qtd_final * preco_item
                        linhas_inst_campo.append({"Categoria": "Instrumentação de Campo", "Item": f"{item_nome_real} ({nome_equip_inst} - {p['nome']})", "Preço Unit.": preco_item, "Qtd": qtd_final, "Custo Total": custo_tot_inst})
                        linhas_pontos.append({"Painel": p['nome'], "Grupo/Equipamento": nome_equip_inst, "Instrumento": item_nome_real, "Quantidade Total": qtd_final, "Entrada Digital (DI)": qtd_final * io_vals["DI"], "Saída Digital (DO)": qtd_final * io_vals["DO"], "Entrada Analógica (AI)": qtd_final * io_vals["AI"], "Saída Analógica (AO)": qtd_final * io_vals["AO"]})
                        
                        if is_siemens: custo_base_siemens += custo_tot_inst
                        elif is_mercato: custo_base_mercato += custo_tot_inst
                        else: custo_base_schneider += custo_tot_inst
                
                # ADICIONA CHAVE APENAS SE FOR MOTOR E NÃO FOR MONITORAMENTO PASSIVO
                if tem_motor and not is_monitoramento:
                    raw_di_painel += (2 * mult)
                    raw_di_g_single += 2
                    linhas_pontos.append({"Painel": p['nome'], "Grupo/Equipamento": nome_equip, "Instrumento": "Chave Seletora Auto/Manual (Painel Elétrico)", "Quantidade Total": mult, "Entrada Digital (DI)": 2 * mult, "Saída Digital (DO)": 0, "Entrada Analógica (AI)": 0, "Saída Analógica (AO)": 0})

                if is_mercato:
                    reserva_g_ui = math.ceil((raw_ai_g_single + raw_di_g_single) * 0.2) if tem_sobra_20 else 0
                    reserva_g_ao = math.ceil(raw_ao_g_single * 0.2) if tem_sobra_20 else 0
                    reserva_g_do = math.ceil(raw_do_g_single * 0.2) if tem_sobra_20 else 0
                    ui_nec = raw_ai_g_single + raw_di_g_single + reserva_g_ui
                    ao_nec = raw_ao_g_single + reserva_g_ao
                    do_nec = raw_do_g_single + reserva_g_do
                    
                    modelo_mcp = dimensionar_mercato(ui_nec, ao_nec, do_nec, is_compressor_sys)
                    if modelo_mcp:
                        controladores_desc_lista.append(f"{mult}x {modelo_mcp.replace('Mercato - ', '')}")
                        p_hw = st.session_state.precos_banco.get(modelo_mcp, 1650.0)
                        linhas_hardware.append({"Categoria": "Hardware e Painéis", "Item": f"{modelo_mcp} ({g['nome_grupo']} - {p['nome']})", "Preço Unit.": p_hw, "Qtd": mult, "Custo Total": mult * p_hw})
                        custo_base_mercato += (mult * p_hw)

            # INCLUSÃO DO CUSTO DE CALIBRAÇÃO (Se Houver)
            if total_pontos_calibracao > 0:
                pr_calib = st.session_state.precos_banco.get("Serviço de Calibração (Por Ponto Analógico)", 180.0)
                linhas_servicos.append({
                    "Categoria": "Serviços de Lógica", 
                    "Item": f"Calibração de Instrumentos Analógicos ({p['nome']})", 
                    "Preço Unit.": pr_calib, "Qtd": total_pontos_calibracao, "Custo Total": total_pontos_calibracao * pr_calib
                })

            reserva_ai_painel = math.ceil(raw_ai_painel * 0.2) if tem_sobra_20 else 0
            reserva_ao_painel = math.ceil(raw_ao_painel * 0.2) if tem_sobra_20 else 0
            reserva_di_painel = math.ceil(raw_di_painel * 0.2) if tem_sobra_20 else 0
            reserva_do_painel = math.ceil(raw_do_painel * 0.2) if tem_sobra_20 else 0
            
            hw_ai_painel = raw_ai_painel + reserva_ai_painel
            hw_ao_painel = raw_ao_painel + reserva_ao_painel
            hw_di_painel = raw_di_painel + reserva_di_painel
            hw_do_painel = raw_do_painel + reserva_do_painel
            
            tot_io_painel_hw = hw_ai_painel + hw_ao_painel + hw_di_painel + hw_do_painel
            
            if reserva_ai_painel > 0 or reserva_ao_painel > 0 or reserva_di_painel > 0 or reserva_do_painel > 0:
                linhas_pontos.append({"Painel": p['nome'], "Grupo/Equipamento": "Reserva Técnica (20%)", "Instrumento": "Pontos de Sobra Física do Quadro", "Quantidade Total": "-", "Entrada Digital (DI)": reserva_di_painel, "Saída Digital (DO)": reserva_do_painel, "Entrada Analógica (AI)": reserva_ai_painel, "Saída Analógica (AO)": reserva_ao_painel})

            if is_siemens:
                total_ai_siemens += raw_ai_painel; total_ao_siemens += raw_ao_painel; total_di_siemens += raw_di_painel; total_do_siemens += raw_do_painel
            elif is_mercato:
                total_io_mercato += (raw_ai_painel + raw_ao_painel + raw_di_painel + raw_do_painel)
            else:
                total_ai_schneider += raw_ai_painel; total_ao_schneider += raw_ao_painel; total_di_schneider += raw_di_painel; total_do_schneider += raw_do_painel
            
            if tot_io_painel_hw > 0:
                if is_mercato:
                    nome_caixa, preco_caixa = calcular_painel_fisico(qtd_equipamentos_painel)
                    linhas_hardware.append({"Categoria": "Hardware e Painéis", "Item": f"Estrutura Física: {nome_caixa} ({p['nome']})", "Preço Unit.": preco_caixa, "Qtd": 1, "Custo Total": preco_caixa})
                    custo_base_mercato += preco_caixa
                        
                else:
                    nome_caixa, preco_caixa = calcular_painel_fisico(tot_io_painel_hw/15)
                    linhas_hardware.append({"Categoria": "Hardware e Painéis", "Item": f"Estrutura Física: {nome_caixa} ({p['nome']})", "Preço Unit.": preco_caixa, "Qtd": 1, "Custo Total": preco_caixa})
                    
                    if is_siemens:
                        custo_base_siemens += preco_caixa
                        if is_siemens_1200: hw_s = dimensionar_siemens_1200(hw_ai_painel, hw_ao_painel, hw_di_painel, hw_do_painel)
                        else: hw_s = dimensionar_siemens_1500(hw_ai_painel, hw_ao_painel, hw_di_painel, hw_do_painel)
                        for i_hw, q_hw in hw_s.items():
                            if q_hw > 0:
                                controladores_desc_lista.append(f"{q_hw}x {i_hw.replace('Siemens - ', '')}")
                                pr = st.session_state.precos_banco.get(i_hw, 0.0)
                                linhas_hardware.append({"Categoria": "Hardware e Painéis", "Item": f"{i_hw} ({p['nome']})", "Preço Unit.": pr, "Qtd": q_hw, "Custo Total": q_hw * pr})
                                custo_base_siemens += (q_hw * pr)
                    else:
                        custo_base_schneider += preco_caixa
                        c36, c24, c18, c15 = dimensionar_controladores(tot_io_painel_hw)
                        if c36 > 0: 
                            controladores_desc_lista.append(f"{c36}x Controlador MP-C-36A")
                            linhas_hardware.append({"Categoria": "Hardware e Painéis", "Item": f"Controlador MP-C-36A ({p['nome']})", "Preço Unit.": st.session_state.precos_banco.get("MP-C-36A", 9459.0), "Qtd": c36, "Custo Total": c36 * st.session_state.precos_banco.get("MP-C-36A", 9459.0)})
                            custo_base_schneider += (c36 * st.session_state.precos_banco.get("MP-C-36A", 9459.0))
                        if c24 > 0: 
                            controladores_desc_lista.append(f"{c24}x Controlador MP-C-24A")
                            linhas_hardware.append({"Categoria": "Hardware e Painéis", "Item": f"Controlador MP-C-24A ({p['nome']})", "Preço Unit.": st.session_state.precos_banco.get("MP-C-24A", 7290.0), "Qtd": c24, "Custo Total": c24 * st.session_state.precos_banco.get("MP-C-24A", 7290.0)})
                            custo_base_schneider += (c24 * st.session_state.precos_banco.get("MP-C-24A", 7290.0))
                        if c18 > 0: 
                            controladores_desc_lista.append(f"{c18}x Controlador MP-C-18A")
                            linhas_hardware.append({"Categoria": "Hardware e Painéis", "Item": f"Controlador MP-C-18A ({p['nome']})", "Preço Unit.": st.session_state.precos_banco.get("MP-C-18A", 5185.0), "Qtd": c18, "Custo Total": c18 * st.session_state.precos_banco.get("MP-C-18A", 5185.0)})
                            custo_base_schneider += (c18 * st.session_state.precos_banco.get("MP-C-18A", 5185.0))
                        if c15 > 0: 
                            controladores_desc_lista.append(f"{c15}x Controlador MP-C-15A")
                            linhas_hardware.append({"Categoria": "Hardware e Painéis", "Item": f"Controlador MP-C-15A ({p['nome']})", "Preço Unit.": st.session_state.precos_banco.get("MP-C-15A", 4649.0), "Qtd": c15, "Custo Total": c15 * st.session_state.precos_banco.get("MP-C-15A", 4649.0)})
                            custo_base_schneider += (c15 * st.session_state.precos_banco.get("MP-C-15A", 4649.0))
                
                if p.get('ihm') and "Cego" not in p['ihm']:
                    preco_ihm = st.session_state.precos_banco.get(p['ihm'], 0.0)
                    if preco_ihm > 0: 
                        linhas_hardware.append({"Categoria": "Hardware e Painéis", "Item": f"Interface: {p['ihm']} ({p['nome']})", "Preço Unit.": preco_ihm, "Qtd": 1, "Custo Total": preco_ihm})
                        if is_mercato: custo_base_mercato += preco_ihm
                        elif is_siemens: custo_base_siemens += preco_ihm
                        else: custo_base_schneider += preco_ihm

                s_type = p.get('supervisorio', "Sem Supervisório")
                if s_type != "Sem Supervisório":
                    if is_schneider:
                        pr_as = st.session_state.precos_banco.get("Schneider - Servidor de Automação (SpaceLogic AS-P/AS-B)", 9500.0)
                        controladores_desc_lista.append("1x Servidor de Automação AS-P/AS-B")
                        linhas_hardware.append({"Categoria": "Hardware e Painéis", "Item": f"Servidor de Automação AS-P/AS-B ({p['nome']})", "Preço Unit.": pr_as, "Qtd": 1, "Custo Total": pr_as})
                        custo_base_schneider += pr_as
                        
                    chave_soft = (s_type, tipo_cfr_painel)
                    if chave_soft not in softwares_incluidos: softwares_incluidos[chave_soft] = 0
                    softwares_incluidos[chave_soft] += (raw_ai_painel + raw_ao_painel + raw_di_painel + raw_do_painel)

            # BLINDAGEM DO NAMERROR (Geração de Texto)
            ihm_desc = f"com IHM instalada na porta, com display de {p['ihm'].replace('Mercato - ', '').replace('IHM Padrão ', '').replace('IHM Premium ', '').replace('IHM Básica ', '')}" if "Cego" not in p['ihm'] else "sem interface IHM instalada"
            
            if "Sem" in str(p.get('supervisorio', 'Sem')): 
                sup_desc = "Stand-alone (sem supervisório)"
            elif "EBO" in str(p.get('supervisorio', '')): 
                sup_desc = "integrado ao sistema supervisório EBO"
            else: 
                sup_desc = "integrado ao sistema supervisório"
            
            texto_filtro = ""
            bullet_filtros = ""
            if tem_filtro_pdt:
                texto_filtro = "monitoramento contínuo da saturação dos filtros"
                bullet_filtros = "• Monitoramento contínuo e alarmes de saturação de filtros.\n"
            elif tem_filtro_psh:
                texto_filtro = "monitoramento para alarme devido à saturação dos filtros"
                bullet_filtros = "• Monitoramento para alarme devido à saturação de filtros.\n"
                
            res_desc_intro = "controle da resistência elétrica de aquecimento" if tem_resistencia else ""
            
            componentes_intro = []
            if texto_filtro: componentes_intro.append(texto_filtro)
            if res_desc_intro: componentes_intro.append(res_desc_intro)
            texto_intro_extra = ", incluindo " + " e ".join(componentes_intro) if componentes_intro else ""

            # Garantindo a existência das variáveis de nome antes da formatação
            str_eqs_nome = ", ".join(lista_equip_nomes) if lista_equip_nomes else "Equipamentos do Quadro"
            str_ctrls_desc = ", ".join(controladores_desc_lista) if controladores_desc_lista else "Controladores"
            
            txt_p = (
                f"Sistema de automação dedicado para controle de {str_eqs_nome}{texto_intro_extra}.\n\n"
                f"O sistema contempla quadro de automação [TAG: {p['nome']}] {ihm_desc}, "
                f"baseado na tecnologia {arquitetura_atual.replace(' - Linha mais econômica', '')} ({str_ctrls_desc}), operando no modo {sup_desc}, "
                f"permitindo a visualização em tempo real e o controle dos seguintes parâmetros operacionais gerais:\n\n"
                f"• Status de operação dos equipamentos.\n"
                f"{bullet_filtros}"
            )
            
            if tem_resistencia: 
                txt_p += "• Status e acionamento da resistência elétrica.\n"
                
            str_insts = ", ".join(list(lista_instrumentos_nomes)) if lista_instrumentos_nomes else "Instrumentos diversos"
            txt_p += (
                f"• Leitura de instrumentos de campo diversos ({str_insts}).\n"
                f"• Condições gerais de funcionamento.\n\n"
                f"A solução proporciona maior confiabilidade operacional, facilidade de manutenção e gestão eficiente dos ativos térmicos e de controle de ar."
            )
            
            if calibracao_ativa:
                txt_p += "\n\nDestaca-se que somente os instrumentos analógicos de medição passarão por processo de calibração aferida."
            
            if tipo_cfr_painel == 'CFR21 Part 11 - Qualificável':
                txt_p += "\n\nO sistema será fornecido de forma Qualificável conforme normas CFR 21 Part 11, atendendo a todos os requisitos técnicos e de software para que o cliente realize a qualificação posterior."
            elif tipo_cfr_painel == 'CFR21 Part 11 - Qualificado':
                txt_p += "\n\nO sistema será integralmente Qualificado conforme normas CFR 21 Part 11, com a entrega de todos os protocolos pela equipe especializada da SIARCON."

            descritivo_linhas_excel.append(txt_p)
            
            txt_com = f"**Sistema completo de automação [TAG: {p['nome']}]**, construído com arquitetura de controladores **{arquitetura_atual.replace(' - Linha mais econômica', '')}** ({str_ctrls_desc}). O sistema operará de forma **{sup_desc}**, {ihm_desc}.\n\n"
            
            if tipo_cfr_painel == 'CFR21 Part 11 - Qualificável':
                txt_com += "O sistema fornecido possuirá as licenças e os parâmetros necessários para ser totalmente **Qualificável (CFR 21 Part 11)**. A SIARCON garantirá todos os requisitos técnicos de software, deixando o ambiente pronto para que a qualificação final seja realizada por empresa à escolha do cliente.\n\n"
            elif tipo_cfr_painel == 'CFR21 Part 11 - Qualificado':
                txt_com += "O sistema contemplado será integralmente **Qualificado (CFR 21 Part 11)**. A SIARCON executará e entregará toda a documentação comprobatória e a execução dos protocolos e validações pertinentes, garantindo a certificação total do ambiente de supervisão.\n\n"
            
            txt_com += "O quadro de automação será responsável pela aquisição de dados e controle da seguinte instrumentação de campo:\n\n"
            
            for desc_inst, qt_inst, func_inst in lista_instrumentos_detalhados:
                txt_com += f"• **{qt_inst}x {func_inst}:** {desc_inst}\n"
                
            txt_com += "\n**Lógica de Operação do Sistema:**\n"
            txt_com += "O sistema realizará o controle da vazão de ar de forma constante, efetuando os ajustes necessários no inversor para que a vazão volumétrica seja mantida, independentemente do nível de saturação dos filtros no tempo. O controlador modulará proporcionalmente a válvula da serpentina (ou estágios do compressor) para atingir os parâmetros exatos de setpoint térmico demandados pelo ambiente."
            
            if tem_resistencia:
                txt_com += " A resistência elétrica de aquecimento será acionada por malha de controle PID dedicada, permitindo ajuste fino de temperatura e desumidificação, possuindo intertravamento de segurança via termostato mecânico de proteção e confirmação de fluxo de ar."
                
            descritivo_comercial_linhas.append(txt_com)

        texto_descritivo_final = "\n\n----------------------------------------------------\n\n".join(descritivo_linhas_excel)

        for (s_name, t_cfr), pts_total in softwares_incluidos.items():
            b_k, p_k = "", ""
            if "SEM certificação" in s_name: 
                b_k, p_k = "Licença Supervisório - SEM CFR-21 (Base)", "Licença Supervisório - SEM CFR-21 (Por Ponto I/O)"
            elif "COM certificação" in s_name: 
                b_k, p_k = "Licença Supervisório - COM CFR-21 (Base)", "Licença Supervisório - COM CFR-21 (Por Ponto I/O)"
            else: 
                b_k, p_k = "Licença Supervisório - Schneider EBO (Base)", "Licença Supervisório - Schneider EBO (Por Ponto I/O)"
            
            p_base = st.session_state.precos_banco.get(b_k, 23000.0)
            p_pto = st.session_state.precos_banco.get(p_k, 100.0)
            linhas_software.append({"Categoria": "Software de Supervisão", "Item": f"Licença Base: {s_name}", "Preço Unit.": p_base, "Qtd": 1, "Custo Total": p_base})
            
            if p_pto > 0 and pts_total > 0:
                linhas_software.append({"Categoria": "Software de Supervisão", "Item": f"Pontos Licenciados no Software ({pts_total} canais ativos)", "Preço Unit.": p_pto, "Qtd": pts_total, "Custo Total": pts_total * p_pto})

            if "COM certificação" in s_name:
                custo_cfr_unit = 0.0
                if t_cfr == "CFR21 Part 11 - Qualificável":
                    if pts_total <= 100: custo_cfr_unit = st.session_state.precos_banco.get("CFR21 Qualificável - Até 100 pts", 70.0)
                    elif pts_total <= 250: custo_cfr_unit = st.session_state.precos_banco.get("CFR21 Qualificável - 101 a 250 pts", 50.0)
                    else: custo_cfr_unit = st.session_state.precos_banco.get("CFR21 Qualificável - Acima de 250 pts", 30.0)
                    
                    linhas_servicos.append({"Categoria": "Serviços de Lógica", "Item": f"Preparação do Sistema Qualificável (CFR21) - Por Ponto", "Preço Unit.": custo_cfr_unit, "Qtd": pts_total, "Custo Total": pts_total * custo_cfr_unit})
                    
                elif t_cfr == "CFR21 Part 11 - Qualificado":
                    if pts_total <= 30: custo_cfr_unit = st.session_state.precos_banco.get("CFR21 Qualificado - Até 30 pts", 400.0)
                    elif pts_total <= 60: custo_cfr_unit = st.session_state.precos_banco.get("CFR21 Qualificado - 31 a 60 pts", 350.0)
                    elif pts_total <= 99: custo_cfr_unit = st.session_state.precos_banco.get("CFR21 Qualificado - 61 a 99 pts", 320.0)
                    elif pts_total <= 150: custo_cfr_unit = st.session_state.precos_banco.get("CFR21 Qualificado - 100 a 150 pts", 290.0)
                    elif pts_total <= 200: custo_cfr_unit = st.session_state.precos_banco.get("CFR21 Qualificado - 151 a 200 pts", 250.0)
                    elif pts_total <= 250: custo_cfr_unit = st.session_state.precos_banco.get("CFR21 Qualificado - 201 a 250 pts", 220.0)
                    else: custo_cfr_unit = st.session_state.precos_banco.get("CFR21 Qualificado - Acima de 250 pts", 200.0)

                    linhas_servicos.append({"Categoria": "Serviços de Lógica", "Item": f"Execução de Qualificação Integral (CFR21) - Por Ponto", "Preço Unit.": custo_cfr_unit, "Qtd": pts_total, "Custo Total": pts_total * custo_cfr_unit})

        for av in st.session_state.orcamento:
            linhas_inst_campo.append({"Categoria": "Instrumentação de Campo", "Item": av['Item'], "Preço Unit.": av['Custo_Total']/av['Quantidade'] if av['Quantidade']>0 else 0, "Qtd": av['Quantidade'], "Custo Total": av['Custo_Total']})

        df_inst = pd.DataFrame(linhas_inst_campo)
        df_hw = pd.DataFrame(linhas_hardware)
        df_sw = pd.DataFrame(linhas_software)
        
        subtotal_inst = df_inst['Custo Total'].sum() if not df_inst.empty else 0
        subtotal_hw = df_hw['Custo Total'].sum() if not df_hw.empty else 0
        subtotal_sw = df_sw['Custo Total'].sum() if not df_sw.empty else 0
        
        if (total_ai_schneider + total_ao_schneider) > 0:
            pr_ai_sch = st.session_state.precos_banco.get("Custo AI/AO", 565.0)
            linhas_servicos.append({"Categoria": "Serviços de Lógica", "Item": "Serviços de lógica: Pontos Analógicos", "Preço Unit.": pr_ai_sch, "Qtd": (total_ai_schneider + total_ao_schneider), "Custo Total": (total_ai_schneider + total_ao_schneider) * pr_ai_sch})
        if (total_di_schneider + total_do_schneider) > 0:
            pr_di_sch = st.session_state.precos_banco.get("Custo DI/DO", 120.0)
            linhas_servicos.append({"Categoria": "Serviços de Lógica", "Item": "Serviços de lógica: Pontos Digitais", "Preço Unit.": pr_di_sch, "Qtd": (total_di_schneider + total_do_schneider), "Custo Total": (total_di_schneider + total_do_schneider) * pr_di_sch})
        if (total_ai_siemens + total_ao_siemens) > 0:
            pr_ai_siem = st.session_state.precos_banco.get("Siemens - Serviço Custo AI/AO", 750.0)
            linhas_servicos.append({"Categoria": "Serviços de Lógica", "Item": "Serviços de lógica (Siemens): Pontos Analógicos", "Preço Unit.": pr_ai_siem, "Qtd": (total_ai_siemens + total_ao_siemens), "Custo Total": (total_ai_siemens + total_ao_siemens) * pr_ai_siem})
        if (total_di_siemens + total_do_siemens) > 0:
            pr_di_siem = st.session_state.precos_banco.get("Siemens - Serviço Custo DI/DO", 180.0)
            linhas_servicos.append({"Categoria": "Serviços de Lógica", "Item": "Serviços de lógica (Siemens): Pontos Digitais", "Preço Unit.": pr_di_siem, "Qtd": (total_di_siemens + total_do_siemens), "Custo Total": (total_di_siemens + total_do_siemens) * pr_di_siem})
        if total_io_mercato > 0:
            pr_serv_merc = st.session_state.precos_banco.get("Mercato - Serviço Parametrização por Ponto", 80.0)
            linhas_servicos.append({"Categoria": "Serviços de Lógica", "Item": "Parametrização de Pontos (Mercato)", "Preço Unit.": pr_serv_merc, "Qtd": total_io_mercato, "Custo Total": total_io_mercato * pr_serv_merc})

        mao_de_obra_extra = (custo_base_schneider * 0.25) + (custo_base_siemens * 0.35) + (custo_base_mercato * 0.10)
        if mao_de_obra_extra > 0:
            linhas_servicos.append({"Categoria": "Serviços de Lógica", "Item": "Demais programações e desenvolvimentos", "Preço Unit.": mao_de_obra_extra, "Qtd": 1, "Custo Total": mao_de_obra_extra})
            
        df_serv = pd.DataFrame(linhas_servicos)
        subtotal_serv = df_serv['Custo Total'].sum() if not df_serv.empty else 0
        total_geral = subtotal_inst + subtotal_hw + subtotal_sw + subtotal_serv
        
        if total_geral > 0:
            st.markdown("### Resumo Estruturado")
            c1, c2, c3 = st.columns(3)
            c1.info(f"**Subtotal Instrumentação:**\nR$ {subtotal_inst:,.2f}")
            c2.warning(f"**Subtotal Hardware:**\nR$ {subtotal_hw:,.2f}")
            c3.success(f"**CUSTO TOTAL ESTIMADO:**\nR$ {total_geral:,.2f}")

            # PROTEÇÃO CONTRA A TUPLA VAZIA (O erro dos "None" na tela corrigido)
            expl = []
            if not df_inst.empty:
                expl.append(pd.DataFrame([{"Categoria": "", "Item": "INSTRUMENTAÇÃO DE CAMPO", "Preço Unit.": "-", "Qtd": "-", "Custo Total": "-"}]))
                df_inst_grouped = df_inst.groupby(['Categoria', 'Item'], as_index=False).agg({'Preço Unit.': 'first', 'Qtd': 'sum', 'Custo Total': 'sum'})
                expl.append(df_inst_grouped)
                expl.append(pd.DataFrame([{"Categoria": "SUBTOTAL", "Item": "INSTRUMENTAÇÃO DE CAMPO", "Preço Unit.": "-", "Qtd": "-", "Custo Total": subtotal_inst}]))
            
            if not df_hw.empty:
                expl.append(pd.DataFrame([{"Categoria": "", "Item": "HARDWARE E PAINÉIS", "Preço Unit.": "-", "Qtd": "-", "Custo Total": "-"}]))
                df_hw_grouped = df_hw.groupby(['Categoria', 'Item'], as_index=False).agg({'Preço Unit.': 'first', 'Qtd': 'sum', 'Custo Total': 'sum'})
                expl.append(df_hw_grouped)
                expl.append(pd.DataFrame([{"Categoria": "SUBTOTAL", "Item": "HARDWARE E PAINÉIS", "Preço Unit.": "-", "Qtd": "-", "Custo Total": subtotal_hw}]))
            
            if not df_sw.empty:
                expl.append(pd.DataFrame([{"Categoria": "", "Item": "SOFTWARE", "Preço Unit.": "-", "Qtd": "-", "Custo Total": "-"}]))
                df_sw_grouped = df_sw.groupby(['Categoria', 'Item'], as_index=False).agg({'Preço Unit.': 'first', 'Qtd': 'sum', 'Custo Total': 'sum'})
                expl.append(df_sw_grouped)
                expl.append(pd.DataFrame([{"Categoria": "SUBTOTAL", "Item": "SOFTWARE", "Preço Unit.": "-", "Qtd": "-", "Custo Total": subtotal_sw}]))
            
            if not df_serv.empty:
                expl.append(pd.DataFrame([{"Categoria": "", "Item": "SERVIÇOS E LÓGICA", "Preço Unit.": "-", "Qtd": "-", "Custo Total": "-"}]))
                expl.append(df_serv)
                expl.append(pd.DataFrame([{"Categoria": "SUBTOTAL", "Item": "SERVIÇOS E LÓGICA", "Preço Unit.": "-", "Qtd": "-", "Custo Total": subtotal_serv}]))
            
            expl.append(pd.DataFrame([{"Categoria": "TOTAL GERAL", "Item": "ORÇAMENTO COMPLETO", "Preço Unit.": "-", "Qtd": "-", "Custo Total": total_geral}]))
            df_exportacao = pd.concat(expl, ignore_index=True)
            
            def format_currency(val):
                try: 
                    return f"R$ {float(val):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                except: 
                    return val
                    
            df_display = df_exportacao.copy()
            df_display['Preço Unit.'] = df_display['Preço Unit.'].apply(format_currency)
            df_display['Custo Total'] = df_display['Custo Total'].apply(format_currency)
            st.dataframe(df_display, use_container_width=True)
            
            with st.expander("📄 Gerar Descritivo Detalhado para Proposta Comercial", expanded=False):
                st.markdown("<div style='background-color:#E3F2FD; padding:20px; border-radius:10px;'>", unsafe_allow_html=True)
                for t_com in descritivo_comercial_linhas:
                    st.markdown(t_com)
                    st.markdown("<hr style='border:1px solid #B0BEC5'>", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)
            
            df_pontos = pd.DataFrame(linhas_pontos)
            if not df_pontos.empty:
                total_qtd = pd.to_numeric(df_pontos['Quantidade Total'], errors='coerce').fillna(0).sum()
                linha_total = pd.DataFrame([{"Painel": "TOTAL GERAL", "Grupo/Equipamento": "-", "Instrumento": "-", "Quantidade Total": total_qtd, "Entrada Digital (DI)": df_pontos['Entrada Digital (DI)'].sum(), "Saída Digital (DO)": df_pontos['Saída Digital (DO)'].sum(), "Entrada Analógica (AI)": df_pontos['Entrada Analógica (AI)'].sum(), "Saída Analógica (AO)": df_pontos['Saída Analógica (AO)'].sum()}])
                df_pontos = pd.concat([df_pontos, linha_total], ignore_index=True)

            buffer = io.BytesIO()
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Detalhamento Financeiro"
            ws1.views.sheetView[0].showGridLines = True
            
            ws1.row_dimensions[1].height = 35
            ws1.row_dimensions[2].height = 25
            ws1.row_dimensions[3].height = 25
            ws1.row_dimensions[4].height = 25
            
            nome_projeto_header = st.session_state.nome_projeto_orcamento if st.session_state.nome_projeto_orcamento else "PROJETO NÃO NOMEADO"
            
            ws1.merge_cells("C1:E1")
            ws1.cell(row=1, column=3, value="DESCRIÇÃO TÉCNICA E ORÇAMENTÁRIA DE SISTEMAS DE AUTOMAÇÃO").font = Font(name="Arial", size=12, bold=True, color="1C8590")
            ws1.cell(row=1, column=3).alignment = Alignment(horizontal="center", vertical="center")
            
            ws1.merge_cells("C2:E2")
            ws1.cell(row=2, column=3, value=f"PROJETO: {nome_projeto_header.upper()}").font = Font(name="Arial", size=10, bold=True, color="333333")
            
            ws1.merge_cells("C3:E3")
            ws1.cell(row=3, column=3, value=f"DATA/HORA EMISSÃO: {datetime.now(fuso_br).strftime('%d/%m/%Y %H:%M:%S')}").font = Font(name="Arial", size=10, color="555555")
            
            ws1.merge_cells("C4:E4")
            ws1.cell(row=4, column=3, value=f"RESPONSÁVEL TÉCNICO: {st.session_state.nome_exibicao.upper()}").font = Font(name="Arial", size=10, color="555555")
            
            for r in range(2, 5): 
                ws1.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
            
            start_row = 6
            if ARQUIVO_LOGO:
                try:
                    from openpyxl.drawing.image import Image as OpenpyxlImage
                    img = OpenpyxlImage(ARQUIVO_LOGO)
                    img.width = 180
                    img.height = 50
                    ws1.add_image(img, "A1")
                except: pass
                
            fill_header = PatternFill(start_color="1C8590", end_color="1C8590", fill_type="solid")
            font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            border_thin = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
            
            for r_idx, row in enumerate(dataframe_to_rows(df_exportacao, index=False, header=True), start=start_row):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws1.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border_thin
                
                if r_idx == start_row:
                    for c in range(1, 6):
                        ws1.cell(row=r_idx, column=c).fill = fill_header
                        ws1.cell(row=r_idx, column=c).font = font_header
                        ws1.cell(row=r_idx, column=c).alignment = Alignment(horizontal="center", vertical="center")
                else:
                    is_subtotal = "SUBTOTAL" in str(ws1.cell(row=r_idx, column=1).value) or "TOTAL GERAL" in str(ws1.cell(row=r_idx, column=1).value)
                    is_title = (str(ws1.cell(row=r_idx, column=1).value) == "" and str(ws1.cell(row=r_idx, column=2).value) in ["INSTRUMENTAÇÃO DE CAMPO", "HARDWARE E PAINÉIS", "SOFTWARE", "SERVIÇOS E LÓGICA"])
                    
                    for c_idx in range(1, 6):
                        c_cell = ws1.cell(row=r_idx, column=c_idx)
                        c_cell.font = Font(name="Arial", size=10, bold=(is_subtotal or is_title))
                        
                        if is_title:
                            c_cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                            if c_idx == 2:
                                c_cell.alignment = Alignment(horizontal="center", vertical="center")
                            elif c_idx > 2:
                                c_cell.value = ""
                        elif is_subtotal:
                            c_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                            if c_idx in [3, 5]:
                                if str(c_cell.value).strip() != "-":
                                    try: 
                                        c_cell.value = float(c_cell.value)
                                        c_cell.number_format = '"R$" #,##0.00'
                                    except: pass
                                c_cell.alignment = Alignment(horizontal="right")
                        else:
                            if c_idx in [3, 5]:
                                if str(c_cell.value).strip() != "-":
                                    try: 
                                        c_cell.value = float(c_cell.value)
                                        c_cell.number_format = '"R$" #,##0.00'
                                    except: pass
                                c_cell.alignment = Alignment(horizontal="right")
                            elif c_idx == 4:
                                c_cell.alignment = Alignment(horizontal="center")
                                
                    if is_title:
                        ws1.merge_cells(start_row=r_idx, start_column=2, end_row=r_idx, end_column=5)
            
            end_row_table = start_row + len(df_exportacao) + 2
            num_linhas_texto = len(texto_descritivo_final.split('\n'))
            tamanho_caixa = max(10, num_linhas_texto + 2) 
            
            ws1.merge_cells(start_row=end_row_table, start_column=1, end_row=end_row_table+tamanho_caixa, end_column=5)
            cell_desc = ws1.cell(row=end_row_table, column=1, value=texto_descritivo_final)
            cell_desc.font = Font(name="Arial", size=10, italic=False, color="333333")
            cell_desc.alignment = Alignment(vertical="top", wrap_text=True)
            cell_desc.fill = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
            
            for r in range(end_row_table, end_row_table+tamanho_caixa+1):
                for c in range(1, 6): ws1.cell(row=r, column=c).border = border_thin
            
            for col in ws1.columns:
                max_len = max(len(str(cell.value or '')) for cell in col if cell.row <= start_row + len(df_exportacao))
                ws1.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)

            if not df_pontos.empty:
                ws2 = wb.create_sheet(title="Matriz de IO")
                ws2.views.sheetView[0].showGridLines = True
                for r_idx, row in enumerate(dataframe_to_rows(df_pontos, index=False, header=True), start=1):
                    for c_idx, value in enumerate(row, start=1):
                        cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = border_thin
                        if r_idx == 1: 
                            cell.fill = fill_header
                            cell.font = font_header
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.font = Font(name="Arial", size=10)
                            if c_idx >= 4: cell.alignment = Alignment(horizontal="center")
                for col in ws2.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    ws2.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)

            # ABA ADICIONAL EXCLUSIVA NO EXCEL: LISTA PARA COTAÇÃO SEPARADA POR MARCA
            ws3 = wb.create_sheet(title="Lista para Cotação")
            ws3.views.sheetView[0].showGridLines = True
            
            ws3.cell(row=1, column=1, value="LISTA DE MATERIAIS PARA COMPRAS E COTAÇÃO EXTERNA").font = Font(name="Arial", size=12, bold=True, color="1C8590")
            ws3.cell(row=2, column=1, value=f"ÚLTIMA ATUALIZAÇÃO DA BASE DE PREÇOS: {st.session_state.data_precos_atualizada}").font = Font(name="Arial", size=9, italic=True)
            
            headers_cot = ["Fabricante", "Item / Modelo", "Quantidade Total Necessária", "Unidade"]
            for c_idx, h_text in enumerate(headers_cot, start=1):
                c_cell = ws3.cell(row=4, column=c_idx, value=h_text)
                c_cell.fill = fill_header
                c_cell.font = font_header
                c_cell.alignment = Alignment(horizontal="center")
                c_cell.border = border_thin
            
            it_row = 5
            todos_itens_cotacao = []
            if not df_hw.empty:
                for _, r in df_hw.iterrows(): todos_itens_cotacao.append(r['Item'])
            if not df_inst.empty:
                for _, r in df_inst.iterrows(): todos_itens_cotacao.append(r['Item'])
                
            marcas = {"SIEMENS": [], "SCHNEIDER": [], "MERCATO": [], "OUTROS / GENÉRICOS": []}
            for it in set(todos_itens_cotacao):
                cnt = todos_itens_cotacao.count(it)
                it_upper = it.upper()
                if "SIEMENS" in it_upper: marcas["SIEMENS"].append((it, cnt, "un"))
                elif "SCHNEIDER" in it_upper or "MP-C" in it_upper or "SPACELOGIC" in it_upper: marcas["SCHNEIDER"].append((it, cnt, "un"))
                elif "MERCATO" in it_upper or "MCP-" in it_upper or "MFC" in it_upper or "MDX" in it_upper: marcas["MERCATO"].append((it, cnt, "un"))
                else: marcas["OUTROS / GENÉRICOS"].append((it, cnt, "un"))
            
            for m_name, items_list in marcas.items():
                if items_list:
                    for name_i, q_i, uni_i in items_list:
                        ws3.cell(row=it_row, column=1, value=m_name).alignment = Alignment(horizontal="center")
                        ws3.cell(row=it_row, column=2, value=name_i)
                        ws3.cell(row=it_row, column=3, value=q_i).alignment = Alignment(horizontal="center")
                        ws3.cell(row=it_row, column=4, value=uni_i).alignment = Alignment(horizontal="center")
                        for col_c in range(1, 5): 
                            ws3.cell(row=it_row, column=col_c).border = border_thin
                            ws3.cell(row=it_row, column=col_c).font = Font(name="Arial", size=10)
                        it_row += 1
                        
            for col in ws3.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws3.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)
            
            wb.save(buffer)
            buffer.seek(0)
            
            st.download_button(label="📥 Exportar Orçamento Final para Excel", data=buffer.getvalue(), file_name="orcamento_dimensionado.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.markdown("---")
            
            if st.button("☁️ Salvar Orçamento Final e Gerar Revisão", type="primary", use_container_width=True):
                if not st.session_state.nome_projeto_orcamento: 
                    st.warning("⚠️ Atenção: Preencha o 'Nome do Orçamento / Projeto' antes de salvar.")
                else:
                    try:
                        sh = conectar_google_sheets()
                        try: 
                            ws_h = sh.worksheet("Historico_Orcamentos")
                        except: 
                            ws_h = sh.add_worksheet("Historico_Orcamentos", 1000, 8)
                            ws_h.append_row(["Data/Hora", "Nome do Projeto", "Revisão", "Subtotal Hardware", "Serviços de Lógica", "Custo Total Estimado", "Configuracao_JSON", "Usuário"])
                            
                        todas_linhas_hist = ws_h.get_all_values()
                        contagem_revisoes = sum(1 for r in todas_linhas_hist[1:] if r[1].strip().upper() == st.session_state.nome_projeto_orcamento.strip().upper())
                        revisao_atual = f"R-{contagem_revisoes:02d}"
                        
                        nova_linha = [datetime.now(fuso_br).strftime("%d/%m/%Y %H:%M:%S"), st.session_state.nome_projeto_orcamento, revisao_atual, f"R$ {subtotal_hw:.2f}".replace('.', ','), f"R$ {subtotal_serv:.2f}".replace('.', ','), f"R$ {total_geral:.2f}".replace('.', ','), json.dumps(st.session_state.paineis_auto), st.session_state.usuario_logado]
                        ws_h.append_row(nova_linha)
                        st.cache_data.clear()
                        st.success(f"✅ Sucesso! Orçamento para '{st.session_state.nome_projeto_orcamento}' salvo com a revisão {revisao_atual}!")
                    except Exception as e: 
                        st.error(f"Erro ao salvar no banco: {e}")

# ==============================================================================
# MÓDULO 3: LEVANTAMENTO DE HIDRÁULICA (CAVALETES)
# ==============================================================================
elif st.session_state.menu_selecionado == "💧 Levantamento de Hidráulica":
    import collections
    import re
    import math
    import uuid
    import io
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    st.markdown("""<style>[data-testid="stVerticalBlockBorderWrapper"] { border-radius: 8px; border-left: 4px solid #1C8590 !important; background-color: rgba(28, 133, 144, 0.03); }</style>""", unsafe_allow_html=True)
    st.title("💧 Engenharia e Custos - Cavaletes Hidráulicos")
    
    c_proj1, c_proj2 = st.columns([3, 1])
    st.session_state.nome_projeto_orcamento = c_proj1.text_input("🏷️ Nome do Projeto:", value=st.session_state.nome_projeto_orcamento, key="hidro_nome_proj")
    rev_proj = c_proj2.text_input("Revisão", value="R-00", key="hidro_rev_proj")
    st.markdown("---")

    def normalizar_string_busca(texto):
        return re.sub(r'[\s\-\"°Ø\’\']+', '', str(texto)).lower().strip()

    PESOS_SCH40 = {"1/2\"": 1.27, "3/4\"": 1.69, "1\"": 2.50, "1.1/4\"": 3.39, "1.1/2\"": 4.05, "2\"": 5.44, "2.1/2\"": 8.63, "3\"": 11.29, "4\"": 16.07, "5\"": 21.77, "6\"": 28.26, "8\"": 42.55, "10\"": 60.31, "12\"": 79.70}
    bitolas_roscadas = ["1/2\"", "3/4\"", "1\"", "1.1/4\"", "1.1/2\"", "2\""]
    bitolas_soldadas = ["2.1/2\"", "3\"", "4\"", "5\"", "6\"", "8\"", "10\"", "12\""]
    todas_bitolas = bitolas_roscadas + bitolas_soldadas

    # ==============================================================================
    # MOTOR GERADOR DE TEMPLATES INICIAIS (LEGADO)
    # ==============================================================================
    def gerar_templates_matriz():
        templates = []
        for eq in ["UTA", "Fancoil", "Fancolete", "Chiller", "Bomba"]:
            for vias in ["2 Vias", "3 Vias"]:
                if eq in ["Chiller", "Bomba"] and vias == "3 Vias": continue
                for ligacao in ["Roscado", "Flangeado"]:
                    
                    # KIT DINÂMICO (LINHA)
                    templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Tubo em aço carbono Aço carbono preto", "Medida": "Variável {b}", "Qtd": 3.0 if eq in ["UTA", "Chiller"] else 2.0})
                    templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Isolamento Espuma elastomérica, espessura 25 mm", "Medida": "Variável {b}", "Qtd": 2.0})
                    templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Rechapeamento chapa de alumínio liso, espessura 0,5 mm", "Medida": "Variável {b}", "Qtd": 2.0})
                    templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Válvula balanceadora", "Medida": "Variável {b}", "Qtd": 1.0})
                    
                    # CORREÇÃO APLICADA AQUI: A válvula de controle volta a ser {b-1}
                    templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": f"Válvula {vias[0]} vias, motorizada com atuador proporcional", "Medida": "Variável {b-1}", "Qtd": 1.0})
                    
                    if ligacao == "Roscado":
                        templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Curva 90° Ferro maleável galvanizado", "Medida": "Variável {b}", "Qtd": 6.0 if vias == "3 Vias" else 4.0})
                        templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Válvula de bloqueio tipo gaveta Corpo de bronze, catelo roscado internos em bronze, haste fixa.", "Medida": "Variável {b}", "Qtd": 3.0 if vias == "3 Vias" else 2.0})
                        templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "União com assento cônico Ferro maleável galvanizado", "Medida": "Variável {b}", "Qtd": 6.0 if vias == "3 Vias" else 5.0})
                        templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Filtro \"Y\" Corpo, tampa e tampão de bronze filtro de aço inoxidável", "Medida": "Variável {b}", "Qtd": 1.0})
                        templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Conexão T 90° Ferro maleável galvanizado", "Medida": "Variável {b}", "Qtd": 2.0 if vias == "3 Vias" else 1.0})
                        templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Niple duplo Ferro maleável galvanizado", "Medida": "Variável {b}", "Qtd": 6.0 if vias == "3 Vias" else 4.0})
                        templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Luva de redução - solda/rosca Aço carbono preto", "Medida": "Variável {b} x 1/2\"", "Qtd": 2.0})
                    else:
                        templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Curva 90° Aço carbono preto SCH40", "Medida": "Variável {b}", "Qtd": 6.0 if vias == "3 Vias" else 4.0})
                        templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Válvula borboleta flangeada", "Medida": "Variável {b}", "Qtd": 3.0 if vias == "3 Vias" else 2.0})
                        templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Flange sobreposto tipo slip-on aço forjado", "Medida": "Variável {b}", "Qtd": 14.0 if vias == "3 Vias" else 12.0})
                        templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Filtro Y Flangeado Ferro Fundido", "Medida": "Variável {b}", "Qtd": 1.0})
                        templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Dinâmico", "Item Base": "Conexão T 90° Aço carbono preto SCH40", "Medida": "Variável {b}", "Qtd": 2.0 if vias == "3 Vias" else 1.0})

                    # KIT FIXO
                    templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Fixo", "Item Base": "Poço para termômetro Latão haste = 38mm", "Medida": "1/2\"", "Qtd": 2.0})
                    templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Fixo", "Item Base": "Termômetro tipo capela. Escala 0° a 40°C", "Medida": "3/8\"", "Qtd": 2.0})
                    templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Fixo", "Item Base": "Robinete com alivio para manômetro Latão", "Medida": "1/2\"", "Qtd": 1.0})
                    templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Fixo", "Item Base": "Manometro tipo dial circular glicerinado, escala 0 a 4kgf/cm²", "Medida": "1/2\"", "Qtd": 1.0})
                    templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Fixo", "Item Base": "Luva Ferro maleável galvanizado", "Medida": "1/2\"", "Qtd": 2.0})
                    templates.append({"Equipamento": eq, "Vias": vias, "Ligação": ligacao, "Grupo": "Kit Fixo", "Item Base": "Niple duplo Ferro maleável galvanizado", "Medida": "1/4\"", "Qtd": 2.0})
        return templates

    # --- PONTES COM O GOOGLE SHEETS PARA OS TEMPLATES ---
    def carregar_templates_do_banco():
        try:
            sh = conectar_google_sheets()
            try:
                ws = sh.worksheet("Templates_Hidraulica")
                dados = ws.get_all_records()
                if len(dados) > 0:
                    return dados
            except: pass 
        except: pass
        return gerar_templates_matriz() 

    def salvar_templates_no_banco(matriz):
        try:
            sh = conectar_google_sheets()
            try:
                ws = sh.worksheet("Templates_Hidraulica")
            except:
                ws = sh.add_worksheet(title="Templates_Hidraulica", rows="2000", cols="10")
            
            ws.clear()
            if matriz:
                df = pd.DataFrame(matriz)
                ws.append_rows([df.columns.tolist()] + df.values.tolist())
        except Exception as e:
            st.warning(f"⚠️ Padrão salvo na memória, mas falhou ao gravar no Google: {e}")

    # ====================================================================
    # 🛑 INICIALIZAÇÃO BLINDADA (ÚNICA E DEFINITIVA)
    # ====================================================================
    if 'matriz_templates' not in st.session_state:
        st.session_state.matriz_templates = carregar_templates_do_banco()
        
    if 'trigger_refresh' not in st.session_state:
        st.session_state.trigger_refresh = 0

    # AS 4 ABAS DO APLICATIVO'

    def obter_bitola_menor(bitola_atual, passos=1):
        lista = ["1/2\"", "3/4\"", "1\"", "1.1/4\"", "1.1/2\"", "2\"", "2.1/2\"", "3\"", "4\"", "5\"", "6\"", "8\"", "10\"", "12\""]
        if bitola_atual in lista:
            idx = lista.index(bitola_atual)
            return lista[idx - passos] if idx >= passos else lista[0]
        return bitola_atual

    def construir_nome_peca(item_base, medida, bitola_cavalete):
        if medida == "Nenhum": return item_base
        bitola_b1 = obter_bitola_menor(bitola_cavalete, 1)
        bitola_b2 = obter_bitola_menor(bitola_cavalete, 2)
        
        # Limpa a palavra "Variável " e substitui as tags matemáticas com precisão
        medida_final = medida.replace("Variável ", "")
        medida_final = medida_final.replace("{b-2}", bitola_b2)
        medida_final = medida_final.replace("{b-1}", bitola_b1)
        medida_final = medida_final.replace("{b}", bitola_cavalete)
        
        return f"{item_base} - Ø {medida_final}"

    # SINCRONIZAÇÃO INTELIGENTE DA TABELA DE PREÇOS
    if st.session_state.get('versao_banco_hidro') != 'v15':
        banco_temp = {}
        for template in st.session_state.matriz_templates:
            bitolas_alvo = bitolas_roscadas if template["Ligação"] == "Roscado" else bitolas_soldadas
            for b in bitolas_alvo:
                nome_final = construir_nome_peca(template["Item Base"], template["Medida"], b)
                if nome_final not in banco_temp:
                    unid = "m" if "Tubo" in nome_final or "Isolamento" in nome_final or "Rechapeamento" in nome_final else "pç"
                    preco = round(PESOS_SCH40.get(b, 2.0) * 12.50, 2) if "Tubo" in nome_final else 150.0
                    banco_temp[nome_final] = {"Item / Componente": nome_final, "Preço Unitário (R$)": preco, "Unidade": unid}
        
        banco_temp["Mão de Obra de Montagem Hidráulica (Por Polegada)"] = {"Item / Componente": "Mão de Obra de Montagem Hidráulica (Por Polegada)", "Preço Unitário (R$)": 120.0, "Unidade": "pol"}
        banco_temp["Mão de Obra de Isolamento Térmico (Por Polegada)"] = {"Item / Componente": "Mão de Obra de Isolamento Térmico (Por Polegada)", "Preço Unitário (R$)": 95.0, "Unidade": "pol"}

        try:
            sh_hidro = conectar_google_sheets()
            aba_h = sh_hidro.worksheet("Precos_Hidraulica_Itens").get_all_records()
            for row in aba_h:
                chave = normalizar_string_busca(row.get("Item / Componente", ""))
                for key_banco in banco_temp.keys():
                    if normalizar_string_busca(key_banco) == chave:
                        banco_temp[key_banco]["Preço Unitário (R$)"] = float(row.get("Preço Unitário (R$)", 0.0))
        except: pass
        
        st.session_state.banco_precos_hidraulica = list(banco_temp.values())
        st.session_state.versao_banco_hidro = 'v15'

    if 'cavaletes_selecionados' not in st.session_state: st.session_state.cavaletes_selecionados = []

    def dimensionar_bitola_pelo_abaco(vazao_m3h, perda_mmca_m, tipo_sistema):
        C = 130.0 if "Fechado" in tipo_sistema else 120.0
        q_m3s = vazao_m3h / 3600.0
        hf_m_m = perda_mmca_m / 1000.0
        if hf_m_m <= 0 or q_m3s <= 0: return "1/2\""
        numerador = 10.67 * (q_m3s ** 1.852)
        denominador = (C ** 1.852) * hf_m_m
        d_interno_calc_mm = ((numerador / denominador) ** (1 / 4.8704)) * 1000.0
        tabela_sch40 = [(15.8, "1/2\""), (20.9, "3/4\""), (26.6, "1\""), (35.1, "1.1/4\""), (40.9, "1.1/2\""), (52.5, "2\""), (62.7, "2.1/2\""), (77.9, "3\""), (102.3, "4\""), (128.2, "5\""), (154.1, "6\""), (202.7, "8\""), (254.5, "10\""), (303.2, "12\"")]
        bitola_selecionada = "12\""
        for int_mm, bit_nom in tabela_sch40:
            if d_interno_calc_mm <= int_mm: bitola_selecionada = bit_nom; break
        return bitola_selecionada

    # AS 4 ABAS DO APLICATIVO
    aba_cadastro_hidro, aba_precos_hidro, aba_padroes_hidro, aba_resumo_hidro = st.tabs([
        "🔧 Dimensionamento", "💲 Tabela de Preços", "📝 Central de Padrões (Templates)", "📊 Resumos / BOM"
    ])

    with aba_cadastro_hidro:
        st.subheader("Configuração Estrutural de Hidráulica")
        
        tipo_sistema = st.radio("Tipo de Sistema da Instalação:", ["Fechado (Água Gelada/Quente)", "Aberto (Água de Condensação/Torre)"], horizontal=True)
        st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
        metodo_dimensionamento = st.radio("Método de dimensionamento:", ["📏 Definir diretamente por Bitola comercial", "🌊 Dimensionar automaticamente por Vazão"], horizontal=True)
        st.markdown("<br>", unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns(3)
        tipo_equip = col1.selectbox("Tipo de Equipamento:", ["UTA", "Fancoil", "Fancolete", "Chiller", "Bomba"])
        if tipo_equip in ["Chiller", "Bomba"]: 
            tipo_vias = col2.selectbox("Válvula:", ["2 Vias"], disabled=True)
        else: 
            tipo_vias = col2.selectbox("Válvula:", ["2 Vias", "3 Vias"])
        
        if "Bitola comercial" in metodo_dimensionamento:
            bitola_final = col3.selectbox("Bitola comercial:", todas_bitolas)
            vazao_calculada = 0.0
        else:
            vazao_calculada = col3.number_input("Vazão de Água (m³/h):", min_value=0.1, step=0.5, value=5.0)
            tipo_aplicacao = st.radio("Exigência Acústica:", ["Escritório / Áreas Críticas", "Indústria / Áreas Técnicas"], horizontal=True)
            perda_carga = st.slider("Perda de Carga (mmCA/m):", 10, 40, 25) if "Escritório" in tipo_aplicacao else st.slider("Perda de Carga (mmCA/m):", 41, 90 if "Aberto" in tipo_sistema else 100, 70)
            bitola_final = dimensionar_bitola_pelo_abaco(vazao_calculada, perda_carga, tipo_sistema)
            st.info(f"📈 Bitola comercial recomendada: **{bitola_final}**.")
            
        col_q, col_t = st.columns([1, 2])
        qtd = col_q.number_input("Quantidade de conjuntos:", min_value=1, step=1, value=1)
        tag_equip = col_t.text_input("TAG identificadora (Opcional):", placeholder="Ex: UTA-01, CH-01...")
        
        # --- BOTÃO ANTIGO (USANDO TEMPLATES MANUAIS) ---
        if st.button("➕ Adicionar Cavalete ao Levantamento", type="secondary"):
            dict_pol = {"1/4\"": 0.25, "3/8\"": 0.375, "1/2\"": 0.5, "3/4\"": 0.75, "1\"": 1.0, "1.1/4\"": 1.25, "1.1/2\"": 1.5, "2\"": 2.0, "2.1/2\"": 2.5, "3\"": 3.0, "4\"": 4.0, "5\"": 5.0, "6\"": 6.0, "8\"": 8.0, "10\"": 10.0, "12\"": 12.0}
            pol_dec = dict_pol.get(bitola_final, 1.0)
            
            is_aberto = "Aberto" in tipo_sistema
            ligacao = "Roscado" if bitola_final in bitolas_roscadas else "Flangeado"
            
            composicao_kit = []
            regras_ativas = [t for t in st.session_state.matriz_templates if t["Equipamento"] == tipo_equip and t["Vias"] == tipo_vias and t["Ligação"] == ligacao]
            
            for regra in regras_ativas:
                nome_final = construir_nome_peca(regra["Item Base"], regra["Medida"], bitola_final)
                if is_aberto and ("Isolamento" in nome_final or "Rechapeamento" in nome_final): continue
                if regra["Qtd"] > 0:
                    composicao_kit.append({"nome": nome_final, "qtd": regra["Qtd"]})
            
            dict_precos_memoria = {normalizar_string_busca(row["Item / Componente"]): float(row["Preço Unitário (R$)"]) for row in st.session_state.banco_precos_hidraulica}
            custo_material_total_kit = sum(dict_precos_memoria.get(normalizar_string_busca(comp["nome"]), 0.0) * comp["qtd"] for comp in composicao_kit)
            mo_mont_calculado = dict_precos_memoria.get(normalizar_string_busca("Mão de Obra de Montagem Hidráulica (Por Polegada)"), 120.0) * pol_dec * 12.0
            
            if is_aberto:
                mo_isol_calculado = 0.0
            else:
                preco_base_isol_metro = dict_precos_memoria.get(normalizar_string_busca("Mão de Obra de Isolamento Térmico (Por Polegada)"), 95.0) * pol_dec
                qtd_tubos_linear = sum(c["qtd"] for c in composicao_kit if "tubo" in c["nome"].lower())
                qtd_conexoes = sum(c["qtd"] for c in composicao_kit if any(x in c["nome"].lower() for x in ["curva", "conexão t", "redução", "cotovelo"]))
                qtd_valvulas = sum(c["qtd"] for c in composicao_kit if any(x in c["nome"].lower() for x in ["válvula", "filtro"]))
                metragem_equivalente = qtd_tubos_linear + (qtd_conexoes * 1.5) + (qtd_valvulas * 2.0)
                mo_isol_calculado = preco_base_isol_metro * metragem_equivalente
            
            st.session_state.cavaletes_selecionados.append({
                "id": str(uuid.uuid4()), "tag": tag_equip if tag_equip else "S/ TAG",
                "equipamento": tipo_equip, "vias": tipo_vias, "bitola": bitola_final, "quantidade": qtd,
                "vazao": vazao_calculada, "sistema": tipo_sistema, "custo_mat_unit": custo_material_total_kit,
                "mo_mont_unit": mo_mont_calculado, "mo_isol_unit": mo_isol_calculado, "composicao": composicao_kit
            })
            st.toast(f"✅ Conjunto Ø {bitola_final} adicionado!", icon="👍")
            st.rerun()

        
        # ====================================================================
        # LISTA DE CAVALETES ADICIONADOS
        # ====================================================================
        st.markdown("---")
        st.markdown("### 📋 Cavaletes Adicionados no Projeto")
        if not st.session_state.cavaletes_selecionados: 
            st.info("Nenhum item adicionado no levantamento.")
        else:
            for idx, cav in enumerate(st.session_state.cavaletes_selecionados):
                c_inf, c_tg, c_qt, c_rm = st.columns([4, 3, 2, 2])
                sys_lbl = " [Aberto]" if "Aberto" in cav.get("sistema", "") else ""
                c_inf.write(f"**Cavalete {cav.get('equipamento', 'EQ')} ({cav.get('vias', 'N/A')})** - Ø {cav.get('bitola', '')}{sys_lbl}")
                c_tg.write(f"TAG: `{cav.get('tag', 'S/ TAG')}`")
                c_qt.write(f"Qtd: **{cav.get('quantidade', 1)} cjs**")
                
                if c_rm.button("🗑️", key=f"rm_h_cv_{cav.get('id', idx)}"):
                    st.session_state.cavaletes_selecionados.pop(idx)
                    st.rerun()
                    
            if st.button("🗑️ Excluir Todos", type="secondary"):
                st.session_state.cavaletes_selecionados = []
                st.rerun()

    with aba_padroes_hidro:
        st.header("⚙️ Central de Padrões e Receitas (Templates)")
        st.caption("Selecione o equipamento e a variação que deseja editar. Suas alterações serão gravadas na nuvem (Google Sheets).")
        
        f_eq, f_vi, f_li = st.columns(3)
        sel_eq = f_eq.selectbox("1. Equipamento:", ["UTA", "Fancoil", "Fancolete", "Chiller", "Bomba"], key="tpl_eq")
        sel_vi = f_vi.selectbox("2. Vias:", ["2 Vias"] if sel_eq in ["Chiller", "Bomba"] else ["2 Vias", "3 Vias"], key="tpl_vi")
        sel_li = f_li.selectbox("3. Ligação (Tamanho):", ["Roscado", "Flangeado"], help="Roscado (1/2\" a 2\") | Flangeado (2.1/2\" a 12\")", key="tpl_li")
        
        st.markdown(f"#### Editando Receita: **{sel_eq} | {sel_vi} | {sel_li}**")
        
        df_templates = pd.DataFrame(st.session_state.matriz_templates)
        filtro = (df_templates["Equipamento"] == sel_eq) & (df_templates["Vias"] == sel_vi) & (df_templates["Ligação"] == sel_li)
        df_editavel = df_templates[filtro][["Grupo", "Item Base", "Medida", "Qtd"]].copy()
        
        # A chave tem um gatilho de refresh dinâmico para não bugar a memória da tabela
        df_editado = st.data_editor(
            df_editavel,
            column_config={
                "Grupo": st.column_config.SelectboxColumn("Grupo do Kit", options=["Kit Dinâmico", "Kit Fixo"], required=True),
                "Item Base": st.column_config.TextColumn("Descrição Base (sem o diâmetro)", required=True),
                "Medida": st.column_config.SelectboxColumn("Medida (Bitola)", options=["Variável {b}", "Variável {b-1}", "Variável {b-2}", "Variável {b} x {b-1}", "Variável {b} x {b-2}", "Variável {b-1} x {b-2}", "Variável {b} x 1/2\"", "Variável {b} x 3/4\"", "1/2\"", "3/4\"", "3/8\"", "1/4\"", "Nenhum"], required=True),
                "Qtd": st.column_config.NumberColumn("Quantidade", min_value=0.0, format="%.2f")
            },
            num_rows="dynamic", use_container_width=True, key=f"grid_tpl_{sel_eq}_{sel_vi}_{sel_li}_{st.session_state.get('trigger_refresh', 0)}"
        )
        
        if st.button("💾 Salvar Esta Tabela na Nuvem", type="primary"):
            nova_matriz = [row for row in st.session_state.matriz_templates if not (row["Equipamento"] == sel_eq and row["Vias"] == sel_vi and row["Ligação"] == sel_li)]
            for _, row in df_editado.iterrows():
                if pd.notna(row["Item Base"]) and str(row["Item Base"]).strip() != "":
                    nova_matriz.append({"Equipamento": sel_eq, "Vias": sel_vi, "Ligação": sel_li, "Grupo": row["Grupo"], "Item Base": str(row["Item Base"]).strip(), "Medida": row["Medida"], "Qtd": float(row["Qtd"]) if pd.notna(row["Qtd"]) else 1.0})
            
            st.session_state.matriz_templates = nova_matriz
            st.session_state.versao_banco_hidro = 'forcar_recalculo'
            st.session_state.trigger_refresh = st.session_state.get('trigger_refresh', 0) + 1 
            
            with st.spinner("Gravando no Google Sheets..."):
                salvar_templates_no_banco(nova_matriz)
                st.success("✅ Padrão salvo com sucesso no Banco de Dados!")
                st.rerun() 
            
        st.markdown("---")
        with st.expander("➕ Inserir Novo Item nesta Receita", expanded=False):
            nomes_base_existentes = set()
            for item in st.session_state.banco_precos_hidraulica:
                nome_bruto = item["Item / Componente"]
                base = nome_bruto.split("- Ø")[0].strip() if "- Ø" in nome_bruto else nome_bruto
                nomes_base_existentes.add(base)
            nomes_base_existentes = sorted(list(nomes_base_existentes))

            with st.form("form_add_item_template"):
                c_a1, c_a2, c_a3 = st.columns(3)
                add_grp = c_a1.selectbox("Grupo", ["Kit Dinâmico", "Kit Fixo"], help="Kit Dinâmico para peças que mudam a bitola. Fixo para peças estáticas.")
                add_base = c_a2.selectbox("1. Selecionar Existente", ["-- Selecione --"] + nomes_base_existentes)
                add_novo = c_a3.text_input("2. OU Digite um Nome Novo")
                
                c_a4, c_a5, c_a6 = st.columns(3)
                add_med = c_a4.selectbox("Medida", ["Variável {b}", "Variável {b-1}", "Variável {b-2}", "Variável {b} x {b-1}", "Variável {b} x {b-2}", "Variável {b-1} x {b-2}", "Variável {b} x 1/2\"", "Variável {b} x 3/4\"", "1/2\"", "3/4\"", "3/8\"", "1/4\"", "Nenhum"])
                add_qtd = c_a5.number_input("Quantidade", min_value=0.01, value=1.0)
                
                if st.form_submit_button("Inserir na Receita Atual"):
                    nome_escolhido = add_novo if add_novo.strip() != "" else (add_base if add_base != "-- Selecione --" else "")
                    if nome_escolhido != "":
                        st.session_state.matriz_templates.append({"Equipamento": sel_eq, "Vias": sel_vi, "Ligação": sel_li, "Grupo": add_grp, "Item Base": nome_escolhido, "Medida": add_med, "Qtd": add_qtd})
                        st.session_state.versao_banco_hidro = 'forcar_recalculo'
                        st.session_state.trigger_refresh = st.session_state.get('trigger_refresh', 0) + 1 
                        
                        salvar_templates_no_banco(st.session_state.matriz_templates)
                        st.toast(f"Item inserido e salvo na nuvem!", icon="✅")
                        st.rerun()
                    else:
                        st.error("Selecione um item do banco ou digite um nome novo.")

        st.markdown("---")
        st.markdown("#### 🔄 Clonar Padrão")
        st.caption("Gostou de como ficou essa receita? Copie-a idêntica para outro cenário para não ter que refazer tudo.")
        
        c_clon1, c_clon2, c_clon3 = st.columns(3)
        clon_eq = c_clon1.selectbox("Copiar PARA Equipamento:", ["UTA", "Fancoil", "Fancolete", "Chiller", "Bomba"], index=["UTA", "Fancoil", "Fancolete", "Chiller", "Bomba"].index(sel_eq), key="clon_eq")
        clon_vi = c_clon2.selectbox("Copiar PARA Vias:", ["2 Vias", "3 Vias"], index=["2 Vias", "3 Vias"].index(sel_vi) if sel_vi in ["2 Vias", "3 Vias"] else 0, key="clon_vi")
        clon_li = c_clon3.selectbox("Copiar PARA Ligação:", ["Roscado", "Flangeado"], index=["Roscado", "Flangeado"].index(sel_li), key="clon_li")
        
        if st.button("📋 Executar Clonagem e Salvar na Nuvem", type="secondary"):
            if clon_eq in ["Chiller", "Bomba"] and clon_vi == "3 Vias":
                st.error("⚠️ Atenção: Chiller e Bomba só trabalham com configuração de 2 Vias no sistema.")
            else:
                matriz_sem_alvo = [row for row in st.session_state.matriz_templates if not (row["Equipamento"] == clon_eq and row["Vias"] == clon_vi and row["Ligação"] == clon_li)]
                for _, row in df_editado.iterrows():
                    matriz_sem_alvo.append({"Equipamento": clon_eq, "Vias": clon_vi, "Ligação": clon_li, "Grupo": row["Grupo"], "Item Base": row["Item Base"], "Medida": row["Medida"], "Qtd": row["Qtd"]})
                
                st.session_state.matriz_templates = matriz_sem_alvo
                st.session_state.trigger_refresh = st.session_state.get('trigger_refresh', 0) + 1 
                
                with st.spinner("Clonando no Google Sheets..."):
                    salvar_templates_no_banco(matriz_sem_alvo)
                    st.toast(f"Receita clonada com sucesso para {clon_eq} | {clon_vi} | {clon_li}!", icon="📋")
                    st.rerun()
            
        st.markdown("---")
        with st.expander("➕ Inserir Novo Item nesta Receita", expanded=False):
            nomes_base_existentes = set()
            for item in st.session_state.banco_precos_hidraulica:
                nome_bruto = item["Item / Componente"]
                base = nome_bruto.split("- Ø")[0].strip() if "- Ø" in nome_bruto else nome_bruto
                nomes_base_existentes.add(base)
            nomes_base_existentes = sorted(list(nomes_base_existentes))

            with st.form("form_add_item_template"):
                c_a1, c_a2, c_a3 = st.columns(3)
                add_grp = c_a1.selectbox("Grupo", ["Kit Dinâmico", "Kit Fixo"], help="Kit Dinâmico para peças que mudam a bitola. Fixo para peças estáticas.")
                add_base = c_a2.selectbox("1. Selecionar Existente", ["-- Selecione --"] + nomes_base_existentes)
                add_novo = c_a3.text_input("2. OU Digite um Nome Novo")
                
                c_a4, c_a5, c_a6 = st.columns(3)
                add_med = c_a4.selectbox("Medida", ["Variável {b}", "Variável {b-1}", "Variável {b-2}", "Variável {b} x {b-1}", "Variável {b} x {b-2}", "Variável {b-1} x {b-2}", "Variável {b} x 1/2\"", "Variável {b} x 3/4\"", "1/2\"", "3/4\"", "3/8\"", "1/4\"", "Nenhum"])
                add_qtd = c_a5.number_input("Quantidade", min_value=0.01, value=1.0)
                
                if st.form_submit_button("Inserir na Receita Atual"):
                    nome_escolhido = add_novo if add_novo.strip() != "" else (add_base if add_base != "-- Selecione --" else "")
                    if nome_escolhido != "":
                        st.session_state.matriz_templates.append({"Equipamento": sel_eq, "Vias": sel_vi, "Ligação": sel_li, "Grupo": add_grp, "Item Base": nome_escolhido, "Medida": add_med, "Qtd": add_qtd})
                        st.session_state.versao_banco_hidro = 'forcar_recalculo'
                        
                        salvar_templates_no_banco(st.session_state.matriz_templates)
                        st.toast(f"Item inserido e salvo na nuvem!", icon="✅")
                        st.rerun()
                    else:
                        st.error("Selecione um item do banco ou digite um nome novo.")

        st.markdown("---")
        st.markdown("#### 🔄 Clonar Padrão")
        st.caption("Gostou de como ficou essa receita? Copie-a idêntica para outro cenário para não ter que refazer tudo.")
        
        c_clon1, c_clon2, c_clon3 = st.columns(3)
        clon_eq = c_clon1.selectbox("Copiar PARA Equipamento:", ["UTA", "Fancoil", "Fancolete", "Chiller", "Bomba"], index=["UTA", "Fancoil", "Fancolete", "Chiller", "Bomba"].index(sel_eq), key="clon_eq")
        clon_vi = c_clon2.selectbox("Copiar PARA Vias:", ["2 Vias", "3 Vias"], index=["2 Vias", "3 Vias"].index(sel_vi) if sel_vi in ["2 Vias", "3 Vias"] else 0, key="clon_vi")
        clon_li = c_clon3.selectbox("Copiar PARA Ligação:", ["Roscado", "Flangeado"], index=["Roscado", "Flangeado"].index(sel_li), key="clon_li")
        
        if st.button("📋 Executar Clonagem e Salvar na Nuvem", type="secondary"):
            if clon_eq in ["Chiller", "Bomba"] and clon_vi == "3 Vias":
                st.error("⚠️ Atenção: Chiller e Bomba só trabalham com configuração de 2 Vias no sistema.")
            else:
                matriz_sem_alvo = [row for row in st.session_state.matriz_templates if not (row["Equipamento"] == clon_eq and row["Vias"] == clon_vi and row["Ligação"] == clon_li)]
                for _, row in df_editado.iterrows():
                    matriz_sem_alvo.append({"Equipamento": clon_eq, "Vias": clon_vi, "Ligação": clon_li, "Grupo": row["Grupo"], "Item Base": row["Item Base"], "Medida": row["Medida"], "Qtd": row["Qtd"]})
                
                st.session_state.matriz_templates = matriz_sem_alvo
                with st.spinner("Clonando no Google Sheets..."):
                    salvar_templates_no_banco(matriz_sem_alvo)
                    st.toast(f"Receita clonada com sucesso para {clon_eq} | {clon_vi} | {clon_li}!", icon="📋")
                    st.rerun()

    with aba_precos_hidro:
        st.header("Gestão Sênior de Preços (Componentes Abertos)")
        st.markdown("### ⚖️ Precificação Inteligente de Tubos (Por Kg)")
        col_kg1, col_kg2 = st.columns([1, 2])
        novo_preco_kg = col_kg1.number_input("Preço do Aço Carbono (R$/kg):", min_value=0.1, value=12.50, step=0.50)
        
        if col_kg2.button("🔄 Converter R$/kg para R$/m e Atualizar Tubos", type="secondary"):
            for item in st.session_state.banco_precos_hidraulica:
                if "Tubo em aço carbono" in item["Item / Componente"] and "sem costura" in item["Item / Componente"]:
                    bitola_str = item["Item / Componente"].split("Ø ")[-1].strip()
                    if bitola_str in PESOS_SCH40:
                        item["Preço Unitário (R$)"] = round(PESOS_SCH40[bitola_str] * novo_preco_kg, 2)
            st.toast("✅ Preços convertidos com sucesso!", icon="👍")
            st.rerun()
            
        st.markdown("---")
        st.markdown("### 🔄 Sincronização em Lote (Excel)")
        
        ch1, ch2 = st.columns(2)
        
        with ch1:
            st.markdown("#### 📥 Exportar para Cotar")
            def gerar_planilha_itens_hidro(com_precos):
                buf = io.BytesIO(); wb = openpyxl.Workbook(); ws = wb.active
                ws.append(["Item / Componente", "Preço Unitário (R$)", "Unidade"])
                for r in st.session_state.banco_precos_hidraulica:
                    ws.append([r["Item / Componente"], r["Preço Unitário (R$)"] if com_precos else "", r["Unidade"]])
                wb.save(buf); buf.seek(0); return buf
            st.download_button("Baixar Planilha (Excel)", data=gerar_planilha_itens_hidro(True), file_name="Precos_Hidro.xlsx", use_container_width=True)
            
        with ch2:
            st.markdown("#### 📂 Devolver Planilha Cotada")
            upl_hidro = st.file_uploader("Selecione a planilha:", type=["xlsx", "xls"], label_visibility="collapsed")
            
            if st.button("💾 Salvar Planilha no Banco", type="primary", use_container_width=True):
                if upl_hidro is not None:
                    try:
                        df_up = pd.read_excel(upl_hidro)
                        df_up.rename(columns=lambda x: str(x).strip(), inplace=True)
                        
                        if "Item / Componente" in df_up.columns:
                            df_up = df_up.dropna(subset=["Item / Componente"])
                            def limpar_preco(val):
                                if isinstance(val, str):
                                    try: return float(val.replace("R$", "").replace(".", "").replace(",", ".").strip())
                                    except: return 0.0
                                return float(val) if pd.notnull(val) else 0.0
                                
                            df_up['Preço Unitário (R$)'] = df_up['Preço Unitário (R$)'].apply(limpar_preco)
                            df_up = df_up.fillna("")
                            
                            st.session_state.banco_precos_hidraulica = df_up.to_dict('records')
                            
                            import datetime
                            st.session_state['data_ultima_atualizacao'] = datetime.datetime.now().strftime("%d/%m/%Y às %H:%M:%S")
                            
                            try:
                                sh_cloud = conectar_google_sheets()
                                ws_ci = sh_cloud.worksheet("Precos_Hidraulica_Itens")
                                ws_ci.clear()
                                linhas = [df_up.columns.tolist()] + df_up.values.tolist()
                                ws_ci.append_rows(linhas)
                                st.success("✅ Valores gravados com sucesso na Nuvem!")
                            except Exception as e:
                                st.warning(f"⚠️ Valores atualizados na tela, mas falha ao sincronizar com o Google: {e}")
                    except Exception as e:
                        st.error(f"Erro na leitura do arquivo: {e}")
                else:
                    st.warning("⚠️ Anexe um arquivo antes de salvar.")

        st.markdown("---")
        st.subheader("📚 Itens Cadastrados no Banco de Dados")
        
        if 'data_ultima_atualizacao' in st.session_state:
            st.caption(f"🕒 **Última atualização:** {st.session_state['data_ultima_atualizacao']}")
        else:
            st.caption("🕒 **Última atualização:** Não registrada nesta sessão")
            
        if not st.session_state.get('banco_precos_hidraulica'):
             try:
                 df_view_hidro = carregar_precos_hidraulica_itens()
                 st.session_state.banco_precos_hidraulica = df_view_hidro.to_dict('records')
             except:
                 st.session_state.banco_precos_hidraulica = []

        if st.session_state.get('banco_precos_hidraulica'):
            df_display = pd.DataFrame(st.session_state.banco_precos_hidraulica)
            
            if not df_display.empty and "Preço Unitário (R$)" in df_display.columns:
                st.info("💡 Você pode editar os preços dando dois cliques nos valores da tabela abaixo!")
                df_editado = st.data_editor(
                    df_display, 
                    use_container_width=True, 
                    hide_index=True, 
                    height=500,
                    column_config={
                        "Preço Unitário (R$)": st.column_config.NumberColumn(
                            "Preço Unitário (R$)",
                            help="Edite o preço aqui",
                            format="R$ %.2f",
                            step=0.01
                        ),
                        "Item / Componente": st.column_config.TextColumn(disabled=True),
                        "Unidade": st.column_config.TextColumn(disabled=True)
                    },
                    key="editor_manual_precos"
                )
                
                if st.button("💾 Salvar Edições Manuais", type="secondary"):
                    st.session_state.banco_precos_hidraulica = df_editado.to_dict('records')
                    import datetime
                    st.session_state['data_ultima_atualizacao'] = datetime.datetime.now().strftime("%d/%m/%Y às %H:%M:%S")
                    
                    try:
                        sh_cloud = conectar_google_sheets()
                        ws_ci = sh_cloud.worksheet("Precos_Hidraulica_Itens")
                        ws_ci.clear()
                        linhas = [df_editado.columns.tolist()] + df_editado.values.tolist()
                        ws_ci.append_rows(linhas)
                        st.success("✅ Edições manuais gravadas no sistema!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erro ao salvar na nuvem: {e}")
            else:
                st.warning("Colunas 'Item / Componente' ou 'Preço Unitário (R$)' não encontradas.")
        else:
            st.warning("Banco de dados vazio.")

    with aba_resumo_hidro:
        st.header("📊 Resumo e Listas de Materiais (BOM)")
        if not st.session_state.cavaletes_selecionados:
            st.warning("Efetue a adição de cavaletes para gerar os quantitativos consolidados.")
        else:
            df_precos_lookup = pd.DataFrame(st.session_state.banco_precos_hidraulica)
            dict_lookup_valores = {normalizar_string_busca(k): v for k, v in zip(df_precos_lookup["Item / Componente"], df_precos_lookup["Preço Unitário (R$)"])}
            dict_lookup_unidades = {normalizar_string_busca(k): v for k, v in zip(df_precos_lookup["Item / Componente"], df_precos_lookup["Unidade"])}
            dict_nomes_originais = {normalizar_string_busca(k): k for k in df_precos_lookup["Item / Componente"]}
            
            total_hidro_material, total_hidro_montagem, total_hidro_isolamento = 0.0, 0.0, 0.0
            resumo_financeiro_quadros = []
            materiais_condensados_lista = {} 
            
            preco_base_mont = dict_lookup_valores.get(normalizar_string_busca("Mão de Obra de Montagem Hidráulica (Por Polegada)"), 120.0)
            preco_base_isol = dict_lookup_valores.get(normalizar_string_busca("Mão de Obra de Isolamento Térmico (Por Polegada)"), 95.0)
            dict_pol = {"1/4\"": 0.25, "3/8\"": 0.375, "1/2\"": 0.5, "3/4\"": 0.75, "1\"": 1.0, "1.1/4\"": 1.25, "1.1/2\"": 1.5, "2\"": 2.0, "2.1/2\"": 2.5, "3\"": 3.0, "4\"": 4.0, "5\"": 5.0, "6\"": 6.0, "8\"": 8.0, "10\"": 10.0, "12\"": 12.0}
            
            for cav in st.session_state.cavaletes_selecionados:
                pol_dec = dict_pol.get(cav["bitola"], 1.0)
                is_aberto = "Aberto" in cav.get("sistema", "")
                
                m_sub_unit = 0.0
                qtd_tubos = 0.0; qtd_conexoes = 0.0; qtd_valvulas = 0.0
                
                for comp_g in cav.get("composicao", []):
                    n_busca = normalizar_string_busca(comp_g["nome"])
                    pr_u = dict_lookup_valores.get(n_busca, 0.0)
                    m_sub_unit += pr_u * comp_g["qtd"]
                    
                    n_real = dict_nomes_originais.get(n_busca, comp_g["nome"])
                    qtd_tot = comp_g["qtd"] * cav["quantidade"]
                    if n_real not in materiais_condensados_lista: 
                        materiais_condensados_lista[n_real] = {'qtd': 0.0, 'tags': set()}
                    materiais_condensados_lista[n_real]['qtd'] += qtd_tot
                    if cav.get('tag', 'S/ TAG') != 'S/ TAG': 
                        materiais_condensados_lista[n_real]['tags'].add(cav.get('tag', 'S/ TAG'))
                    
                    nome_low = comp_g["nome"].lower()
                    if "tubo" in nome_low: qtd_tubos += comp_g["qtd"]
                    elif "curva" in nome_low or "conexão t" in nome_low or "redução" in nome_low or "cotovelo" in nome_low: qtd_conexoes += comp_g["qtd"]
                    elif "válvula" in nome_low or "filtro" in nome_low: qtd_valvulas += comp_g["qtd"]
                
                mo_m_u = preco_base_mont * pol_dec * 12.0
                if is_aberto:
                    mo_i_u = 0.0
                else:
                    metragem_equivalente = qtd_tubos + (qtd_conexoes * 1.5) + (qtd_valvulas * 2.0)
                    mo_i_u = preco_base_isol * pol_dec * metragem_equivalente
                
                cav["mo_mont_unit"] = mo_m_u
                cav["mo_isol_unit"] = mo_i_u
                
                m_sub = m_sub_unit * cav["quantidade"]
                mo_mont_sub = mo_m_u * cav["quantidade"]
                mo_isol_sub = mo_i_u * cav["quantidade"]
                
                total_hidro_material += m_sub
                total_hidro_montagem += mo_mont_sub
                total_hidro_isolamento += mo_isol_sub
                tag_atual = cav.get('tag', 'S/ TAG')
                sys_lbl = " [Aberto]" if "Aberto" in cav.get("sistema", "") else ""
                
                resumo_financeiro_quadros.append({
                    "TAG": tag_atual, 
                    "Conjunto": f"{cav.get('equipamento', '')} ({cav.get('vias', '')}) - Ø {cav.get('bitola', '')}{sys_lbl}",
                    "Qtd": cav["quantidade"], 
                    "Materiais": m_sub, 
                    "M.O. Montagem": mo_mont_sub, 
                    "M.O. Isolamento": mo_isol_sub, 
                    "Subtotal": m_sub + mo_mont_sub + mo_isol_sub
                })
                        
            c1_h, c2_h, c3_h = st.columns(3)
            c1_h.info(f"**Materiais:**\nR$ {total_hidro_material:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            c2_h.warning(f"**Mão de Obra:**\nR$ {(total_hidro_montagem + total_hidro_isolamento):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            c3_h.success(f"**TOTAL:**\nR$ {(total_hidro_material + total_hidro_montagem + total_hidro_isolamento):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            
            with st.expander("ℹ️ Entenda como a Mão de Obra é calculada pelo sistema"):
                st.markdown("""
                A Mão de Obra é parametrizada com base no indexador industrial de **"Polegada Equivalente"**.
                
                **1. Montagem Hidráulica:**
                O tempo e esforço de rosqueamento/soldagem crescem proporcionalmente ao diâmetro do tubo. O cálculo considera a média de pontos de acoplamento dentro de um cavalete.
                *Fórmula:* `Preço Base da Polegada x Diâmetro Nominal (decimal) x Fator de Montagem (12.0)`
                
                **2. Isolamento Térmico (Rechapeamento):**
                Aplicamos a técnica Sênior de **Metragem Equivalente**. Isolar um tubo reto é simples, mas fabricar caixas de válvulas consome muito mais tempo e material.
                O sistema varre a receita do cavalete e aplica os seguintes multiplicadores de esforço sobre o metro linear da polegada:
                *   Tubo Reto = `1.0x`
                *   Curvas, Tês e Reduções = `1.5x` (50% a mais de esforço)
                *   Válvulas e Filtros Y = `2.0x` (O dobro de esforço para caixas artesanais)
                """)

            st.markdown("---")
            st.subheader("📊 Resumo de Custos por Cavalete")
            df_res_quadros_disp = pd.DataFrame(resumo_financeiro_quadros)
            df_res_quadros_disp["Mão de Obra"] = df_res_quadros_disp["M.O. Montagem"] + df_res_quadros_disp["M.O. Isolamento"]
            df_res_quadros_disp = df_res_quadros_disp[["TAG", "Conjunto", "Qtd", "Materiais", "Mão de Obra", "Subtotal"]]
            for col in ["Materiais", "Mão de Obra", "Subtotal"]: df_res_quadros_disp[col] = df_res_quadros_disp[col].apply(lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            st.dataframe(df_res_quadros_disp, use_container_width=True, hide_index=True)
            
            st.markdown("---")
            st.subheader("📋 Lista Explodida (Por Equipamento)")
            for cav in st.session_state.cavaletes_selecionados:
                st.markdown(f"#### {cav.get('tag', 'S/ TAG')} - {cav.get('equipamento', '')} (Ø {cav.get('bitola', '')})")
                
                lista_itens_cav = []
                qtd_conjuntos = cav['quantidade']
                
                for comp in cav['composicao']:
                    n_busca = normalizar_string_busca(comp["nome"])
                    pr_u = dict_lookup_valores.get(n_busca, 0.0)
                    pr_t = pr_u * comp["qtd"] * qtd_conjuntos
                    unidade_item = dict_lookup_unidades.get(n_busca, "un")
                    
                    lista_itens_cav.append({
                        "Componente": comp["nome"],
                        "Unid": unidade_item,
                        "Qtd Unit": comp["qtd"],
                        f"Qtd Total ({qtd_conjuntos} conj.)": comp["qtd"] * qtd_conjuntos,
                        "Valor Unit": f"R$ {pr_u:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                        "Valor Total": f"R$ {pr_t:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    })
                
                if cav.get("mo_mont_unit", 0.0) > 0:
                    mo_m_u = cav["mo_mont_unit"]
                    lista_itens_cav.append({
                        "Componente": "🔧 SERVIÇO: Mão de Obra de Montagem Hidráulica (Por Conjunto)",
                        "Unid": "sv",
                        "Qtd Unit": 1.0,
                        f"Qtd Total ({qtd_conjuntos} conj.)": 1.0 * qtd_conjuntos,
                        "Valor Unit": f"R$ {mo_m_u:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                        "Valor Total": f"R$ {(mo_m_u * qtd_conjuntos):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    })
                    
                if cav.get("mo_isol_unit", 0.0) > 0:
                    mo_i_u = cav["mo_isol_unit"]
                    lista_itens_cav.append({
                        "Componente": "🔧 SERVIÇO: Mão de Obra de Isolamento Térmico e Rechapeamento (Por Conjunto)",
                        "Unid": "sv",
                        "Qtd Unit": 1.0,
                        f"Qtd Total ({qtd_conjuntos} conj.)": 1.0 * qtd_conjuntos,
                        "Valor Unit": f"R$ {mo_i_u:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                        "Valor Total": f"R$ {(mo_i_u * qtd_conjuntos):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    })
                    
                st.dataframe(pd.DataFrame(lista_itens_cav), use_container_width=True, hide_index=True)
                
            st.markdown("---")
            st.subheader("🛒 Lista de Materiais Consolidada (BOM Única)")
            lista_bom = []
            for nome, dados in materiais_condensados_lista.items():
                pr_u = dict_lookup_valores.get(normalizar_string_busca(nome), 0.0)
                is_pc = dict_lookup_unidades.get(normalizar_string_busca(nome), "pç") == "pç"
                match = re.search(r'Ø\s*([\d\./"]+)', nome)
                lista_bom.append({
                    "Bitola": match.group(1) if match else "Geral", "Item / Componente": nome,
                    "Aplicação (TAGs)": ", ".join(sorted(dados['tags'])) if dados['tags'] else "Uso Geral",
                    "Qtd Total": math.ceil(dados['qtd']) if is_pc else round(dados['qtd'], 2),
                    "Unid": dict_lookup_unidades.get(normalizar_string_busca(nome), "un"),
                    "Preço Ref.": f"R$ {pr_u:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                    "Preço Total": f"R$ {(pr_u * dados['qtd']):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                })
                
            df_bom_final = pd.DataFrame(lista_bom).sort_values(by=["Bitola", "Item / Componente"])
            st.dataframe(df_bom_final, use_container_width=True, hide_index=True)
            
            st.markdown("---")
            buf_excel_hidro = io.BytesIO()
            wb_export_h = openpyxl.Workbook()
            
            ws_f = wb_export_h.active; ws_f.title = "Resumo Comercial"
            ws_f.views.sheetView[0].showGridLines = True
            ws_f.row_dimensions[1].height = 35
            ws_f.merge_cells("A1:G1")
            ws_f.cell(row=1, column=1, value="PLANILHA DE FECHAMENTO DE CUSTOS - CAVALETES HIDRÁULICOS").font = Font(name="Arial", size=12, bold=True, color="1C8590")
            ws_f.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
            
            f_head_ex = PatternFill(start_color="1C8590", end_color="1C8590", fill_type="solid")
            font_head_ex = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            b_thin = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
            
            df_res_quadros_excel = pd.DataFrame(resumo_financeiro_quadros)
            for r_i, r_data in enumerate(dataframe_to_rows(df_res_quadros_excel, index=False, header=True), start=3):
                for c_i, val_cell in enumerate(r_data, start=1):
                    celula_ex = ws_f.cell(row=r_i, column=c_i, value=val_cell)
                    celula_ex.border = b_thin
                    if r_i == 3: 
                        celula_ex.fill = f_head_ex; celula_ex.font = font_head_ex; celula_ex.alignment = Alignment(horizontal="center")
                    elif c_i >= 4: 
                        celula_ex.number_format = '"R$" #,##0.00'; celula_ex.alignment = Alignment(horizontal="right")
                            
            custo_total_geral_hidraulica = total_hidro_material + total_hidro_montagem + total_hidro_isolamento
            linha_tot_f = len(df_res_quadros_excel) + 4
            ws_f.merge_cells(start_row=linha_tot_f, start_column=1, end_row=linha_tot_f, end_column=6)
            ws_f.cell(row=linha_tot_f, column=1, value="TOTAL CONSOLIDADO DO ESCOPO HIDRÁULICO:").font = Font(bold=True)
            ws_f.cell(row=linha_tot_f, column=1).alignment = Alignment(horizontal="right")
            ws_f_tot_val = ws_f.cell(row=linha_tot_f, column=7, value=custo_total_geral_hidraulica)
            ws_f_tot_val.font = Font(bold=True, size=11); ws_f_tot_val.number_format = '"R$" #,##0.00'; ws_f_tot_val.border = b_thin
            
            ws_b = wb_export_h.create_sheet(title="Lista de Materiais (BOM)")
            ws_b.views.sheetView[0].showGridLines = True
            for r_i, r_data in enumerate(dataframe_to_rows(df_bom_final, index=False, header=True), start=1):
                for c_i, val_cell in enumerate(r_data, start=1):
                    celula_ex = ws_b.cell(row=r_i, column=c_i, value=val_cell)
                    celula_ex.border = b_thin
                    if r_i == 1: 
                        celula_ex.fill = f_head_ex; celula_ex.font = font_head_ex; celula_ex.alignment = Alignment(horizontal="center")
            
            for sheet_obj in [ws_f, ws_b]:
                for col_obj in sheet_obj.columns:
                    max_len_val = max(len(str(c_val.value or '')) for c_val in col_obj if c_val.row <= 100)
                    sheet_obj.column_dimensions[get_column_letter(col_obj[0].column)].width = max(max_len_val + 3, 12)
                    
            wb_export_h.save(buf_excel_hidro); buf_excel_hidro.seek(0)
            st.download_button(label="📥 Exportar Relatório Consolidado para Excel", data=buf_excel_hidro.getvalue(), file_name="Orcamento_Cavaletes_Consolidado.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, key="btn_export_excel_hidro_v17")
            
            st.markdown("---")
            st.markdown("### 📝 Descritivo Comercial Automático")
            texto_comercial_final = "FORNECIMENTO E MONTAGEM DE CAVALETES HIDRÁULICOS PARA SISTEMAS DE CLIMATIZAÇÃO\n\n"
            texto_comercial_final += "O escopo desta proposta contempla a montagem mecânica e o fornecimento de materiais para os conjuntos hidráulicos descritos abaixo:\n"
            
            agrupado_comercial = {}
            tem_sistema_fechado = False
            
            for cav in st.session_state.cavaletes_selecionados:
                sys_lbl = " [Sistema Aberto]" if "Aberto" in cav.get("sistema", "") else ""
                if not "Aberto" in cav.get("sistema", ""): tem_sistema_fechado = True
                
                chave_com = f"Cavalete para {cav.get('equipamento', 'EQ')} ({cav.get('vias', '2 Vias')}) - Diâmetro de Ø {cav.get('bitola', '1/2\"')}{sys_lbl}"
                if chave_com not in agrupado_comercial: agrupado_comercial[chave_com] = {"qtd": 0, "tags": []}
                agrupado_comercial[chave_com]["qtd"] += cav.get("quantidade", 1)
                tag_limpa = cav.get('tag', 'S/ TAG')
                if tag_limpa != "S/ TAG": agrupado_comercial[chave_com]["tags"].append(tag_limpa)
                    
            for desc_conj, dados_conj in agrupado_comercial.items():
                str_tags_c = f" [TAGs: {', '.join(dados_conj['tags'])}]" if dados_conj['tags'] else ""
                texto_comercial_final += f"• {dados_conj['qtd']} conjunto(s) - {desc_conj}{str_tags_c}, completo com válvulas, tubulações, conexões estruturais e periféricas.\n"
                
            texto_comercial_final += "\n**Serviços Especializados Inclusos:**\n"
            texto_comercial_final += "• Mão de obra qualificada para montagem mecânica, rosqueamento/soldagem e acoplamento das tubulações;\n"
            
            if tem_sistema_fechado:
                texto_comercial_final += "• Fornecimento e aplicação de isolamento térmico em espuma elastomérica de alta densidade (espessura 25mm) com proteção mecânica e acabamento profissional em chapa de alumínio liso."
            
            st.text_area("Texto formatado para inserção direta na Proposta Comercial:", value=texto_comercial_final, height=250)
